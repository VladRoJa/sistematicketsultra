from __future__ import annotations

import argparse
import hashlib
import json
import os
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select

from app import create_app
from app.extensions import db
from app.models.routine_control import RoutineControlMemberORM


BUSINESS_TIMEZONE = ZoneInfo("America/Tijuana")
BACKFILL_DATE_FROM = date(2026, 1, 1)
BACKFILL_DATE_TO = date(2026, 8, 15)
EXPECTED_MANIFEST_MEMBERS = 20313


@dataclass(frozen=True, slots=True)
class ArtifactSpec:
    month: str
    run_id: str
    size_bytes: int
    sha256: str
    expected_rows: int
    expected_unique_ids: int
    expected_extra_duplicate_rows: int


ARTIFACTS = (
    ArtifactSpec(
        month="2026-01",
        run_id="3fc8fa2ed6884652b404b75743c09575",
        size_bytes=758225,
        sha256="6df892a38c5c07eb04033605514810d976a79a7700e236b048756b3e787761e8",
        expected_rows=4475,
        expected_unique_ids=4303,
        expected_extra_duplicate_rows=172,
    ),
    ArtifactSpec(
        month="2026-02",
        run_id="ad191015244048e383ed81531485a312",
        size_bytes=538459,
        sha256="d80ccdf25e45ffebe0a80a76d64dd1218cbf069017655f5f0342c52313d87757",
        expected_rows=3085,
        expected_unique_ids=3081,
        expected_extra_duplicate_rows=4,
    ),
    ArtifactSpec(
        month="2026-03",
        run_id="099a3c4bb27c41e4a97d9e44f43433e6",
        size_bytes=464582,
        sha256="6e099712e2c3bf78be1a002c9e326e2eeb0dbcf1aaf661dc0ca3324c17d2a818",
        expected_rows=2683,
        expected_unique_ids=2683,
        expected_extra_duplicate_rows=0,
    ),
    ArtifactSpec(
        month="2026-04",
        run_id="0d7a918e22e2461f8a49b56f89e1d931",
        size_bytes=371345,
        sha256="8be6d41e063c4c20cd76cb8622cfbe068d9b3df3eee3921acf37965edbf13243",
        expected_rows=2174,
        expected_unique_ids=2174,
        expected_extra_duplicate_rows=0,
    ),
    ArtifactSpec(
        month="2026-05",
        run_id="f33f37649b5646a7ac5186f8bdddafc4",
        size_bytes=343616,
        sha256="60b335024f9d0e68d00d7708fb9d81040b98340d5cb6b7ad8f4c90dfd76ba76d",
        expected_rows=1990,
        expected_unique_ids=1990,
        expected_extra_duplicate_rows=0,
    ),
    ArtifactSpec(
        month="2026-06",
        run_id="1da99208ec8b462196a7a78cff111404",
        size_bytes=426826,
        sha256="ae1082122ff21d990c0179fabd50ca7292baf22ad145bec5926b342b7d943dc1",
        expected_rows=2497,
        expected_unique_ids=2497,
        expected_extra_duplicate_rows=0,
    ),
    ArtifactSpec(
        month="2026-07",
        run_id="9cbc53981c6a4ab4a7fb6285f69450f0",
        size_bytes=424396,
        sha256="bd6617153d3648628fddd19aa6df5c747c019e6faac2326d3c7580b0de280f11",
        expected_rows=2473,
        expected_unique_ids=2473,
        expected_extra_duplicate_rows=0,
    ),
    ArtifactSpec(
        month="2026-08",
        run_id="d49e3a557b0b437683febec6814e17d4",
        size_bytes=191338,
        sha256="0fabec00cfb4db9aabe920c8065eed2f64852e42625621e4015b3c8821e136fc",
        expected_rows=1112,
        expected_unique_ids=1112,
        expected_extra_duplicate_rows=0,
    ),
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()

    with path.open("rb") as source:
        for chunk in iter(
            lambda: source.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.hexdigest()


def _member_id_text(value: Any) -> str:
    if isinstance(value, bool):
        raise RuntimeError("IDSocio booleano no es válido.")

    if isinstance(value, int) and value > 0:
        return str(value)

    if (
        isinstance(value, str)
        and value.isdigit()
        and int(value) > 0
    ):
        return str(int(value))

    raise RuntimeError(
        f"IDSocio inválido en artifact: {value!r}"
    )


def _folio_text(value: Any) -> str:
    if (
        isinstance(value, str)
        and value
        and value.isdigit()
    ):
        return value

    raise RuntimeError(
        "IDFolio inválido; debe conservarse como texto de dígitos."
    )


def _parse_fecha_pago(value: Any) -> datetime:
    if not isinstance(value, str):
        raise RuntimeError(
            "FechaPago debe conservar el formato "
            "dd-mm-YYYY HH:mm:ss."
        )

    try:
        local_value = datetime.strptime(
            value.strip(),
            "%d-%m-%Y %H:%M:%S",
        )
    except ValueError as exc:
        raise RuntimeError(
            f"FechaPago inválida: {value!r}"
        ) from exc

    return local_value.replace(
        tzinfo=BUSINESS_TIMEZONE
    )


def _artifact_path(
    root: Path,
    spec: ArtifactSpec,
) -> Path:
    return (
        root
        / "gasca"
        / "new_members"
        / spec.run_id
        / "gasca-new-members.xlsx"
    )


def _load_artifact(
    root: Path,
    spec: ArtifactSpec,
) -> dict[str, datetime]:
    path = _artifact_path(root, spec)

    if not path.is_file():
        raise RuntimeError(
            f"{spec.month}: artifact no encontrado: {path}"
        )

    actual_size = path.stat().st_size

    if actual_size != spec.size_bytes:
        raise RuntimeError(
            f"{spec.month}: size_bytes cambió. "
            f"Esperado={spec.size_bytes}, actual={actual_size}."
        )

    actual_sha256 = _sha256(path)

    if actual_sha256 != spec.sha256:
        raise RuntimeError(
            f"{spec.month}: SHA-256 cambió. "
            f"Esperado={spec.sha256}, actual={actual_sha256}."
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        if "Socios" not in workbook.sheetnames:
            raise RuntimeError(
                f"{spec.month}: falta hoja Socios."
            )

        worksheet = workbook["Socios"]

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            (),
        )

        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_row)
            if value is not None
        }

        required = {
            "IDSocio",
            "IDFolio",
            "FechaPago",
        }

        missing = required.difference(headers)

        if missing:
            raise RuntimeError(
                f"{spec.month}: faltan headers: "
                f"{sorted(missing)}."
            )

        by_source_id: dict[str, list[datetime]] = (
            defaultdict(list)
        )
        row_count = 0

        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            row_count += 1

            member_id = _member_id_text(
                row[headers["IDSocio"]]
            )
            folio = _folio_text(
                row[headers["IDFolio"]]
            )
            sale_local = _parse_fecha_pago(
                row[headers["FechaPago"]]
            )

            if sale_local.strftime("%Y-%m") != spec.month:
                raise RuntimeError(
                    f"{spec.month}: FechaPago fuera del mes "
                    f"para {member_id}:{folio}: "
                    f"{sale_local.isoformat()}."
                )

            source_record_id = f"{member_id}:{folio}"

            by_source_id[source_record_id].append(
                sale_local
            )

        if row_count != spec.expected_rows:
            raise RuntimeError(
                f"{spec.month}: filas cambiaron. "
                f"Esperado={spec.expected_rows}, "
                f"actual={row_count}."
            )

        unique_count = len(by_source_id)

        if unique_count != spec.expected_unique_ids:
            raise RuntimeError(
                f"{spec.month}: identidades únicas cambiaron. "
                f"Esperado={spec.expected_unique_ids}, "
                f"actual={unique_count}."
            )

        extra_duplicate_rows = sum(
            len(values) - 1
            for values in by_source_id.values()
        )

        if (
            extra_duplicate_rows
            != spec.expected_extra_duplicate_rows
        ):
            raise RuntimeError(
                f"{spec.month}: duplicados cambiaron. "
                "Esperado="
                f"{spec.expected_extra_duplicate_rows}, "
                f"actual={extra_duplicate_rows}."
            )

        resolved: dict[str, datetime] = {}

        for source_record_id, values in by_source_id.items():
            unique_values = {
                value.isoformat()
                for value in values
            }

            if len(unique_values) != 1:
                raise RuntimeError(
                    f"{spec.month}: FechaPago ambiguo para "
                    f"{source_record_id}: "
                    f"{sorted(unique_values)}."
                )

            resolved[source_record_id] = (
                values[0].astimezone(timezone.utc)
            )

        return resolved

    finally:
        workbook.close()


def _canonical_json_hash(
    payload: dict[str, Any],
) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=lambda value: (
            value.isoformat()
            if isinstance(value, (date, datetime))
            else TypeError(
                "Tipo no serializable: "
                f"{type(value).__name__}"
            )
        ),
    ).encode("utf-8")

    return hashlib.sha256(encoded).hexdigest()


def _payload_hash_with_sale_time(
    member: RoutineControlMemberORM,
    sale_at_utc: datetime,
) -> str:
    operational_payload = {
        "external_member_id": member.external_member_id,
        "external_sale_id": member.external_sale_id,
        "source_branch_name": member.source_branch_name,
        "sucursal_id": member.sucursal_id,
        "member_name": member.member_name,
        "email_original": member.email_original,
        "email_normalized": member.email_normalized,
        "phone_original": member.phone_original,
        "phone_normalized": member.phone_normalized,
        "sale_date": member.sale_date,
        "sale_at_utc": sale_at_utc,
    }

    return _canonical_json_hash(
        operational_payload
    )


def _aware_utc(value: datetime) -> datetime:
    if (
        value.tzinfo is None
        or value.utcoffset() is None
    ):
        raise RuntimeError(
            "sale_at_utc persistido debe incluir timezone."
        )

    return value.astimezone(timezone.utc)


def _month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def run_backfill(
    *,
    artifact_root: Path,
    commit: bool,
) -> None:
    artifacts_by_month = {
        spec.month: _load_artifact(
            artifact_root,
            spec,
        )
        for spec in ARTIFACTS
    }

    members = db.session.scalars(
        select(RoutineControlMemberORM)
        .where(
            RoutineControlMemberORM.source_system
            == "gasca",
            RoutineControlMemberORM.sale_date
            >= BACKFILL_DATE_FROM,
            RoutineControlMemberORM.sale_date
            <= BACKFILL_DATE_TO,
        )
        .order_by(
            RoutineControlMemberORM.cohort_month,
            RoutineControlMemberORM.source_record_id,
        )
    ).all()

    if len(members) < EXPECTED_MANIFEST_MEMBERS:
        raise RuntimeError(
            "Faltan filas Gasca dentro del periodo auditado. "
            f"Periodo={BACKFILL_DATE_FROM}..{BACKFILL_DATE_TO}, "
            f"mínimo_esperado={EXPECTED_MANIFEST_MEMBERS}, "
            f"actual={len(members)}."
        )

    members_by_month: dict[
        str,
        dict[str, RoutineControlMemberORM],
    ] = defaultdict(dict)

    for member in members:
        month = _month_key(member.cohort_month)

        if member.source_record_id in members_by_month[month]:
            raise RuntimeError(
                "DB contiene source_record_id duplicado: "
                f"{member.source_record_id}."
            )

        members_by_month[month][
            member.source_record_id
        ] = member

    expected_months = {
        spec.month
        for spec in ARTIFACTS
    }

    actual_months = set(members_by_month)

    if actual_months != expected_months:
        raise RuntimeError(
            "Los meses Gasca de DB cambiaron. "
            f"Esperados={sorted(expected_months)}, "
            f"actuales={sorted(actual_months)}."
        )

    update_plan: list[
        tuple[
            RoutineControlMemberORM,
            datetime,
            str,
        ]
    ] = []

    missing_time = 0
    already_correct_time = 0
    hash_changes = 0
    allowed_extra_db_rows = 0
    manifest_members: list[
        RoutineControlMemberORM
    ] = []
    seen_manifest_ids: set[str] = set()

    for spec in ARTIFACTS:
        artifact_map = artifacts_by_month[spec.month]
        db_map = members_by_month[spec.month]

        artifact_ids = set(artifact_map)
        db_ids = set(db_map)

        artifact_only = artifact_ids - db_ids

        if artifact_only:
            raise RuntimeError(
                f"{spec.month}: faltan identidades del "
                "manifiesto en DB. "
                f"artifact_only={len(artifact_only)}, "
                f"sample={sorted(artifact_only)[:10]}."
            )

        db_only = db_ids - artifact_ids

        unexpected_db_only = {
            source_record_id
            for source_record_id in db_only
            if db_map[source_record_id].sale_date
            != BACKFILL_DATE_TO
        }

        if unexpected_db_only:
            raise RuntimeError(
                f"{spec.month}: aparecieron identidades "
                "históricas fuera del manifiesto. "
                f"count={len(unexpected_db_only)}, "
                f"sample="
                f"{sorted(unexpected_db_only)[:10]}."
            )

        allowed_extra_db_rows += len(db_only)

        for source_record_id in sorted(artifact_ids):
            if source_record_id in seen_manifest_ids:
                raise RuntimeError(
                    "El manifiesto repite una identidad "
                    "entre meses: "
                    f"{source_record_id}."
                )

            seen_manifest_ids.add(source_record_id)

            member = db_map[source_record_id]
            manifest_members.append(member)
            expected_sale_at_utc = artifact_map[
                source_record_id
            ]

            if (
                member.sale_date
                != expected_sale_at_utc
                .astimezone(BUSINESS_TIMEZONE)
                .date()
            ):
                raise RuntimeError(
                    f"{spec.month}: sale_date conflictivo "
                    f"para {source_record_id}. "
                    f"DB={member.sale_date}, "
                    "artifact="
                    f"{expected_sale_at_utc.isoformat()}."
                )

            if member.sale_at_utc is None:
                missing_time += 1
                time_needs_update = True
            else:
                persisted = _aware_utc(
                    member.sale_at_utc
                )

                if persisted != expected_sale_at_utc:
                    raise RuntimeError(
                        f"{spec.month}: sale_at_utc "
                        f"conflictivo para "
                        f"{source_record_id}. "
                        f"DB={persisted.isoformat()}, "
                        "artifact="
                        f"{expected_sale_at_utc.isoformat()}."
                    )

                already_correct_time += 1
                time_needs_update = False

            expected_payload_hash = (
                _payload_hash_with_sale_time(
                    member,
                    expected_sale_at_utc,
                )
            )

            hash_needs_update = (
                member.payload_hash
                != expected_payload_hash
            )

            if hash_needs_update:
                hash_changes += 1

            if (
                time_needs_update
                or hash_needs_update
            ):
                update_plan.append(
                    (
                        member,
                        expected_sale_at_utc,
                        expected_payload_hash,
                    )
                )

    if (
        len(manifest_members)
        != EXPECTED_MANIFEST_MEMBERS
    ):
        raise RuntimeError(
            "El total de identidades procesadas no "
            "coincide con el manifiesto. "
            f"Esperado={EXPECTED_MANIFEST_MEMBERS}, "
            f"actual={len(manifest_members)}."
        )

    print("Routine Control sale_at_utc backfill")
    print(
        {
            "mode": "COMMIT" if commit else "DRY_RUN",
            "artifact_root": str(artifact_root),
            "backfill_date_from": BACKFILL_DATE_FROM.isoformat(),
            "backfill_date_to": BACKFILL_DATE_TO.isoformat(),
            "db_scoped_gasca_rows": len(members),
            "manifest_members": len(manifest_members),
            "allowed_extra_db_rows": (
                allowed_extra_db_rows
            ),
            "missing_sale_at_utc": missing_time,
            "already_correct_sale_at_utc": (
                already_correct_time
            ),
            "payload_hash_changes": hash_changes,
            "rows_to_update": len(update_plan),
        }
    )

    if not commit:
        db.session.rollback()
        print("DRY_RUN_OK: no se modificó la DB.")
        return

    for (
        member,
        expected_sale_at_utc,
        expected_payload_hash,
    ) in update_plan:
        member.sale_at_utc = expected_sale_at_utc
        member.payload_hash = expected_payload_hash

    db.session.flush()

    remaining_manifest_null = sum(
        1
        for member in manifest_members
        if member.sale_at_utc is None
    )

    if remaining_manifest_null != 0:
        raise RuntimeError(
            "Validación post-update falló para "
            "el manifiesto: "
            "remaining_manifest_null="
            f"{remaining_manifest_null}."
        )

    db.session.commit()

    print(
        "BACKFILL_OK",
        {
            "updated": len(update_plan),
            "remaining_manifest_without_sale_at_utc": 0,
            "allowed_extra_db_rows": (
                allowed_extra_db_rows
            ),
        },
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Backfill idempotente de sale_at_utc para "
            "Routine Control usando artifacts Gasca "
            "previamente auditados."
        )
    )

    parser.add_argument(
        "--artifact-root",
        default=os.getenv(
            "ROUTINE_CONTROL_ARTIFACT_DIR",
            "/app/runtime/routine-control/artifacts",
        ),
        help=(
            "Raíz de artifacts de Routine Control. "
            "Default: ROUTINE_CONTROL_ARTIFACT_DIR."
        ),
    )

    parser.add_argument(
        "--commit",
        action="store_true",
        help=(
            "Aplica cambios. Sin este flag el script "
            "solo audita y muestra el plan."
        ),
    )

    args = parser.parse_args()

    app = create_app()

    with app.app_context():
        try:
            run_backfill(
                artifact_root=Path(
                    args.artifact_root
                ).resolve(),
                commit=args.commit,
            )
        except Exception:
            db.session.rollback()
            raise
        finally:
            db.session.remove()


if __name__ == "__main__":
    main()
