from __future__ import annotations

import hashlib
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.routine_control.providers.gasca_member_normalizer import (
    GascaInvalidRequiredValueError,
    GascaMissingHeaderError,
    GascaUnsafeIdentifierError,
    load_gasca_member_commands_from_xlsx,
    normalize_gasca_member_row,
)


class GascaMemberNormalizerTestCase(unittest.TestCase):
    fixture_path = (
        Path(__file__).parent
        / "fixtures"
        / "gasca_socios_nuevos_detallado.xlsx"
    )
    observed_at = datetime(2026, 7, 15, 18, 0, tzinfo=timezone.utc)

    @classmethod
    def setUpClass(cls) -> None:
        workbook = load_workbook(cls.fixture_path, data_only=True)
        cls.sheet_names = workbook.sheetnames
        worksheet = workbook["Socios"]
        cls.headers = tuple(cell.value for cell in worksheet[1])
        cls.fixture_rows = [
            dict(zip(cls.headers, values))
            for values in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                values_only=True,
            )
        ]
        workbook.close()
        cls.batch = load_gasca_member_commands_from_xlsx(
            cls.fixture_path,
            observed_at_utc=cls.observed_at,
            branch_resolver=lambda _branch: 1,
        )

    def _normalize(self, row=None, **kwargs):
        return normalize_gasca_member_row(
            dict(row or self.fixture_rows[0]),
            observed_at_utc=kwargs.get("observed_at_utc", self.observed_at),
            branch_resolver=kwargs.get("branch_resolver", lambda _branch: 1),
        )

    def test_fixture_opens_socios_sheet(self) -> None:
        self.assertEqual(self.sheet_names, ["Socios"])

    def test_fixture_contains_required_headers(self) -> None:
        required = {
            "IDSocio",
            "IDFolio",
            "Sucursal",
            "Nombre",
            "ApellidoPaterno",
            "ApellidoMaterno",
            "Email",
            "FechaPago",
            "FechaCreacion",
        }
        self.assertTrue(required.issubset(self.headers))

    def test_fixture_generates_33_commands(self) -> None:
        self.assertEqual(self.batch.total_source_rows, 33)
        self.assertEqual(len(self.batch.commands), 33)
        self.assertEqual(self.batch.rejected_rows, ())

    def test_idfolio_is_preserved_as_20_digit_string(self) -> None:
        self.assertTrue(
            all(
                isinstance(command.external_sale_id, str)
                and len(command.external_sale_id) == 20
                and command.external_sale_id.isdigit()
                for command in self.batch.commands
            )
        )

    def test_source_record_id_uses_member_and_folio(self) -> None:
        command = self.batch.commands[0]
        self.assertEqual(
            command.source_record_id,
            f"{command.external_member_id}:{command.external_sale_id}",
        )

    def test_source_identity_key_is_stable(self) -> None:
        first = self._normalize()
        second = self._normalize()
        self.assertEqual(first.source_identity_key, second.source_identity_key)

    def test_payload_hash_ignores_observed_at(self) -> None:
        first = self._normalize()
        second = self._normalize(
            observed_at_utc=self.observed_at + timedelta(days=1)
        )
        self.assertEqual(first.payload_hash, second.payload_hash)

    def test_fecha_pago_converts_to_date(self) -> None:
        command = self._normalize()
        expected = datetime.strptime(
            self.fixture_rows[0]["FechaPago"], "%d-%m-%Y %H:%M:%S"
        ).date()
        self.assertEqual(command.sale_date, expected)

    def test_fecha_creacion_does_not_replace_sale_date(self) -> None:
        command = self._normalize()
        self.assertIsNone(command.source_updated_at_utc)
        self.assertEqual(
            command.source_metadata["source_created_at"],
            self.fixture_rows[0]["FechaCreacion"].strip(),
        )

    def test_member_name_collapses_spaces(self) -> None:
        row = dict(self.fixture_rows[0])
        row.update(
            Nombre="  Ana   María ",
            ApellidoPaterno="  López  ",
            ApellidoMaterno=" ",
        )
        self.assertEqual(self._normalize(row).member_name, "Ana María López")

    def test_valid_email_is_normalized(self) -> None:
        row = dict(self.fixture_rows[0])
        row["Email"] = "  User.Name@Example.COM  "
        command = self._normalize(row)
        self.assertEqual(command.email_original, "User.Name@Example.COM")
        self.assertEqual(command.email_normalized, "user.name@example.com")

    def test_empty_email_normalizes_to_none(self) -> None:
        row = dict(self.fixture_rows[0])
        row["Email"] = "   "
        command = self._normalize(row)
        self.assertIsNone(command.email_original)
        self.assertIsNone(command.email_normalized)

    def test_email_sentinel_77_is_not_normalized(self) -> None:
        row = dict(self.fixture_rows[0])
        row["Email"] = "77"
        command = self._normalize(row)
        self.assertEqual(command.email_original, "77")
        self.assertIsNone(command.email_normalized)

    def test_gamil_domain_is_not_corrected(self) -> None:
        row = dict(self.fixture_rows[0])
        row["Email"] = "USER@GAMIL.COM"
        command = self._normalize(row)
        self.assertEqual(command.email_normalized, "user@gamil.com")

    def test_duplicate_emails_are_allowed(self) -> None:
        first_row = dict(self.fixture_rows[0])
        second_row = dict(self.fixture_rows[1])
        first_row["Email"] = second_row["Email"] = "same@example.com"
        first = self._normalize(first_row)
        second = self._normalize(second_row)
        self.assertEqual(first.email_normalized, second.email_normalized)
        self.assertNotEqual(first.source_record_id, second.source_record_id)

    def test_one_unresolved_branch_rejects_only_one_row(self) -> None:
        calls = 0

        def resolver(_branch):
            nonlocal calls
            calls += 1
            return None if calls == 1 else 1

        batch = load_gasca_member_commands_from_xlsx(
            self.fixture_path,
            observed_at_utc=self.observed_at,
            branch_resolver=resolver,
        )
        self.assertEqual(len(batch.commands), 32)
        self.assertEqual(len(batch.rejected_rows), 1)
        self.assertEqual(batch.rejected_rows[0].reason_code, "BRANCH_UNRESOLVED")

    def test_branch_resolver_receives_clean_original_name(self) -> None:
        received = []
        row = dict(self.fixture_rows[0])
        row["Sucursal"] = "  Sucursal Fuente  "
        self._normalize(row, branch_resolver=lambda value: received.append(value) or 1)
        self.assertEqual(received, ["Sucursal Fuente"])

    def test_numeric_idfolio_is_rejected_as_unsafe(self) -> None:
        for unsafe in (12345678901234567890, 1.2345678901234567e19):
            row = dict(self.fixture_rows[0])
            row["IDFolio"] = unsafe
            with self.subTest(unsafe=type(unsafe).__name__):
                with self.assertRaises(GascaUnsafeIdentifierError):
                    self._normalize(row)

    def test_missing_required_header_fails_batch(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "missing_header.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Socios"
            worksheet.append([header for header in self.headers if header != "Email"])
            workbook.save(path)
            workbook.close()
            with self.assertRaises(GascaMissingHeaderError):
                load_gasca_member_commands_from_xlsx(
                    path,
                    observed_at_utc=self.observed_at,
                    branch_resolver=lambda _branch: 1,
                )

    def test_naive_observed_at_is_rejected(self) -> None:
        with self.assertRaises(GascaInvalidRequiredValueError):
            self._normalize(observed_at_utc=datetime(2026, 7, 15, 18, 0))

    def test_observed_at_is_normalized_to_utc(self) -> None:
        source = datetime(
            2026,
            7,
            15,
            11,
            0,
            tzinfo=timezone(timedelta(hours=-7)),
        )
        self.assertEqual(self._normalize(observed_at_utc=source).observed_at_utc, self.observed_at)

    def test_source_metadata_excludes_sensitive_fields(self) -> None:
        metadata = self._normalize().source_metadata
        forbidden = {"Telefono", "Domicilio", "FechaNacimiento", "phone", "address"}
        self.assertTrue(forbidden.isdisjoint(metadata))
        self.assertFalse(any("nacimiento" in key.lower() for key in metadata))

    def test_loader_does_not_write_fixture(self) -> None:
        before = hashlib.sha256(self.fixture_path.read_bytes()).hexdigest()
        load_gasca_member_commands_from_xlsx(
            self.fixture_path,
            observed_at_utc=self.observed_at,
            branch_resolver=lambda _branch: 1,
        )
        after = hashlib.sha256(self.fixture_path.read_bytes()).hexdigest()
        self.assertEqual(before, after)


if __name__ == "__main__":
    unittest.main()

