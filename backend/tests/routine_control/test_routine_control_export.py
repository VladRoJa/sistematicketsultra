from __future__ import annotations

import unittest
from datetime import date

from openpyxl import load_workbook

from app.routine_control.queries.export_service import EXPORT_HEADERS, build_members_export


class RoutineControlExportTest(unittest.TestCase):
    def setUp(self):
        self.item = {
            "external_member_id": "M-1", "external_sale_id": "F-1", "member_name": "Socio",
            "email": "socio@example.com", "branch_name": "Centro", "sale_date": "2026-07-01",
            "classification_status": "CLASSIFIED", "current_status": "CON_RUTINA",
            "first_routine_at": "2026-07-01", "latest_routine_at": "2026-07-02",
            "current_instructor_name": "Ana", "routine_assignment_type": "MISMO_DIA",
            "active_incident_count": 0, "active_evidence_count": 1, "status_version": 2,
            "source_metadata": {"secret": True}, "payload_hash": "do-not-export",
        }
        self.workbook = load_workbook(build_members_export([self.item]))
        self.sheet = self.workbook.active

    def test_headers_and_values_match_contract(self):
        self.assertEqual(tuple(cell.value for cell in self.sheet[1]), EXPORT_HEADERS)
        self.assertEqual(self.sheet["A2"].value, "M-1")
        self.assertEqual(self.sheet["O2"].value, 2)

    def test_dates_are_native_excel_dates(self):
        self.assertIsInstance(self.sheet["F2"].value, (date,))
        self.assertEqual(self.sheet["F2"].number_format, "yyyy-mm-dd")

    def test_filter_freeze_and_visible_headers(self):
        self.assertEqual(self.sheet.freeze_panes, "A2")
        self.assertEqual(self.sheet.auto_filter.ref, "A1:O2")
        self.assertTrue(self.sheet["A1"].font.bold)

    def test_sensitive_metadata_is_not_exported(self):
        values = [cell.value for row in self.sheet.iter_rows() for cell in row]
        serialized = " ".join(str(value) for value in values)
        self.assertNotIn("source_metadata", serialized)
        self.assertNotIn("payload_hash", serialized)
        self.assertNotIn("do-not-export", serialized)

    def test_workbook_contains_no_formulas(self):
        self.assertFalse(any(cell.data_type == "f" for row in self.sheet.iter_rows() for cell in row))


if __name__ == "__main__":
    unittest.main()
