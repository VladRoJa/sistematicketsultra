from __future__ import annotations

import hashlib
import tempfile
import unittest
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, to_excel

from app.routine_control.providers.trainingym_evidence_normalizer import (
    _adapt_trainingym_csv_row,
    _coerce_trainingym_csv_operational_types,
    _reject_csv_non_routine_without_employee,
    _read_trainingym_csv_rows,
    TrainingymInvalidRequiredValueError,
    TrainingymMissingHeaderError,
    TrainingymNormalizationError,
    load_trainingym_evidence_commands_from_csv,
    load_trainingym_evidence_commands_from_xlsx,
    normalize_trainingym_evidence_row,
)


class TrainingymEvidenceNormalizerTestCase(unittest.TestCase):
    fixture_path = Path(__file__).parent / "fixtures" / "trainingym_workout.xlsx"
    observed_at = datetime(2026, 7, 15, 18, 0, tzinfo=timezone.utc)

    @classmethod
    def setUpClass(cls) -> None:
        workbook = load_workbook(cls.fixture_path, data_only=True)
        cls.sheet_names = workbook.sheetnames
        worksheet = workbook["Export"]
        cls.headers = tuple(cell.value for cell in worksheet[1])
        cls.epoch = workbook.epoch
        cls.fixture_rows = [
            dict(zip(cls.headers, values))
            for values in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                values_only=True,
            )
        ]
        workbook.close()
        cls.valid_row = next(
            row
            for row in cls.fixture_rows
            if isinstance(row["id"], int)
            and isinstance(row["NºRutinas"], int)
            and row["NºRutinas"] > 0
            and "automat" not in str(row["Técnico"]).lower()
        )
        cls.batch = load_trainingym_evidence_commands_from_xlsx(
            cls.fixture_path,
            observed_at_utc=cls.observed_at,
            provider_run_id=41,
        )

    def _normalize(self, row=None, **kwargs):
        return normalize_trainingym_evidence_row(
            dict(row or self.valid_row),
            observed_at_utc=kwargs.get("observed_at_utc", self.observed_at),
            provider_run_id=kwargs.get("provider_run_id", 41),
            center_resolver=kwargs.get("center_resolver"),
            excel_epoch=kwargs.get("excel_epoch", self.epoch),
        )

    def test_new_trainingym_csv_reads_bom_semicolon_and_required_headers(self) -> None:
        content = (
            "\ufeff"
            '"ID";"ID externo";"Socio";"Email";"Edad";"Sexo";'
            '"Empleados";"Workouts";"Pesajes";"Total";"Valoración";'
            '"Fecha";"Centro"\n'
            '"24639860";"88669";"SOCIO PRUEBA";"test@example.com";'
            '"42";"F";"TECNICO PRUEBA";"1";"0";"1";"";'
            '"2026-08-15";"UltraGym & Fitness - Centro"\n'
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "rutinas_pesajes.csv"
            path.write_text(
                content,
                encoding="utf-8",
                newline="",
            )

            rows = _read_trainingym_csv_rows(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ID"], "24639860")
        self.assertEqual(rows[0]["ID externo"], "88669")
        self.assertEqual(rows[0]["Empleados"], "TECNICO PRUEBA")
        self.assertEqual(rows[0]["Workouts"], "1")
        self.assertEqual(rows[0]["Pesajes"], "0")
        self.assertEqual(rows[0]["Fecha"], "2026-08-15")
        self.assertEqual(
            rows[0]["Centro"],
            "UltraGym & Fitness - Centro",
        )

    def test_new_trainingym_csv_missing_required_header_fails(self) -> None:
        content = (
            "\ufeff"
            '"ID";"ID externo";"Socio";"Email";"Empleados";'
            '"Workouts";"Pesajes";"Centro"\n'
            '"24639860";"88669";"SOCIO PRUEBA";"test@example.com";'
            '"TECNICO PRUEBA";"1";"0";'
            '"UltraGym & Fitness - Centro"\n'
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "missing_fecha.csv"
            path.write_text(
                content,
                encoding="utf-8",
                newline="",
            )

            with self.assertRaises(
                TrainingymMissingHeaderError
            ) as raised:
                _read_trainingym_csv_rows(path)

        self.assertIn(
            "Fecha",
            str(raised.exception),
        )

    def test_new_trainingym_csv_row_maps_to_canonical_contract(self) -> None:
        source = {
            "ID": "24639860",
            "ID externo": "88669",
            "Socio": "SOCIO PRUEBA",
            "Email": "test@example.com",
            "Edad": "42",
            "Sexo": "F",
            "Empleados": "TECNICO PRUEBA",
            "Workouts": "1",
            "Pesajes": "0",
            "Total": "1",
            "Valoración": "5",
            "Fecha": "2026-08-15",
            "Centro": "UltraGym & Fitness - Centro",
        }

        adapted = _adapt_trainingym_csv_row(source)

        self.assertEqual(adapted["id"], "24639860")
        self.assertEqual(adapted["Idsocioexterno"], "88669")
        self.assertEqual(adapted["NombreApellidos"], "SOCIO PRUEBA")
        self.assertEqual(adapted["Email"], "test@example.com")
        self.assertEqual(adapted["Técnico"], "TECNICO PRUEBA")
        self.assertEqual(adapted["NºRutinas"], "1")
        self.assertEqual(adapted["NºPesajes"], "0")
        self.assertEqual(adapted["Fecha"], "2026-08-15")
        self.assertEqual(
            adapted["Centro Origen"],
            "UltraGym & Fitness - Centro",
        )
        self.assertEqual(adapted["Valoración"], "5")
        self.assertEqual(adapted["Sexo"], "F")
        self.assertEqual(adapted["Total Rutinas-Pesaje"], "1")

        self.assertNotIn("Edad", adapted)
        self.assertNotIn("age", adapted)

    def test_new_trainingym_csv_operational_types_are_coerced(self) -> None:
        adapted = {
            "id": "24639860",
            "Idsocioexterno": "88669",
            "NombreApellidos": "SOCIO PRUEBA",
            "Email": "test@example.com",
            "Técnico": "TECNICO PRUEBA",
            "NºRutinas": "1",
            "NºPesajes": "0",
            "Fecha": "2026-08-15",
            "Centro Origen": "UltraGym & Fitness - Centro",
            "Valoración": "5",
            "Sexo": "F",
            "Total Rutinas-Pesaje": "1",
        }

        coerced = _coerce_trainingym_csv_operational_types(
            adapted
        )

        self.assertEqual(coerced["NºRutinas"], 1)
        self.assertIsInstance(coerced["NºRutinas"], int)

        self.assertEqual(coerced["NºPesajes"], 0)
        self.assertIsInstance(coerced["NºPesajes"], int)

        self.assertEqual(
            coerced["Fecha"],
            date(2026, 8, 15),
        )
        self.assertIsInstance(coerced["Fecha"], date)

        self.assertEqual(
            coerced["Técnico"],
            "TECNICO PRUEBA",
        )
        self.assertEqual(
            coerced["Valoración"],
            "5",
        )

    def test_new_trainingym_csv_row_generates_valid_routine_evidence(self) -> None:
        source = {
            "ID": "24639860",
            "ID externo": "88669",
            "Socio": "SOCIO PRUEBA",
            "Email": "Test@Example.COM",
            "Edad": "42",
            "Sexo": "F",
            "Empleados": "  José   Núñez  ",
            "Workouts": "1",
            "Pesajes": "0",
            "Total": "1",
            "Valoración": "5",
            "Fecha": "2026-08-15",
            "Centro": "  UltraGym   Centro.  ",
        }

        adapted = _adapt_trainingym_csv_row(source)
        coerced = _coerce_trainingym_csv_operational_types(
            adapted
        )

        command = normalize_trainingym_evidence_row(
            coerced,
            observed_at_utc=self.observed_at,
            provider_run_id=41,
        )

        self.assertEqual(command.provider_key, "trainingym")
        self.assertEqual(
            command.provider_member_id,
            "24639860",
        )
        self.assertEqual(
            command.external_member_id,
            "88669",
        )
        self.assertEqual(
            command.member_name_original,
            "SOCIO PRUEBA",
        )
        self.assertEqual(
            command.email_normalized,
            "test@example.com",
        )
        self.assertEqual(
            command.instructor_name,
            "José Núñez",
        )
        self.assertEqual(
            command.instructor_name_normalized,
            "jose nunez",
        )
        self.assertEqual(
            command.routine_activity_date,
            date(2026, 8, 15),
        )
        self.assertEqual(command.routine_count, 1)
        self.assertEqual(command.weighing_count, 0)
        self.assertEqual(
            command.provider_center_name,
            "UltraGym Centro.",
        )
        self.assertEqual(
            command.provider_center_key,
            "ultragym centro",
        )
        self.assertEqual(command.provider_run_id, 41)

    def test_new_trainingym_csv_loader_builds_batch_and_rejects_weighing_only(self) -> None:
        content = (
            "\ufeff"
            '"ID";"ID externo";"Socio";"Email";"Edad";"Sexo";'
            '"Empleados";"Workouts";"Pesajes";"Total";"Valoración";'
            '"Fecha";"Centro"\n'
            '"24639860";"88669";"SOCIO RUTINA";"uno@example.com";'
            '"42";"F";"TECNICO PRUEBA";"1";"0";"1";"";'
            '"2026-08-15";"UltraGym Centro"\n'
            '"24639861";"88670";"SOCIO PESAJE";"dos@example.com";'
            '"39";"M";"";"0";"1";"1";"";'
            '"2026-08-15";"UltraGym Centro"\n'
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "rutinas_pesajes.csv"

            path.write_text(
                content,
                encoding="utf-8",
                newline="",
            )

            batch = load_trainingym_evidence_commands_from_csv(
                path,
                observed_at_utc=self.observed_at,
                provider_run_id=41,
            )

        self.assertEqual(
            batch.total_source_rows,
            2,
        )
        self.assertEqual(
            len(batch.commands),
            1,
        )
        self.assertEqual(
            batch.commands[0].provider_member_id,
            "24639860",
        )
        self.assertEqual(
            batch.commands[0].routine_count,
            1,
        )

        self.assertEqual(
            len(batch.rejected_rows),
            1,
        )
        self.assertEqual(
            batch.rejected_rows[0].row_number,
            3,
        )
        self.assertEqual(
            batch.rejected_rows[0].reason_code,
            "WEIGHING_ONLY",
        )

    def test_fixture_opens_export_sheet(self) -> None:
        self.assertEqual(self.sheet_names, ["Export"])

    def test_fixture_contains_required_headers(self) -> None:
        required = {
            "id",
            "Idsocioexterno",
            "NombreApellidos",
            "Email",
            "Técnico",
            "NºRutinas",
            "NºPesajes",
            "Fecha",
            "Centro Origen",
        }
        self.assertTrue(required.issubset(self.headers))

    def test_fixture_excludes_non_operational_rows(self) -> None:
        reasons = Counter(row.reason_code for row in self.batch.rejected_rows)
        self.assertEqual(reasons["SUMMARY_ROW"], 1)
        self.assertEqual(reasons["EMPTY_ROW"], 1)
        self.assertEqual(reasons["FILTER_DESCRIPTION_ROW"], 1)

    def test_fixture_excludes_automatic_routines(self) -> None:
        reasons = Counter(row.reason_code for row in self.batch.rejected_rows)
        self.assertEqual(reasons["AUTOMATIC_ROUTINE"], 9)

    def test_new_csv_weighing_only_with_blank_employee_is_classified_correctly(self) -> None:
        source = {
            "ID": "24639860",
            "ID externo": "88669",
            "Socio": "SOCIO PRUEBA",
            "Email": "test@example.com",
            "Edad": "42",
            "Sexo": "F",
            "Empleados": "",
            "Workouts": "0",
            "Pesajes": "1",
            "Total": "1",
            "Valoración": "",
            "Fecha": "2026-08-15",
            "Centro": "UltraGym Centro",
        }

        adapted = _adapt_trainingym_csv_row(source)
        coerced = _coerce_trainingym_csv_operational_types(
            adapted
        )

        with self.assertRaises(
            TrainingymNormalizationError
        ) as caught:
            _reject_csv_non_routine_without_employee(
                coerced
            )

        self.assertEqual(
            caught.exception.reason_code,
            "WEIGHING_ONLY",
        )

    def test_weighing_without_routine_is_excluded(self) -> None:
        row = dict(self.valid_row)
        row["NºRutinas"] = 0
        row["NºPesajes"] = 1
        with self.assertRaises(TrainingymNormalizationError) as caught:
            self._normalize(row)
        self.assertEqual(caught.exception.reason_code, "WEIGHING_ONLY")

    def test_no_routine_is_excluded(self) -> None:
        row = dict(self.valid_row)
        row["NºRutinas"] = 0
        row["NºPesajes"] = None
        with self.assertRaises(TrainingymNormalizationError) as caught:
            self._normalize(row)
        self.assertEqual(caught.exception.reason_code, "NO_ROUTINE")

    def test_fixture_generates_human_routine_commands(self) -> None:
        self.assertEqual(self.batch.total_source_rows, 31)
        self.assertEqual(len(self.batch.commands), 19)
        self.assertTrue(all(command.routine_count > 0 for command in self.batch.commands))

    def test_provider_member_id_is_stable_string(self) -> None:
        command = self._normalize()
        self.assertIsInstance(command.provider_member_id, str)
        self.assertEqual(command.provider_member_id, str(self.valid_row["id"]))

    def test_external_member_placeholder_17_becomes_none(self) -> None:
        row = dict(self.valid_row)
        row["Idsocioexterno"] = "17"
        self.assertIsNone(self._normalize(row).external_member_id)

    def test_empty_external_member_id_becomes_none(self) -> None:
        row = dict(self.valid_row)
        row["Idsocioexterno"] = "  "
        self.assertIsNone(self._normalize(row).external_member_id)

    def test_valid_evidence_without_external_member_id_is_accepted(self) -> None:
        row = dict(self.valid_row)
        row["Idsocioexterno"] = None
        command = self._normalize(row)
        self.assertIsNone(command.external_member_id)
        self.assertEqual(command.provider_key, "trainingym")

    def test_excel_datetime_converts_to_date(self) -> None:
        command = self._normalize()
        self.assertIsInstance(self.valid_row["Fecha"], datetime)
        self.assertEqual(command.routine_activity_date, self.valid_row["Fecha"].date())

    def test_excel_serial_uses_supplied_epoch(self) -> None:
        row = dict(self.valid_row)
        expected = date(2026, 2, 3)
        row["Fecha"] = to_excel(expected, epoch=WINDOWS_EPOCH)
        command = self._normalize(row, excel_epoch=WINDOWS_EPOCH)
        self.assertEqual(command.routine_activity_date, expected)

    def test_center_normalization_removes_accents_period_and_extra_spaces(self) -> None:
        row = dict(self.valid_row)
        row["Centro Origen"] = "  Sucursál   Centro.  "
        command = self._normalize(row)
        self.assertEqual(command.provider_center_name, "Sucursál Centro.")
        self.assertEqual(command.provider_center_key, "sucursal centro")

    def test_instructor_normalization_removes_accents_and_extra_spaces(self) -> None:
        row = dict(self.valid_row)
        row["Técnico"] = "  José   Núñez  "
        command = self._normalize(row)
        self.assertEqual(command.instructor_name, "José Núñez")
        self.assertEqual(command.instructor_name_normalized, "jose nunez")

    def test_member_name_is_persisted_and_normalized_deterministically(self) -> None:
        row = dict(self.valid_row)
        row["NombreApellidos"] = "  MARÍA-José   O'Connor.  "
        command = self._normalize(row)
        self.assertEqual(command.member_name_original, "MARÍA-José O'Connor.")
        self.assertEqual(command.member_name_normalized, "maria jose o connor")

    def test_empty_member_name_is_accepted_for_email_matching(self) -> None:
        row = dict(self.valid_row)
        row["NombreApellidos"] = "   "
        command = self._normalize(row)
        self.assertIsNone(command.member_name_original)
        self.assertIsNone(command.member_name_normalized)

    def test_evidence_identity_key_is_stable(self) -> None:
        self.assertEqual(
            self._normalize().evidence_identity_key,
            self._normalize().evidence_identity_key,
        )

    def test_routine_count_does_not_change_identity(self) -> None:
        row = dict(self.valid_row)
        row["NºRutinas"] += 1
        self.assertEqual(
            self._normalize().evidence_identity_key,
            self._normalize(row).evidence_identity_key,
        )

    def test_routine_count_changes_payload_hash(self) -> None:
        row = dict(self.valid_row)
        row["NºRutinas"] += 1
        self.assertNotEqual(
            self._normalize().payload_hash,
            self._normalize(row).payload_hash,
        )

    def test_member_name_changes_payload_hash_but_not_identity(self) -> None:
        row = dict(self.valid_row)
        row["NombreApellidos"] = "Nombre Corregido"
        original = self._normalize()
        changed = self._normalize(row)
        self.assertEqual(
            original.evidence_identity_key,
            changed.evidence_identity_key,
        )
        self.assertNotEqual(original.payload_hash, changed.payload_hash)

    def test_observed_at_does_not_change_payload_hash(self) -> None:
        later = self.observed_at + timedelta(days=1)
        self.assertEqual(
            self._normalize().payload_hash,
            self._normalize(observed_at_utc=later).payload_hash,
        )

    def test_email_is_normalized_but_not_part_of_identity(self) -> None:
        first_row = dict(self.valid_row)
        second_row = dict(self.valid_row)
        first_row["Email"] = " User@Example.COM "
        second_row["Email"] = "another@example.com"
        first = self._normalize(first_row)
        second = self._normalize(second_row)
        self.assertEqual(first.email_normalized, "user@example.com")
        self.assertEqual(first.evidence_identity_key, second.evidence_identity_key)

    def test_source_metadata_excludes_sensitive_fields(self) -> None:
        metadata = self._normalize().source_metadata or {}
        forbidden = {"Movil", "NombreApellidos", "Edad", "mobile", "name", "age"}
        self.assertTrue(forbidden.isdisjoint(metadata))

    def test_provider_run_id_is_propagated(self) -> None:
        self.assertEqual(self._normalize(provider_run_id=987).provider_run_id, 987)

    def test_center_resolver_receives_clean_original_name(self) -> None:
        row = dict(self.valid_row)
        row["Centro Origen"] = "  Centro   Norte.  "
        received = []
        command = self._normalize(
            row,
            center_resolver=lambda value: received.append(value) or 5,
        )
        self.assertEqual(received, ["Centro Norte."])
        self.assertEqual(command.sucursal_id, 5)

    def test_unresolved_optional_center_leaves_sucursal_none(self) -> None:
        command = self._normalize(center_resolver=lambda _center: None)
        self.assertIsNone(command.sucursal_id)

    def test_missing_required_header_fails_batch(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "missing_header.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Export"
            worksheet.append([header for header in self.headers if header != "Fecha"])
            workbook.save(path)
            workbook.close()
            with self.assertRaises(TrainingymMissingHeaderError):
                load_trainingym_evidence_commands_from_xlsx(
                    path,
                    observed_at_utc=self.observed_at,
                    provider_run_id=41,
                )

    def test_missing_member_name_header_fails_batch(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "missing_member_name_header.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Export"
            worksheet.append([
                header
                for header in self.headers
                if header != "NombreApellidos"
            ])
            workbook.save(path)
            workbook.close()
            with self.assertRaises(TrainingymMissingHeaderError):
                load_trainingym_evidence_commands_from_xlsx(
                    path,
                    observed_at_utc=self.observed_at,
                    provider_run_id=41,
                )

    def test_naive_observed_at_is_rejected(self) -> None:
        with self.assertRaises(TrainingymInvalidRequiredValueError):
            self._normalize(observed_at_utc=datetime(2026, 7, 15, 18, 0))

    def test_invalid_provider_run_id_type_is_rejected(self) -> None:
        with self.assertRaises(TrainingymInvalidRequiredValueError):
            self._normalize(provider_run_id="41")

    def test_loader_does_not_write_fixture(self) -> None:
        before = hashlib.sha256(self.fixture_path.read_bytes()).hexdigest()
        load_trainingym_evidence_commands_from_xlsx(
            self.fixture_path,
            observed_at_utc=self.observed_at,
            provider_run_id=41,
        )
        after = hashlib.sha256(self.fixture_path.read_bytes()).hexdigest()
        self.assertEqual(before, after)


if __name__ == "__main__":
    unittest.main()

