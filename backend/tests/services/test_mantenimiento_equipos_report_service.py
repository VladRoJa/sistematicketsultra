from datetime import datetime, timezone
from datetime import timedelta
from types import SimpleNamespace
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.services import mantenimiento_equipos_report_service as report


class _Expression:
    def __init__(self, predicate):
        self.predicate = predicate

    def __or__(self, other):
        return _Expression(
            lambda row: self.predicate(row) or other.predicate(row)
        )

    def __and__(self, other):
        return _Expression(
            lambda row: self.predicate(row) and other.predicate(row)
        )


class _Column:
    def __init__(self, name):
        self.name = name

    def __eq__(self, value):
        return _Expression(lambda row: getattr(row, self.name) == value)

    def isnot(self, value):
        return _Expression(lambda row: getattr(row, self.name) is not value)

    def is_(self, value):
        return _Expression(lambda row: getattr(row, self.name) is value)

    def in_(self, values):
        return _Expression(lambda row: getattr(row, self.name) in values)

    def asc(self):
        return self


class _Query:
    def __init__(self, rows):
        self.rows = list(rows)

    def options(self, *_args):
        return self

    def filter(self, *expressions):
        if any(expression is False for expression in expressions):
            self.rows = []
            return self

        self.rows = [
            row
            for row in self.rows
            if all(expression.predicate(row) for expression in expressions)
        ]
        return self

    def order_by(self, *_args):
        return self

    def all(self):
        return self.rows


def _ticket(
    ticket_id,
    *,
    family=None,
    family_id=None,
    failure=None,
    condition="TRABAJA",
    state="abierto",
    apparatus_id=90,
    created_at=None,
    history=None,
    branch_id=10,
    branch_name="TIJUANA CENTRO",
):
    current_family = SimpleNamespace(key="PESO_LIBRE", nombre="Peso Libre")
    inventory = SimpleNamespace(
        id=apparatus_id,
        nombre="Caminadora comercial",
        codigo_interno="11CCJW21",
        familia_equipo=current_family,
        familia_equipo_id=6,
    )
    branch = SimpleNamespace(sucursal=branch_name)
    return SimpleNamespace(
        id=ticket_id,
        descripcion="Ruido durante operación",
        estado=state,
        departamento_id=1,
        aparato_id=apparatus_id,
        fecha_creacion=created_at
        or datetime(2025, 5, 1, 18, 0, tzinfo=timezone.utc),
        fecha_solucion=None,
        sucursal_id=branch_id,
        sucursal_id_destino=branch_id,
        sucursal_destino=branch,
        sucursal=None,
        inventario=inventory,
        familia_equipo_id=family_id,
        familia_equipo=family,
        falla_mantenimiento=failure,
        condicion_operativa=condition,
        historial_fechas=history or [],
    )


def _fake_ticket_model(rows):
    return SimpleNamespace(
        query=_Query(rows),
        inventario="inventario",
        familia_equipo="familia_equipo",
        falla_mantenimiento="falla_mantenimiento",
        sucursal="sucursal",
        sucursal_destino="sucursal_destino",
        departamento_id=_Column("departamento_id"),
        aparato_id=_Column("aparato_id"),
        familia_equipo_id=_Column("familia_equipo_id"),
        estado=_Column("estado"),
        fecha_creacion=_Column("fecha_creacion"),        sucursal_id=_Column("sucursal_id"),
        sucursal_id_destino=_Column("sucursal_id_destino"),
        id=_Column("id"),
    )


class MantenimientoEquiposReportServiceTest(unittest.TestCase):
    def test_region_resuelve_solo_ids_actuales_desde_la_fuente_oficial(self):
        region_query = MagicMock()
        region_query.filter.return_value.first.return_value = SimpleNamespace(
            id=1
        )
        assignment_query = MagicMock()
        assignment_rows = (
            assignment_query.with_entities.return_value
            .filter.return_value
            .all
        )
        assignment_rows.return_value = [
            (11,),
            (10,),
            (11,),
        ]
        region_model = SimpleNamespace(
            query=region_query,
            id=_Column("id"),
            is_active=_Column("is_active"),
        )
        assignment_model = SimpleNamespace(
            query=assignment_query,
            sucursal_id=_Column("sucursal_id"),
            region_id=_Column("region_id"),
            is_current=_Column("is_current"),
        )

        with (
            patch.object(report, "SuiteRegionORM", region_model),
            patch.object(
                report,
                "SuiteSucursalRegionAssignmentORM",
                assignment_model,
            ),
        ):
            branch_ids = report._obtener_sucursales_ids_region(1)

        self.assertEqual(branch_ids, [10, 11])

    def test_region_inexistente_se_rechaza_desde_la_fuente_oficial(self):
        region_query = MagicMock()
        region_query.filter.return_value.first.return_value = None
        region_model = SimpleNamespace(
            query=region_query,
            id=_Column("id"),
            is_active=_Column("is_active"),
        )

        with patch.object(report, "SuiteRegionORM", region_model):
            with self.assertRaisesRegex(
                report.RegionReporteNoEncontradaError,
                "no existe o está inactiva",
            ):
                report._obtener_sucursales_ids_region(999999)

    def test_consulta_incluye_activo_con_aparato_o_snapshot_historico(self):
        active_2025 = _ticket(1)

        snapshot = SimpleNamespace(
            key="PESO_INTEGRADO",
            nombre="Peso Integrado",
        )
        historical_without_apparatus = _ticket(
            2,
            apparatus_id=None,
            family=snapshot,
            family_id=6,
        )

        without_apparatus_or_snapshot = _ticket(
            3,
            apparatus_id=None,
            family=None,
            family_id=None,
        )

        finalized = _ticket(4, state="finalizado")

        other_department = _ticket(5)
        other_department.departamento_id = 7

        fake_ticket_model = _fake_ticket_model(
            [
                active_2025,
                historical_without_apparatus,
                without_apparatus_or_snapshot,
                finalized,
                other_department,
            ]
        )

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
        ):
            selected = report.obtener_tickets_reporte()

        self.assertEqual(
            [ticket.id for ticket in selected],
            [1, 2],
        )

    def test_todo_conserva_todos_los_tickets_permitidos_en_el_xlsx(self):
        tickets = [
            _ticket(6, branch_id=10, branch_name="REGIÓN A - SUCURSAL 1"),
            _ticket(7, branch_id=20, branch_name="REGIÓN B - SUCURSAL 1"),
        ]
        user = SimpleNamespace(id=20)
        fake_ticket_model = _fake_ticket_model(tickets)

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
            patch.object(
                report,
                "filtrar_tickets_por_usuario",
                return_value=fake_ticket_model.query,
            ) as filter_by_user,
        ):
            selected = report.obtener_tickets_reporte(user=user)

        workbook = load_workbook(report.construir_reporte_xlsx(selected))
        branch_names = {
            row[6]
            for row in list(
                workbook["Tickets"].iter_rows(values_only=True)
            )[1:]
        }

        self.assertEqual(
            branch_names,
            {"REGIÓN A - SUCURSAL 1", "REGIÓN B - SUCURSAL 1"},
        )
        filter_by_user.assert_called_once_with(user)

    def test_region_incluye_todas_sus_sucursales_en_el_xlsx(self):
        tickets = [
            _ticket(8, branch_id=10, branch_name="REGIÓN A - SUCURSAL 1"),
            _ticket(9, branch_id=11, branch_name="REGIÓN A - SUCURSAL 2"),
        ]
        user = SimpleNamespace(id=20)
        fake_ticket_model = _fake_ticket_model(tickets)

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
            patch.object(
                report,
                "filtrar_tickets_por_usuario",
                return_value=fake_ticket_model.query,
            ),
            patch.object(
                report,
                "_obtener_sucursales_ids_region",
                return_value=[10, 11],
            ),
        ):
            selected = report.obtener_tickets_reporte(user=user, region_id=1)

        workbook = load_workbook(report.construir_reporte_xlsx(selected))
        branch_names = {
            row[6]
            for row in list(
                workbook["Tickets"].iter_rows(values_only=True)
            )[1:]
        }

        self.assertEqual(
            branch_names,
            {"REGIÓN A - SUCURSAL 1", "REGIÓN A - SUCURSAL 2"},
        )

    def test_region_excluye_tickets_de_otra_region_en_el_xlsx(self):
        tickets = [
            _ticket(40, branch_id=10, branch_name="REGIÓN A - SUCURSAL 1"),
            _ticket(41, branch_id=20, branch_name="REGIÓN B - SUCURSAL 1"),
        ]
        user = SimpleNamespace(id=20)
        fake_ticket_model = _fake_ticket_model(tickets)

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
            patch.object(
                report,
                "filtrar_tickets_por_usuario",
                return_value=fake_ticket_model.query,
            ),
            patch.object(
                report,
                "_obtener_sucursales_ids_region",
                return_value=[10],
            ),
        ):
            selected = report.obtener_tickets_reporte(user=user, region_id=1)

        workbook = load_workbook(report.construir_reporte_xlsx(selected))
        branch_names = [
            row[6]
            for row in list(
                workbook["Tickets"].iter_rows(values_only=True)
            )[1:]
        ]

        self.assertEqual(branch_names, ["REGIÓN A - SUCURSAL 1"])

    def test_region_respeta_el_scope_previamente_autorizado_del_usuario(self):
        allowed = _ticket(
            50,
            branch_id=10,
            branch_name="REGIÓN A - SUCURSAL PERMITIDA",
        )
        denied = _ticket(
            51,
            branch_id=11,
            branch_name="REGIÓN A - SUCURSAL NO PERMITIDA",
        )
        user = SimpleNamespace(id=20)
        fake_ticket_model = _fake_ticket_model([allowed, denied])

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
            patch.object(
                report,
                "filtrar_tickets_por_usuario",
                return_value=_Query([allowed]),
            ) as filter_by_user,
            patch.object(
                report,
                "_obtener_sucursales_ids_region",
                return_value=[10, 11],
            ),
        ):
            selected = report.obtener_tickets_reporte(user=user, region_id=1)

        workbook = load_workbook(report.construir_reporte_xlsx(selected))
        branch_names = [
            row[6]
            for row in list(
                workbook["Tickets"].iter_rows(values_only=True)
            )[1:]
        ]

        self.assertEqual(branch_names, ["REGIÓN A - SUCURSAL PERMITIDA"])
        filter_by_user.assert_called_once_with(user)

    def test_region_valida_sin_sucursales_genera_xlsx_vacio_con_headers(self):
        user = SimpleNamespace(id=20)
        fake_ticket_model = _fake_ticket_model([])

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
            patch.object(
                report,
                "filtrar_tickets_por_usuario",
                return_value=fake_ticket_model.query,
            ),
            patch.object(
                report,
                "_obtener_sucursales_ids_region",
                return_value=[],
            ),
        ):
            selected = report.obtener_tickets_reporte(user=user, region_id=1)

        workbook = load_workbook(report.construir_reporte_xlsx(selected))

        self.assertEqual(selected, [])
        self.assertEqual(workbook["Tickets"].max_row, 1)
        self.assertEqual(
            tuple(
                cell.value
                for cell in workbook["Tickets"][1]
            ),
            report.TICKETS_HEADERS,
        )

    def test_reporte_usa_snapshot_y_no_familia_actual_del_aparato(self):
        snapshot = SimpleNamespace(key="CAMINADORA", nombre="Caminadora")
        failure = SimpleNamespace(nombre="Banda desgastada")
        ticket = _ticket(
            10,
            family=snapshot,
            family_id=1,
            failure=failure,
        )

        workbook = load_workbook(report.construir_reporte_xlsx([ticket]))

        tickets_row = list(workbook["Tickets"].iter_rows(values_only=True))[1]
        self.assertEqual(tickets_row[7], "Caminadora")
        self.assertEqual(workbook["Caminadoras"].max_row, 2)
        self.assertEqual(workbook["Peso Libre"].max_row, 1)

    def test_ticket_legacy_se_muestra_sin_clasificar_historico(self):
        ticket = _ticket(
            11,
            family=None,
            family_id=None,
            failure=None,
            condition=None,
        )

        workbook = load_workbook(report.construir_reporte_xlsx([ticket]))
        row = list(workbook["Tickets"].iter_rows(values_only=True))[1]

        self.assertEqual(row[7], report.LEGACY_UNCLASSIFIED_FAMILY)
        self.assertEqual(row[8], report.PENDING_CAPTURE)
        self.assertEqual(row[9], report.PENDING_CAPTURE)
        self.assertEqual(row[10], report.NO_COMMITMENT)

    def test_plan_excluye_rechazo_y_ordena_compromisos(self):
        snapshot = SimpleNamespace(key="CAMINADORA", nombre="Caminadora")
        history = [
            {
                "fechaCambio": "2026-01-03T10:00:00Z",
                "motivo": "Segundo compromiso",
            },
            {
                "fechaCambio": "2026-01-02T10:00:00Z",
                "motivo": "No aceptado por gerente",
                "tipo": "rechazo_cierre_gerente",
            },
            {
                "fechaCambio": "2026-01-01T10:00:00Z",
                "motivo": "Primer compromiso",
            },
        ]
        ticket = _ticket(12, family=snapshot, family_id=1, history=history)

        workbook = load_workbook(report.construir_reporte_xlsx([ticket]))
        row = list(workbook["Tickets"].iter_rows(values_only=True))[1]

        self.assertEqual(
            row[11],
            "Primer compromiso\nSegundo compromiso",
        )

    def test_mismo_aparato_con_dos_tickets_cuenta_dos_fallas(self):
        snapshot = SimpleNamespace(key="CAMINADORA", nombre="Caminadora")
        tickets = [
            _ticket(20, family=snapshot, family_id=1),
            _ticket(21, family=snapshot, family_id=1),
        ]

        workbook = load_workbook(report.construir_reporte_xlsx(tickets))
        summary_row = list(
            workbook["Fallas por familia"].iter_rows(values_only=True)
        )[1]

        self.assertEqual(summary_row[1], 2)
        self.assertEqual(summary_row[2], 2)
        self.assertEqual(workbook["Caminadoras"].max_row, 3)

    def test_resumen_muestra_sin_clasificar_y_total_cuadra(self):
        snapshot = SimpleNamespace(
            key="CAMINADORA",
            nombre="Caminadora",
        )

        classified = _ticket(
            30,
            family=snapshot,
            family_id=1,
        )
        unclassified = _ticket(
            31,
            family=None,
            family_id=None,
            failure=None,
            condition=None,
        )

        workbook = load_workbook(
            report.construir_reporte_xlsx(
                [classified, unclassified]
            )
        )

        rows = list(
            workbook["Fallas por familia"].iter_rows(values_only=True)
        )
        headers = rows[0]
        summary = rows[1]

        self.assertEqual(headers[-1], "Sin clasificar")
        self.assertEqual(summary[1], 2)
        self.assertEqual(summary[2], 1)
        self.assertEqual(summary[-1], 1)
        self.assertEqual(
            summary[1],
            sum(value or 0 for value in summary[2:]),
        )



    def test_resumen_agrega_total_de_todas_las_sucursales(self):
        caminadora = SimpleNamespace(
            key="CAMINADORA",
            nombre="Caminadora",
        )
        peso_libre = SimpleNamespace(
            key="PESO_LIBRE",
            nombre="Peso Libre",
        )

        tickets = [
            _ticket(
                60,
                branch_id=10,
                branch_name="SUCURSAL A",
                family=caminadora,
                family_id=1,
            ),
            _ticket(
                61,
                branch_id=10,
                branch_name="SUCURSAL A",
                family=None,
                family_id=None,
                failure=None,
                condition=None,
            ),
            _ticket(
                62,
                branch_id=20,
                branch_name="SUCURSAL B",
                family=peso_libre,
                family_id=6,
            ),
        ]

        workbook = load_workbook(
            report.construir_reporte_xlsx(tickets)
        )

        rows = list(
            workbook["Fallas por familia"].iter_rows(values_only=True)
        )

        total = rows[-1]

        self.assertEqual(total[0], "TOTAL")
        self.assertEqual(total[1], 3)
        self.assertEqual(total[2], 1)
        self.assertEqual(total[7], 1)
        self.assertEqual(total[9], 1)
        self.assertEqual(
            total[1],
            sum(value or 0 for value in total[2:]),
        )

    def test_historico_incluye_activos_y_finalizados(self):
        activo = _ticket(
            70,
            state="abierto",
        )
        finalizado = _ticket(
            71,
            state="finalizado",
        )

        otro_departamento = _ticket(
            72,
            state="finalizado",
        )
        otro_departamento.departamento_id = 7

        sin_equipo_ni_snapshot = _ticket(
            73,
            state="finalizado",
            apparatus_id=None,
            family=None,
            family_id=None,
        )

        fake_ticket_model = _fake_ticket_model(
            [
                activo,
                finalizado,
                otro_departamento,
                sin_equipo_ni_snapshot,
            ]
        )

        with (
            patch.object(report, "Ticket", fake_ticket_model),
            patch.object(report, "joinedload", side_effect=lambda value: value),
        ):
            selected = report.obtener_tickets_historico_reporte()

        self.assertEqual(
            [ticket.id for ticket in selected],
            [70, 71],
        )

    def test_historico_mensual_separa_creacion_finalizacion_y_timezone(self):
        caminadora = SimpleNamespace(
            key="CAMINADORA",
            nombre="Caminadora",
        )
        peso_libre = SimpleNamespace(
            key="PESO_LIBRE",
            nombre="Peso Libre",
        )

        julio_por_timezone = _ticket(
            80,
            family=caminadora,
            family_id=1,
            created_at=datetime(
                2026, 8, 1, 6, 30, tzinfo=timezone.utc
            ),
        )

        agosto_finalizado_en_septiembre = _ticket(
            81,
            family=peso_libre,
            family_id=6,
            state="finalizado",
            created_at=datetime(
                2026, 8, 10, 18, 0, tzinfo=timezone.utc
            ),
        )
        agosto_finalizado_en_septiembre.fecha_finalizado = datetime(
            2026, 9, 5, 18, 0, tzinfo=timezone.utc
        )

        octubre_sin_clasificar = _ticket(
            82,
            family=None,
            family_id=None,
            created_at=datetime(
                2026, 10, 10, 18, 0, tzinfo=timezone.utc
            ),
        )

        rows = report._monthly_history_rows(
            [
                julio_por_timezone,
                agosto_finalizado_en_septiembre,
                octubre_sin_clasificar,
            ],
            now=datetime(
                2026, 10, 20, 18, 0, tzinfo=timezone.utc
            ),
        )

        self.assertEqual(
            [row[0].strftime("%Y-%m") for row in rows],
            [
                "2026-07",
                "2026-08",
                "2026-09",
                "2026-10",
            ],
        )

        julio, agosto, septiembre, octubre = rows

        self.assertEqual(julio[1], 1)
        self.assertEqual(julio[2], 1)
        self.assertEqual(julio[10], 0)

        self.assertEqual(agosto[1], 1)
        self.assertEqual(agosto[7], 1)
        self.assertEqual(agosto[10], 0)

        self.assertEqual(septiembre[1], 0)
        self.assertEqual(
            sum(septiembre[2:10]),
            0,
        )
        self.assertEqual(septiembre[10], 1)

        self.assertEqual(octubre[1], 1)
        self.assertEqual(octubre[9], 1)
        self.assertEqual(octubre[10], 0)

        for row in rows:
            self.assertEqual(
                row[1],
                sum(value or 0 for value in row[2:10]),
            )

    def test_xlsx_agrega_hoja_historico_mensual_separada_del_backlog(self):
        backlog_ticket = _ticket(
            90,
            branch_name="BACKLOG ACTUAL",
        )
        historical_ticket = _ticket(
            91,
            state="finalizado",
            branch_name="HISTORICO",
        )

        monthly_row = (
            datetime(2026, 8, 1),
            3,
            1,
            0,
            0,
            0,
            0,
            1,
            0,
            1,
            2,
        )

        with patch.object(
            report,
            "_monthly_history_rows",
            return_value=[monthly_row],
        ) as monthly_rows:
            workbook = load_workbook(
                report.construir_reporte_xlsx(
                    tickets=[backlog_ticket],
                    historical_tickets=[historical_ticket],
                )
            )

        self.assertIn("Histórico mensual", workbook.sheetnames)

        self.assertEqual(
            workbook.sheetnames[:4],
            [
                "Tickets",
                "Fallas por familia",
                "Por validar",
                "Histórico mensual",
            ],
        )

        rows = list(
            workbook["Histórico mensual"].iter_rows(values_only=True)
        )

        self.assertEqual(
            rows[0],
            (
                "Mes",
                "Fallas",
                "Caminadoras",
                "Elípticas",
                "Escaladoras",
                "Recumbentes",
                "Spinning",
                "Peso Libre",
                "Peso Integrado",
                "Sin clasificar",
                "Finalizados",
            ),
        )

        self.assertEqual(rows[1], monthly_row)

        monthly_rows.assert_called_once()
        self.assertEqual(
            monthly_rows.call_args.args[0],
            [historical_ticket],
        )

        self.assertIn(
            "mmmm yyyy",
            workbook["Histórico mensual"]["A2"].number_format,
        )

        tickets_rows = list(
            workbook["Tickets"].iter_rows(values_only=True)
        )
        self.assertEqual(tickets_rows[1][0], 90)

    def test_por_validar_filtra_ordena_y_calcula_antiguedad(self):
        now = datetime.now(timezone.utc)

        antiguo = _ticket(
            100,
            state="por_validar",
            branch_name="SUCURSAL A",
            created_at=now,
        )
        antiguo.fecha_finalizado = now - timedelta(
            days=5,
            hours=2,
        )

        reciente = _ticket(
            101,
            state="por_validar",
            branch_name="SUCURSAL B",
            created_at=now,
        )
        reciente.fecha_finalizado = now - timedelta(
            days=2,
            hours=2,
        )

        en_progreso = _ticket(
            102,
            state="en progreso",
            branch_name="SUCURSAL C",
            created_at=now,
        )
        en_progreso.fecha_finalizado = now - timedelta(days=10)

        solo_historico = _ticket(
            103,
            state="por_validar",
            branch_name="SUCURSAL HISTORICA",
            created_at=now,
        )
        solo_historico.fecha_finalizado = now - timedelta(days=20)

        self.assertEqual(
            report._validation_wait_days(antiguo, now=now),
            5,
        )
        self.assertEqual(
            report._validation_wait_days(reciente, now=now),
            2,
        )

        workbook = load_workbook(
            report.construir_reporte_xlsx(
                tickets=[
                    reciente,
                    en_progreso,
                    antiguo,
                ],
                historical_tickets=[
                    solo_historico,
                ],
            )
        )

        self.assertIn("Por validar", workbook.sheetnames)

        rows = list(
            workbook["Por validar"].iter_rows(values_only=True)
        )

        self.assertEqual(
            rows[0],
            report.POR_VALIDAR_HEADERS,
        )

        data_rows = rows[1:]

        self.assertEqual(
            [row[1] for row in data_rows],
            [100, 101],
        )

        self.assertEqual(
            [row[9] for row in data_rows],
            [5, 2],
        )

        self.assertNotIn(
            102,
            [row[1] for row in data_rows],
        )
        self.assertNotIn(
            103,
            [row[1] for row in data_rows],
        )

        self.assertTrue(data_rows[0][8])
        self.assertTrue(data_rows[1][8])

if __name__ == "__main__":
    unittest.main()