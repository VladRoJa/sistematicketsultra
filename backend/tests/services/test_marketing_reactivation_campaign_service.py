from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
import json
from pathlib import Path
import re
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import UniqueConstraint, select
from sqlalchemy.dialects import postgresql

from app.models import (
    MarketingReactivationCampaignORM,
    MarketingReactivationCampaignRecipientORM,
)
from app.models.warehouse import SociosVencidosCarteraORM
from app.services import marketing_reactivation_service as service


NOW = datetime(2026, 9, 3, 18, 0, tzinfo=timezone.utc)


def _candidate(
    row_id: int,
    *,
    phone: str | None,
    status: str = "CONTACT_HISTORY_UNKNOWN",
    reason: str = "NO_OUTBOUND_EVIDENCE",
    tariff: str = "1 MES $899",
    tariff_category: str | None = "Mensualidad",
    tariff_group: str | None = "REACTIVATE",
    tariff_classified: bool = True,
):
    return {
        "vencido_row_id": row_id,
        "pin": f"PIN-{row_id}",
        "nombre": f"Socio {row_id}",
        "sucursal": "CENTRO",
        "telefono": phone,
        "correo": None,
        "fecha_vencimiento": "2026-08-23",
        "fecha_ultimo_pago": None,
        "tarifa": tariff,
        "tarifa_categoria": tariff_category,
        "tarifa_group": tariff_group,
        "tarifa_classified": tariff_classified,
        "adeudo": None,
        "status": status,
        "reason": reason,
        "active_status": "NOT_FOUND",
        "active_id_socio": None,
        "iventas_contact_id": None,
        "latest_outbound_at_utc": None,
    }


def _candidate_response(rows):
    return {
        "sources": {
            "date_from": "2026-08-23",
            "date_to": "2026-08-23",
            "activos_snapshot_id": 8,
            "iventas_sync_run_id": 26,
            "iventas_period_key": "IVENTAS-2026-08",
        },
        "summary": {"total_rows": len(rows)},
        "rows": rows,
    }


def _filters(**overrides):
    values = {
        "iventas_period_key": "IVENTAS-2026-08",
        "sucursal": None,
        "operational_status": "ALL",
        "search": None,
        "tarifa": None,
        "tariff_group": None,
    }
    values.update(overrides)
    return values


class WriteSession:
    def __init__(self, *, flush_error: Exception | None = None):
        self.added = []
        self.commits = 0
        self.rollbacks = 0
        self.flush_error = flush_error
        self.next_campaign_id = 1

    def add(self, row):
        if isinstance(row, MarketingReactivationCampaignORM) and row.id is None:
            row.id = self.next_campaign_id
            self.next_campaign_id += 1
        self.added.append(row)

    def flush(self):
        if self.flush_error is not None:
            raise self.flush_error

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1


def test_preview_uses_one_engine_and_classifies_every_decision(monkeypatch):
    rows = [
        _candidate(1, phone="686 100 0001"),
        _candidate(
            2,
            phone="686 100 0002",
            status="EXCLUDED_ACTIVE",
            reason="ACTIVE_CONFIRMED",
        ),
        _candidate(3, phone="123"),
        _candidate(
            4,
            phone="686 100 0004",
            status="REVIEW_ACTIVE_MATCH",
            reason="ACTIVE_REVIEW",
        ),
        _candidate(5, phone="686 100 0005"),
        _candidate(6, phone="6861000005"),
        _candidate(7, phone="686 100 0007"),
    ]
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **kwargs: _candidate_response(rows),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **kwargs: {"6861000007": NOW},
    )
    session = SimpleNamespace()

    result = service.preview_marketing_reactivation_campaign(
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        campaign_cooldown_days=30,
        session=session,
        now=NOW,
    )

    assert result["summary"] == {
        "total_candidates": 7,
        "eligible": 1,
        "excluded_active": 1,
        "excluded_invalid_phone": 1,
        "review_identity": 1,
        "duplicate_phone": 2,
        "excluded_tariff": 0,
        "domiciliated_flow": 0,
        "review_tariff": 0,
        "excluded_recent_campaign": 1,
        "review": 3,
    }
    assert not hasattr(session, "added")


def test_preview_resolves_complete_server_side_segment(monkeypatch):
    calls = {}

    def fake_segment(**kwargs):
        calls.update(kwargs)
        return _candidate_response([_candidate(1, phone="6861000001")])

    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        fake_segment,
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **kwargs: {},
    )

    service.preview_marketing_reactivation_campaign(
        date_from="2026-08-01",
        date_to="2026-08-31",
        filters=_filters(
            sucursal="CENTRO",
            tarifa="1 MES $899",
            tariff_group="REACTIVATE",
            operational_status="WORK_PENDING",
            search="socio",
        ),
        session=SimpleNamespace(),
        now=NOW,
    )

    assert calls["allowed_sucursal_keys"] is None
    assert calls["query"].sucursal == "CENTRO"
    assert calls["query"].tarifa == "1 MES $899"
    assert calls["query"].tariff_group == "REACTIVATE"
    assert calls["query"].operational_status == "WORK_PENDING"
    assert calls["query"].search == "socio"


def test_campaign_plan_rejects_requested_branch_outside_backend_scope():
    with pytest.raises(
        service.MarketingReactivationValidationError,
        match="fuera del alcance",
    ):
        service.preview_marketing_reactivation_campaign(
            date_from="2026-08-01",
            date_to="2026-08-31",
            filters=_filters(sucursal="NORTE"),
            allowed_sucursal_keys=("CENTRO",),
            session=SimpleNamespace(),
            now=NOW,
        )


def test_preview_propagates_allowed_scope_to_complete_segment(monkeypatch):
    calls = {}
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **kwargs: calls.update(kwargs) or _candidate_response([]),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **_kwargs: {},
    )

    service.preview_marketing_reactivation_campaign(
        date_from="2026-08-01",
        date_to="2026-08-31",
        filters=_filters(),
        allowed_sucursal_keys=("CENTRO",),
        session=SimpleNamespace(),
        now=NOW,
    )

    assert calls["allowed_sucursal_keys"] == ("CENTRO",)


def test_campaign_freezes_filtered_branch_as_effective_scope(monkeypatch):
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **_kwargs: _candidate_response([]),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **_kwargs: {},
    )

    result = service._build_campaign_plan(
        date_from="2026-08-01",
        date_to="2026-08-31",
        filters=_filters(sucursal="CENTRO"),
        campaign_cooldown_days=None,
        allowed_sucursal_keys=("CENTRO", "NORTE"),
        session=SimpleNamespace(),
        now=NOW,
    )

    assert result["scope"] == {
        "is_global": False,
        "allowed_sucursal_keys": ["CENTRO"],
    }


@pytest.mark.parametrize(
    (
        "tariff",
        "category",
        "tariff_group",
        "expected_eligibility",
        "expected_reason",
    ),
    [
        ("PASE GYMPASS", "Agregadora", "EXCLUDE", "EXCLUDED_TARIFF", "TARIFF_EXCLUDED"),
        ("PASE TOTAL PASS", "Agregadora", "EXCLUDE", "EXCLUDED_TARIFF", "TARIFF_EXCLUDED"),
        ("1 DIA", "Diario", "EXCLUDE", "EXCLUDED_TARIFF", "TARIFF_EXCLUDED"),
        ("1 MES $899", "Mensualidad", "REACTIVATE", "ELIGIBLE", None),
        (
            "DOMICILIADO 12 MESES $499",
            "Domiciliado",
            "DOMICILIATED_FLOW",
            "EXCLUDED_TARIFF",
            "TARIFF_DOMICILIATED_FLOW",
        ),
        ("SEMANA", "Trimestre", "REVIEW", "REVIEW", "REVIEW_TARIFF"),
        ("SEMANA $250", "Semana", "REVIEW", "REVIEW", "REVIEW_TARIFF"),
        ("SEMANA $299", "Semana", "REVIEW", "REVIEW", "REVIEW_TARIFF"),
        ("TARIFA INEXISTENTE", None, None, "REVIEW", "REVIEW_TARIFF"),
    ],
)
def test_tariff_classification_controls_campaign_eligibility(
    tariff,
    category,
    tariff_group,
    expected_eligibility,
    expected_reason,
):
    row = _candidate(
        1,
        phone="6861000001",
        tariff=tariff,
        tariff_category=category,
        tariff_group=tariff_group,
        tariff_classified=tariff_group is not None,
    )

    eligibility, reason = service._resolve_campaign_eligibility(
        row=row,
        phone_mx10="6861000001",
        duplicate_phone=False,
        last_sent_at=None,
        campaign_cooldown_days=None,
        now=NOW,
    )

    assert (eligibility, reason) == (expected_eligibility, expected_reason)


def test_tariff_decision_precedes_phone_checks_but_follows_identity():
    unknown = _candidate(
        1,
        phone=None,
        tariff="DESCONOCIDA",
        tariff_category=None,
        tariff_group=None,
        tariff_classified=False,
    )
    assert service._resolve_campaign_eligibility(
        row=unknown,
        phone_mx10=None,
        duplicate_phone=False,
        last_sent_at=None,
        campaign_cooldown_days=None,
        now=NOW,
    ) == ("REVIEW", "REVIEW_TARIFF")

    identity_review = {
        **unknown,
        "status": "REVIEW_ACTIVE_MATCH",
        "reason": "ACTIVE_REVIEW",
    }
    assert service._resolve_campaign_eligibility(
        row=identity_review,
        phone_mx10=None,
        duplicate_phone=False,
        last_sent_at=None,
        campaign_cooldown_days=None,
        now=NOW,
    ) == ("REVIEW", "REVIEW_IDENTITY")


def test_preview_reports_separate_tariff_audit_counts(monkeypatch):
    rows = [
        _candidate(
            1,
            phone="6861000001",
            tariff="1 DIA",
            tariff_category="Diario",
            tariff_group="EXCLUDE",
        ),
        _candidate(
            2,
            phone="6861000002",
            tariff="DOMICILIADO 12 MESES $499",
            tariff_category="Domiciliado",
            tariff_group="DOMICILIATED_FLOW",
        ),
        _candidate(
            3,
            phone="6861000003",
            tariff="SEMANA $250",
            tariff_category="Semana",
            tariff_group="REVIEW",
        ),
    ]
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **kwargs: _candidate_response(rows),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **kwargs: {},
    )

    result = service.preview_marketing_reactivation_campaign(
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        session=SimpleNamespace(),
        now=NOW,
    )

    assert result["summary"]["excluded_tariff"] == 2
    assert result["summary"]["domiciliated_flow"] == 1
    assert result["summary"]["review_tariff"] == 1
    assert result["summary"]["review"] == 1


def test_tariff_normalization_is_exact_and_does_not_fuzzy_match():
    assert service.normalize_reactivation_tariff_key(
        "  pase\t gympass  "
    ) == "PASE GYMPASS"
    assert service.normalize_reactivation_tariff_key(
        "１２ meses en línea $4990"
    ) == "12 MESES EN LÍNEA $4990"
    assert service.normalize_reactivation_tariff_key(
        "PASE GYMPASS EXTRA"
    ) != "PASE GYMPASS"


def test_candidate_serialization_uses_normalized_tariff_catalog():
    catalog_row = SimpleNamespace(
        tarifa_key="PASE GYMPASS",
        categoria_tarifa="Agregadora",
        reactivation_group="EXCLUDE",
    )
    vencido_row = SimpleNamespace(
        pin="PIN-1",
        nombre="Socio 1",
        sucursal_raw="CENTRO",
        telefono_raw="6861000001",
        correo_raw=None,
        fecha_vencimiento_date=date(2026, 8, 23),
        fecha_ultimo_pago_local=None,
        tarifa="  pase   gympass ",
        adeudo=None,
    )
    candidate = SimpleNamespace(
        vencido_row_id=1,
        status="CONTACT_HISTORY_UNKNOWN",
        reason="NO_OUTBOUND_EVIDENCE",
        active_status="NOT_FOUND",
        active_id_socio=None,
        iventas_contact_id=None,
        latest_outbound_at_utc=None,
    )

    result = service._serialize_candidate(
        candidate=candidate,
        vencido_row=vencido_row,
        tariff_catalog={"PASE GYMPASS": catalog_row},
    )

    assert result["tarifa"] == "  pase   gympass "
    assert result["tarifa_categoria"] == "Agregadora"
    assert result["tarifa_group"] == "EXCLUDE"
    assert result["tarifa_classified"] is True


def test_cooldown_is_not_applied_without_explicit_configuration(monkeypatch):
    rows = [_candidate(1, phone="6861000001")]
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **kwargs: _candidate_response(rows),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **kwargs: {"6861000001": NOW},
    )

    result = service.preview_marketing_reactivation_campaign(
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        session=SimpleNamespace(),
        now=NOW,
    )

    assert result["summary"]["eligible"] == 1
    assert result["summary"]["excluded_recent_campaign"] == 0


def test_no_match_current_run_is_operational_evidence_not_never_contacted(monkeypatch):
    rows = [
        _candidate(
            1,
            phone="6861000001",
            reason="NO_MATCH_CURRENT_IVENTAS_RUN",
        )
    ]
    monkeypatch.setattr(
        service,
        "_build_marketing_reactivation_campaign_segment",
        lambda **kwargs: _candidate_response(rows),
    )
    monkeypatch.setattr(
        service,
        "_read_last_suite_campaign_sent_at",
        lambda **kwargs: {},
    )

    plan = service._build_campaign_plan(
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        campaign_cooldown_days=None,
        allowed_sucursal_keys=None,
        session=SimpleNamespace(),
        now=NOW,
    )

    assert plan["decision_rows"][0]["operational_status"] == (
        "NO_CONTACT_IN_PERIOD"
    )
    assert plan["decision_rows"][0]["campaign_eligibility"] == "ELIGIBLE"


def test_create_campaign_is_atomic_and_snapshots_only_eligible_rows(monkeypatch):
    plan = {
        "sources": _candidate_response([])["sources"],
        "filters": _filters(),
        "summary": {"total_candidates": 1, "eligible": 1},
        "eligible_rows": [
            {
                **_candidate(10, phone="6861000010"),
                "phone_mx10": "6861000010",
                "operational_status": "NO_OUTBOUND_MESSAGE",
            }
        ],
    }
    monkeypatch.setattr(service, "_build_campaign_plan", lambda **kwargs: plan)
    session = WriteSession()

    result = service.create_marketing_reactivation_campaign(
        name="Septiembre Centro",
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        notes="Exportación externa",
        created_by_user_id=3,
        session=session,
        now=NOW,
    )

    assert result["status"] == "DRAFT"
    assert result["recipient_count"] == 1
    assert session.commits == 1
    assert session.rollbacks == 0
    recipients = [
        row
        for row in session.added
        if isinstance(row, MarketingReactivationCampaignRecipientORM)
    ]
    assert len(recipients) == 1
    assert recipients[0].phone_mx10 == "6861000010"
    assert recipients[0].inclusion_status == "ELIGIBLE"
    campaign = next(
        row for row in session.added if isinstance(row, MarketingReactivationCampaignORM)
    )
    assert campaign.filters_json["scope"] == {
        "is_global": True,
        "allowed_sucursal_keys": None,
    }


def test_create_rolls_back_complete_campaign_when_flush_fails(monkeypatch):
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: {
            "sources": _candidate_response([])["sources"],
            "filters": _filters(),
            "summary": {"eligible": 1},
            "eligible_rows": [
                {
                    **_candidate(1, phone="6861000001"),
                    "phone_mx10": "6861000001",
                    "operational_status": "NO_OUTBOUND_MESSAGE",
                }
            ],
        },
    )
    session = WriteSession(flush_error=RuntimeError("flush failed"))

    with pytest.raises(RuntimeError, match="flush failed"):
        service.create_marketing_reactivation_campaign(
            name="Lote",
            date_from="2026-08-23",
            date_to="2026-08-23",
            filters=_filters(),
            created_by_user_id=3,
            session=session,
            now=NOW,
        )

    assert session.commits == 0
    assert session.rollbacks == 1


def test_empty_campaign_is_rejected_without_writes(monkeypatch):
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: {
            "sources": _candidate_response([])["sources"],
            "filters": _filters(),
            "summary": {"eligible": 0},
            "eligible_rows": [],
        },
    )
    session = WriteSession()

    with pytest.raises(
        service.MarketingReactivationValidationError,
        match="no contiene destinatarios",
    ):
        service.create_marketing_reactivation_campaign(
            name="Vacía",
            date_from="2026-08-23",
            date_to="2026-08-23",
            filters=_filters(),
            created_by_user_id=3,
            session=session,
            now=NOW,
        )

    assert session.added == []
    assert session.commits == 0


def test_same_phone_is_allowed_in_different_campaigns(monkeypatch):
    plan = {
        "sources": _candidate_response([])["sources"],
        "filters": _filters(),
        "summary": {"eligible": 1},
        "eligible_rows": [
            {
                **_candidate(1, phone="6861000001"),
                "phone_mx10": "6861000001",
                "operational_status": "NO_OUTBOUND_MESSAGE",
            }
        ],
    }
    monkeypatch.setattr(service, "_build_campaign_plan", lambda **kwargs: plan)
    session = WriteSession()

    first = service.create_marketing_reactivation_campaign(
        name="Primera",
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        created_by_user_id=3,
        session=session,
        now=NOW,
    )
    second = service.create_marketing_reactivation_campaign(
        name="Segunda",
        date_from="2026-08-23",
        date_to="2026-08-23",
        filters=_filters(),
        created_by_user_id=3,
        session=session,
        now=NOW,
    )

    assert first["id"] != second["id"]
    recipients = [
        row
        for row in session.added
        if isinstance(row, MarketingReactivationCampaignRecipientORM)
    ]
    assert {row.campaign_id for row in recipients} == {first["id"], second["id"]}
    assert {row.phone_mx10 for row in recipients} == {"6861000001"}


def test_recipient_unique_constraint_is_scoped_to_campaign_and_phone():
    constraints = {
        constraint.name: tuple(column.name for column in constraint.columns)
        for constraint in MarketingReactivationCampaignRecipientORM.__table__.constraints
        if isinstance(constraint, UniqueConstraint)
    }

    assert constraints[
        "uq_marketing_reactivation_recipients_campaign_phone"
    ] == ("campaign_id", "phone_mx10")


def _campaign(status="DRAFT"):
    recipient = SimpleNamespace(
        id=1,
        socios_vencidos_cartera_id=10,
        phone_mx10="6861000010",
        member_name="Ana Pérez",
        sucursal="CENTRO",
        fecha_vencimiento_date=date(2026, 8, 23),
        tarifa="ANUAL",
        inclusion_status="ELIGIBLE",
        exclusion_reason=None,
        operational_status="NO_OUTBOUND_MESSAGE",
        operational_reason="NO_OUTBOUND_EVIDENCE",
        created_at=NOW,
    )
    return SimpleNamespace(
        id=9,
        name="Campaña 9",
        status=status,
        date_from=date(2026, 8, 23),
        date_to=date(2026, 8, 23),
        created_by_user_id=3,
        created_by_user=SimpleNamespace(username="admin"),
        created_at=NOW,
        updated_at=NOW,
        exported_at=None,
        sent_at=None,
        notes=None,
        filters_json={"filters": _filters()},
        recipient_count=1,
        recipients=[recipient],
    )


def test_export_builds_xlsx_and_transitions_draft_after_success(monkeypatch):
    campaign = _campaign()
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **kwargs: campaign)
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: {
            "eligible_rows": [
                {"vencido_row_id": 10, "phone_mx10": "6861000010"}
            ]
        },
    )

    file_bytes, filename = service.export_marketing_reactivation_campaign(
        campaign_id=9,
        session=session,
        now=NOW,
    )

    workbook = load_workbook(BytesIO(file_bytes), read_only=True)
    rows = list(workbook["Destinatarios"].iter_rows(values_only=True))
    assert rows[0] == (
        "Nombre",
        "Teléfono",
        "Sucursal",
        "Fecha vencimiento",
        "Tarifa",
    )
    assert rows[1] == (
        "Ana Pérez",
        "6861000010",
        "CENTRO",
        "2026-08-23",
        "ANUAL",
    )
    assert filename == "reactivacion_campana_9.xlsx"
    assert campaign.status == "EXPORTED"
    assert campaign.exported_at == NOW
    assert session.commits == 1


def test_export_does_not_mark_exported_when_revalidation_changes(monkeypatch):
    campaign = _campaign()
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **kwargs: campaign)
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: {"eligible_rows": []},
    )

    with pytest.raises(service.MarketingReactivationConflictError):
        service.export_marketing_reactivation_campaign(
            campaign_id=9,
            session=session,
            now=NOW,
        )

    assert campaign.status == "DRAFT"
    assert campaign.exported_at is None
    assert session.commits == 0


def test_export_never_expands_or_shrinks_frozen_campaign_scope(monkeypatch):
    campaign = _campaign()
    campaign.filters_json["scope"] = {
        "is_global": False,
        "allowed_sucursal_keys": ["CENTRO", "NORTE"],
    }
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **_kwargs: campaign)

    with pytest.raises(
        service.MarketingReactivationValidationError,
        match="fuera del alcance actual",
    ):
        service.export_marketing_reactivation_campaign(
            campaign_id=9,
            allowed_sucursal_keys=("CENTRO",),
            session=session,
            now=NOW,
        )

    assert session.commits == 0


def test_export_revalidates_with_original_frozen_scope(monkeypatch):
    campaign = _campaign()
    campaign.filters_json["scope"] = {
        "is_global": False,
        "allowed_sucursal_keys": ["CENTRO"],
    }
    calls = {}
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **_kwargs: campaign)
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: calls.update(kwargs) or {
            "eligible_rows": [
                {"vencido_row_id": 10, "phone_mx10": "6861000010"}
            ]
        },
    )

    service.export_marketing_reactivation_campaign(
        campaign_id=9,
        allowed_sucursal_keys=None,
        session=session,
        now=NOW,
    )

    assert calls["allowed_sucursal_keys"] == ("CENTRO",)


def test_export_rejects_frozen_recipient_outside_campaign_scope(monkeypatch):
    campaign = _campaign()
    campaign.filters_json["scope"] = {
        "is_global": False,
        "allowed_sucursal_keys": ["CENTRO"],
    }
    campaign.recipients[0].sucursal = "NORTE"
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **_kwargs: campaign)

    with pytest.raises(
        service.MarketingReactivationValidationError,
        match="destinatarios fuera",
    ):
        service.export_marketing_reactivation_campaign(
            campaign_id=9,
            allowed_sucursal_keys=("CENTRO",),
            session=session,
            now=NOW,
        )

    assert session.commits == 0


def test_export_writer_failure_does_not_mark_exported(monkeypatch):
    campaign = _campaign()
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **kwargs: campaign)
    monkeypatch.setattr(
        service,
        "_build_campaign_plan",
        lambda **kwargs: {
            "eligible_rows": [
                {"vencido_row_id": 10, "phone_mx10": "6861000010"}
            ]
        },
    )

    class BrokenWorkbook:
        def __init__(self, **kwargs):
            pass

        def create_sheet(self, title):
            return SimpleNamespace(append=lambda row: None)

        def save(self, output):
            raise RuntimeError("xlsx failed")

    monkeypatch.setattr(service, "Workbook", BrokenWorkbook)

    with pytest.raises(RuntimeError, match="xlsx failed"):
        service.export_marketing_reactivation_campaign(
            campaign_id=9,
            session=session,
            now=NOW,
        )

    assert campaign.status == "DRAFT"
    assert campaign.exported_at is None
    assert session.commits == 0


def test_mark_sent_requires_exported_and_uses_small_transaction(monkeypatch):
    campaign = _campaign(status="EXPORTED")
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **kwargs: campaign)

    result = service.mark_marketing_reactivation_campaign_sent(
        campaign_id=9,
        session=session,
        now=NOW,
    )

    assert result["status"] == "SENT"
    assert campaign.sent_at == NOW
    assert session.commits == 1


def test_draft_cannot_transition_directly_to_sent(monkeypatch):
    campaign = _campaign(status="DRAFT")
    session = WriteSession()
    monkeypatch.setattr(service, "_read_campaign", lambda **kwargs: campaign)

    with pytest.raises(service.MarketingReactivationInvalidTransitionError):
        service.mark_marketing_reactivation_campaign_sent(
            campaign_id=9,
            session=session,
            now=NOW,
        )

    assert campaign.status == "DRAFT"
    assert session.commits == 0


@pytest.mark.parametrize(
    ("date_from", "date_to"),
    [("invalid", "2026-08-23"), ("2026-08-24", "2026-08-23")],
)
def test_preview_validates_date_range(date_from, date_to):
    with pytest.raises(service.MarketingReactivationValidationError):
        service.preview_marketing_reactivation_campaign(
            date_from=date_from,
            date_to=date_to,
            filters=_filters(),
            session=SimpleNamespace(),
            now=NOW,
        )


class AggregateQuery:
    def __init__(self, rows):
        self.rows = rows
        self.criteria = []

    def filter(self, *criteria):
        self.criteria.extend(criteria)
        return self

    def with_entities(self, *args):
        return self

    def group_by(self, *args):
        return self

    def order_by(self, *args):
        return self

    def all(self):
        return self.rows


class AggregateSession:
    def __init__(self, rows):
        self.query_object = AggregateQuery(rows)

    def query(self, *entities):
        return self.query_object


def test_tariff_lookup_filters_range_in_sql_and_enriches_unknowns(monkeypatch):
    session = AggregateSession([("  anualidad  ", 4), ("DESCONOCIDA", 2), (None, 1)])
    catalog_row = SimpleNamespace(
        tarifa_key="ANUALIDAD",
        categoria_tarifa="Anualidad",
        reactivation_group="REACTIVATE",
    )
    monkeypatch.setattr(
        service,
        "_read_active_tariff_catalog",
        lambda **kwargs: {"ANUALIDAD": catalog_row},
    )
    query_calls = {}

    def fake_operational_query(**kwargs):
        query_calls.update(kwargs)
        return session.query_object

    monkeypatch.setattr(
        service,
        "build_latest_operational_episode_query",
        fake_operational_query,
    )

    result = service.list_marketing_reactivation_tariffs(
        date_from=date(2026, 8, 1),
        date_to=date(2026, 8, 31),
        session=session,
    )

    assert query_calls["date_from"] == date(2026, 8, 1)
    assert query_calls["date_to"] == date(2026, 8, 31)
    assert result["rows"] == [
        {
            "tarifa": "  anualidad  ",
            "count": 4,
            "classified": True,
            "categoria_tarifa": "Anualidad",
            "reactivation_group": "REACTIVATE",
        },
        {
            "tarifa": "DESCONOCIDA",
            "count": 2,
            "classified": False,
            "categoria_tarifa": None,
            "reactivation_group": None,
        },
        {
            "tarifa": None,
            "count": 1,
            "classified": False,
            "categoria_tarifa": None,
            "reactivation_group": None,
        },
    ]


def test_seed_has_expected_tariffs_without_normalized_duplicates():
    backend_root = Path(__file__).resolve().parents[2]
    seed_path = (
        backend_root
        / "data"
        / "reference"
        / "reactivacion_tarifas_edmundo_seed.json"
    )
    document = json.loads(seed_path.read_text(encoding="utf-8"))
    tariffs = document["tariffs"]
    by_raw = {row["tarifa_raw"]: row for row in tariffs}
    keys = [
        service.normalize_reactivation_tariff_key(row["tarifa_raw"])
        for row in tariffs
    ]

    assert document["source"]["distinct_tariffs"] == 161
    assert document["source"]["conflicts"] == 0
    assert len(tariffs) == len(set(keys)) == 161
    assert len({row["categoria_tarifa"] for row in tariffs}) == 16
    assert by_raw["PASE GYMPASS"]["reactivation_group"] == "EXCLUDE"
    assert by_raw["PASE TOTAL PASS"]["reactivation_group"] == "EXCLUDE"
    assert by_raw["1 DIA"]["reactivation_group"] == "EXCLUDE"
    assert by_raw["1 MES $899"]["reactivation_group"] == "REACTIVATE"
    assert by_raw["DOMICILIADO 12 MESES $499"]["reactivation_group"] == (
        "DOMICILIATED_FLOW"
    )
    assert by_raw["SEMANA"] == {
        "tarifa_raw": "SEMANA",
        "categoria_tarifa": "Trimestre",
        "reactivation_group": "REVIEW",
        "source_count": 62,
    }
    assert by_raw["SEMANA $299"]["reactivation_group"] == "REVIEW"


def test_alembic_has_single_head_for_operational_index_migration():
    versions_path = Path(__file__).resolve().parents[2] / "migrations" / "versions"
    revisions: set[str] = set()
    parents: set[str] = set()
    revision_pattern = re.compile(r'^revision\s*=\s*["\']([^"\']+)["\']', re.M)
    parent_pattern = re.compile(r'^down_revision\s*=\s*(.+)$', re.M)
    quoted_pattern = re.compile(r'["\']([^"\']+)["\']')
    for migration_path in versions_path.glob("*.py"):
        source = migration_path.read_text(encoding="utf-8")
        revision_match = revision_pattern.search(source)
        if revision_match:
            revisions.add(revision_match.group(1))
        parent_match = parent_pattern.search(source)
        if parent_match:
            parents.update(quoted_pattern.findall(parent_match.group(1)))

    assert revisions - parents == {"f7c9a2d4e6b1"}
