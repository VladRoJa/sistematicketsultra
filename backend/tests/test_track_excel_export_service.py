from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.warehouse.services import track_excel_export_service as service


def test_daily_mart_raw_includes_official_branch_income_projection(
    monkeypatch,
):
    row = SimpleNamespace(
        sucursal_canon="TEC_MXL",
        ingreso_real_total_mtd=Decimal("80000"),
        ingreso_real_mtd=Decimal("70000"),
    )

    resolved_version = SimpleNamespace(
        id=123,
        version_type="cierre_canonico",
        status="success",
        generated_at_utc=None,
        finished_at_utc=None,
        started_at_utc=None,
    )

    def fake_projection(**kwargs):
        assert kwargs["sucursal_canon"] == "TEC_MXL"
        assert kwargs["target_month"] == date(2026, 8, 1)
        assert kwargs["cutoff_day"] == 27
        assert kwargs["current_income_mtd"] == Decimal("80000")

        return {
            "status": "available",
            "projected_close": "887460.2043673072",
        }

    monkeypatch.setattr(
        service,
        "build_branch_income_projection_summary",
        fake_projection,
    )

    excel_bytes = service.build_track_daily_mart_excel(
        track_date=date(2026, 8, 27),
        generation_mode="official_closed_day",
        resolved_version=resolved_version,
        rows=[row],
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
        data_only=False,
    )

    worksheet = workbook["Daily Mart Raw"]

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    assert "ingreso_proyectado_cierre" in headers
    assert "ingreso_proyeccion_status" in headers

    projection_value = worksheet.cell(
        row=2,
        column=headers["ingreso_proyectado_cierre"],
    ).value

    projection_status = worksheet.cell(
        row=2,
        column=headers["ingreso_proyeccion_status"],
    ).value

    assert projection_value == 887460.2043673072
    assert projection_status == "available"
