from collections import OrderedDict
from datetime import datetime, timezone
from io import BytesIO

from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload
import pytz

from app.models.suite_governance import (
    SuiteRegionORM,
    SuiteSucursalRegionAssignmentORM,
)
from app.models.ticket_model import Ticket
from app.utils.ticket_filters import filtrar_tickets_por_usuario


BUSINESS_TIMEZONE = pytz.timezone("America/Tijuana")
ACTIVE_TICKET_STATUSES = ("abierto", "en progreso", "por_validar")
LEGACY_UNCLASSIFIED_FAMILY = "SIN CLASIFICAR HISTÓRICO"
PENDING_CAPTURE = "Pendiente de captura"
NO_COMMITMENT = "Sin compromiso"

TICKETS_HEADERS = (
    "Ticket ID",
    "Aparato/Dispositivo",
    "Código Interno",
    "Descripción",
    "Estado",
    "Fecha Creación",
    "Sucursal",
    "Familia",
    "Falla detectada",
    "Condición",
    "Reparación",
    "Observación / Plan de trabajo",
)

SUMMARY_FAMILIES = (
    ("CAMINADORA", "Caminadoras"),
    ("ELIPTICA", "Elípticas"),
    ("ESCALADORA", "Escaladoras"),
    ("RECUMBENTE", "Recumbentes"),
    ("SPINNING", "Spinning"),
    ("PESO_LIBRE", "Peso Libre"),
    ("PESO_INTEGRADO", "Peso Integrado"),
)

FAMILY_SHEETS = (
    ("CAMINADORA", "Caminadoras"),
    ("ELIPTICA", "Elípticas"),
    ("ESCALADORA", "Escaladoras"),
    ("SPINNING", "Spinning"),
    ("RECUMBENTE", "Recumbentes"),
    ("PESO_INTEGRADO", "Peso Integrado"),
    ("PESO_LIBRE", "Peso Libre"),
)

FAMILY_HEADERS = (
    "Sucursal",
    "Ticket ID",
    "ID Equipo",
    "Falla",
    "Condición",
    "Reparación",
    "Observación / Plan de trabajo",
)


POR_VALIDAR_HEADERS = (
    "Sucursal",
    "Ticket ID",
    "Aparato/Dispositivo",
    "Código Interno",
    "Descripción",
    "Familia",
    "Falla detectada",
    "Condición",
    "Desde por validar",
    "Días por validar",
    "Reparación",
    "Observación / Plan de trabajo",
)

class RegionReporteNoEncontradaError(ValueError):
    pass


def listar_regiones_reporte():
    return (
        SuiteRegionORM.query
        .filter(
            SuiteRegionORM.is_active.is_(True),
            SuiteRegionORM.sucursal_assignments.any(
                SuiteSucursalRegionAssignmentORM.is_current.is_(True)
            ),
        )
        .order_by(SuiteRegionORM.region_label.asc())
        .all()
    )


def obtener_region_reporte(region_id):
    region = (
        SuiteRegionORM.query
        .filter(
            SuiteRegionORM.id == region_id,
            SuiteRegionORM.is_active.is_(True),
        )
        .first()
    )
    if not region:
        raise RegionReporteNoEncontradaError(
            "La región solicitada no existe o está inactiva."
        )
    return region


def _obtener_sucursales_ids_region(region_id):
    obtener_region_reporte(region_id)

    rows = (
        SuiteSucursalRegionAssignmentORM.query
        .with_entities(SuiteSucursalRegionAssignmentORM.sucursal_id)
        .filter(
            SuiteSucursalRegionAssignmentORM.region_id == region_id,
            SuiteSucursalRegionAssignmentORM.is_current.is_(True),
        )
        .all()
    )
    return sorted({int(row[0]) for row in rows})


def _query_tickets_mantenimiento(user=None):
    query = (
        filtrar_tickets_por_usuario(user)
        if user is not None
        else Ticket.query
    )

    return (
        query
        .options(
            joinedload(Ticket.inventario),
            joinedload(Ticket.familia_equipo),
            joinedload(Ticket.falla_mantenimiento),
            joinedload(Ticket.sucursal),
            joinedload(Ticket.sucursal_destino),
        )
        .filter(
            Ticket.departamento_id == 1,
            (
                Ticket.aparato_id.isnot(None)
                | Ticket.familia_equipo_id.isnot(None)
            ),
        )
    )


def _aplicar_filtro_region(query, region_id):
    if region_id is None:
        return query

    sucursales_ids = _obtener_sucursales_ids_region(region_id)
    if not sucursales_ids:
        return query.filter(False)

    return query.filter(
        Ticket.sucursal_id_destino.in_(sucursales_ids)
        | (
            Ticket.sucursal_id_destino.is_(None)
            & Ticket.sucursal_id.in_(sucursales_ids)
        )
    )


def obtener_tickets_reporte(user=None, region_id=None):
    query = _query_tickets_mantenimiento(user).filter(
        Ticket.estado.in_(ACTIVE_TICKET_STATUSES)
    )
    query = _aplicar_filtro_region(query, region_id)

    return (
        query
        .order_by(Ticket.sucursal_id_destino.asc(), Ticket.id.asc())
        .all()
    )


def obtener_tickets_historico_reporte(user=None, region_id=None):
    query = _query_tickets_mantenimiento(user)
    query = _aplicar_filtro_region(query, region_id)

    return (
        query
        .order_by(Ticket.fecha_creacion.asc(), Ticket.id.asc())
        .all()
    )


def _safe_branch_name(ticket):
    branch = ticket.sucursal_destino or ticket.sucursal
    if not branch:
        return "Sin sucursal"

    return str(
        getattr(branch, "sucursal", None)
        or getattr(branch, "nombre", None)
        or getattr(branch, "nombre_sucursal", None)
        or "Sin sucursal"
    ).strip()


def _to_business_datetime(value):
    if not value:
        return None
    if isinstance(value, str):
        value = parser.isoparse(value)
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(BUSINESS_TIMEZONE)


def _format_created_at(value):
    local_value = _to_business_datetime(value)
    return local_value.strftime("%d/%m/%Y %H:%M") if local_value else ""


def _format_commitment(value):
    local_value = _to_business_datetime(value)
    return local_value.strftime("%d/%m/%Y") if local_value else NO_COMMITMENT


def _history_sort_key(item):
    value = (
        item.get("fechaCambio")
        or item.get("fecha_cambio")
        or item.get("fecha")
    )
    try:
        parsed = parser.isoparse(value)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except (TypeError, ValueError, OverflowError):
        return datetime.min.replace(tzinfo=timezone.utc)


def plan_trabajo_desde_historial(history):
    normal_entries = []
    for item in history or []:
        if not isinstance(item, dict):
            continue
        if item.get("tipo") or item.get("evento"):
            continue

        motive = str(item.get("motivo") or "").strip()
        if motive:
            normal_entries.append((item, motive))

    normal_entries.sort(key=lambda pair: _history_sort_key(pair[0]))
    return "\n".join(motive for _, motive in normal_entries)


def _snapshot_family_key(ticket):
    family = ticket.familia_equipo
    return str(getattr(family, "key", "") or "").strip().upper() or None


def _snapshot_family_name(ticket):
    family = ticket.familia_equipo
    if not ticket.familia_equipo_id or not family:
        return LEGACY_UNCLASSIFIED_FAMILY
    return str(family.nombre or LEGACY_UNCLASSIFIED_FAMILY).strip()


def _failure_name(ticket):
    failure = ticket.falla_mantenimiento
    return str(failure.nombre).strip() if failure else PENDING_CAPTURE


def _condition_name(ticket):
    return str(ticket.condicion_operativa or PENDING_CAPTURE).strip()


def _equipment_name(ticket):
    inventory = ticket.inventario
    return str(getattr(inventory, "nombre", "") or "").strip()


def _equipment_code(ticket):
    inventory = ticket.inventario
    return str(getattr(inventory, "codigo_interno", "") or "").strip()


def _ticket_row(ticket):
    return (
        ticket.id,
        _equipment_name(ticket),
        _equipment_code(ticket),
        str(ticket.descripcion or "").strip(),
        str(ticket.estado or "").strip(),
        _format_created_at(ticket.fecha_creacion),
        _safe_branch_name(ticket),
        _snapshot_family_name(ticket),
        _failure_name(ticket),
        _condition_name(ticket),
        _format_commitment(ticket.fecha_solucion),
        plan_trabajo_desde_historial(ticket.historial_fechas),
    )


def _family_row(ticket):
    return (
        _safe_branch_name(ticket),
        ticket.id,
        _equipment_code(ticket),
        _failure_name(ticket),
        _condition_name(ticket),
        _format_commitment(ticket.fecha_solucion),
        plan_trabajo_desde_historial(ticket.historial_fechas),
    )


def _validation_wait_days(ticket, now=None):
    pending_since = _to_business_datetime(
        getattr(ticket, "fecha_finalizado", None)
    )
    if pending_since is None:
        return None

    now_local = (
        _to_business_datetime(now)
        if now is not None
        else datetime.now(timezone.utc).astimezone(BUSINESS_TIMEZONE)
    )

    return max(0, (now_local - pending_since).days)


def _por_validar_sort_key(ticket):
    pending_since = _to_business_datetime(
        getattr(ticket, "fecha_finalizado", None)
    )

    return (
        pending_since is None,
        pending_since.timestamp() if pending_since else float("inf"),
        _safe_branch_name(ticket).casefold(),
        int(getattr(ticket, "id", 0) or 0),
    )


def _por_validar_row(ticket):
    return (
        _safe_branch_name(ticket),
        ticket.id,
        _equipment_name(ticket),
        _equipment_code(ticket),
        str(ticket.descripcion or "").strip(),
        _snapshot_family_name(ticket),
        _failure_name(ticket),
        _condition_name(ticket),
        _format_created_at(
            getattr(ticket, "fecha_finalizado", None)
        ),
        _validation_wait_days(ticket),
        _format_commitment(ticket.fecha_solucion),
        plan_trabajo_desde_historial(ticket.historial_fechas),
    )

def _summary_rows(tickets):
    rows_by_branch = OrderedDict()
    family_keys = {
        family_key
        for family_key, _ in SUMMARY_FAMILIES
    }

    for ticket in tickets:
        branch_name = _safe_branch_name(ticket)
        if branch_name not in rows_by_branch:
            rows_by_branch[branch_name] = {
                "total": 0,
                "unclassified": 0,
                **{family_key: 0 for family_key, _ in SUMMARY_FAMILIES},
            }

        counts = rows_by_branch[branch_name]
        counts["total"] += 1

        family_key = _snapshot_family_key(ticket)
        if family_key in family_keys:
            counts[family_key] += 1
        else:
            counts["unclassified"] += 1

    rows = []
    for branch_name in sorted(rows_by_branch, key=str.casefold):
        counts = rows_by_branch[branch_name]
        rows.append(
            (
                branch_name,
                counts["total"],
                *(
                    counts[family_key]
                    for family_key, _ in SUMMARY_FAMILIES
                ),
                counts["unclassified"],
            )
        )
    return rows


def _month_key(value):
    local_value = _to_business_datetime(value)
    if not local_value:
        return None
    return local_value.year, local_value.month


def _next_month_key(month_key):
    year, month = month_key
    if month == 12:
        return year + 1, 1
    return year, month + 1


def _monthly_history_rows(tickets, now=None):
    family_keys = {
        family_key
        for family_key, _ in SUMMARY_FAMILIES
    }
    created_by_month = {}
    finalized_by_month = {}
    first_month = None

    for ticket in tickets:
        created_month = _month_key(
            getattr(ticket, "fecha_creacion", None)
        )

        if created_month is not None:
            if first_month is None or created_month < first_month:
                first_month = created_month

            counts = created_by_month.setdefault(
                created_month,
                {
                    "total": 0,
                    "unclassified": 0,
                    **{
                        family_key: 0
                        for family_key, _ in SUMMARY_FAMILIES
                    },
                },
            )

            counts["total"] += 1

            family_key = _snapshot_family_key(ticket)
            if family_key in family_keys:
                counts[family_key] += 1
            else:
                counts["unclassified"] += 1

        estado = str(
            getattr(ticket, "estado", "") or ""
        ).strip().lower()

        if estado == "finalizado":
            finalized_month = _month_key(
                getattr(ticket, "fecha_finalizado", None)
            )
            if finalized_month is not None:
                finalized_by_month[finalized_month] = (
                    finalized_by_month.get(finalized_month, 0) + 1
                )

    if first_month is None:
        return []

    if now is None:
        now_local = datetime.now(timezone.utc).astimezone(
            BUSINESS_TIMEZONE
        )
    else:
        now_local = _to_business_datetime(now)

    last_month = (now_local.year, now_local.month)

    rows = []
    month_key = first_month

    while month_key <= last_month:
        counts = created_by_month.get(
            month_key,
            {
                "total": 0,
                "unclassified": 0,
                **{
                    family_key: 0
                    for family_key, _ in SUMMARY_FAMILIES
                },
            },
        )

        year, month = month_key
        rows.append(
            (
                datetime(year, month, 1),
                counts["total"],
                *(
                    counts[family_key]
                    for family_key, _ in SUMMARY_FAMILIES
                ),
                counts["unclassified"],
                finalized_by_month.get(month_key, 0),
            )
        )

        month_key = _next_month_key(month_key)

    return rows

def _style_worksheet(worksheet, headers, widths):
    header_fill = PatternFill("solid", fgColor="E54525")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def construir_reporte_xlsx(
    tickets=None,
    *,
    historical_tickets=None,
    user=None,
    region_id=None,
):
    tickets_were_provided = tickets is not None

    tickets = list(
        tickets
        if tickets is not None
        else obtener_tickets_reporte(user=user, region_id=region_id)
    )

    historical_tickets = list(
        historical_tickets
        if historical_tickets is not None
        else (
            tickets
            if tickets_were_provided
            else obtener_tickets_historico_reporte(
                user=user,
                region_id=region_id,
            )
        )
    )
    workbook = Workbook()
    tickets_sheet = workbook.active
    tickets_sheet.title = "Tickets"
    tickets_sheet.append(TICKETS_HEADERS)
    for ticket in tickets:
        tickets_sheet.append(_ticket_row(ticket))
    _style_worksheet(
        tickets_sheet,
        TICKETS_HEADERS,
        (11, 28, 18, 42, 16, 20, 22, 24, 30, 20, 18, 52),
    )

    summary_sheet = workbook.create_sheet("Fallas por familia")
    summary_headers = (
        "Sucursal",
        "Fallas",
        *(label for _, label in SUMMARY_FAMILIES),
        "Sin clasificar",
    )
    summary_sheet.append(summary_headers)
    summary_rows = _summary_rows(tickets)
    for row in summary_rows:
        summary_sheet.append(row)

    total_row = (
        "TOTAL",
        *(
            sum((row[column_index] or 0) for row in summary_rows)
            for column_index in range(1, len(summary_headers))
        ),
    )
    summary_sheet.append(total_row)

    _style_worksheet(
        summary_sheet,
        summary_headers,
        (24, 12, 14, 14, 14, 14, 14, 14, 16, 16),
    )

    validation_sheet = workbook.create_sheet("Por validar")
    validation_sheet.append(POR_VALIDAR_HEADERS)

    validation_tickets = sorted(
        (
            ticket
            for ticket in tickets
            if str(ticket.estado or "").strip().lower()
            == "por_validar"
        ),
        key=_por_validar_sort_key,
    )

    for ticket in validation_tickets:
        validation_sheet.append(_por_validar_row(ticket))

    _style_worksheet(
        validation_sheet,
        POR_VALIDAR_HEADERS,
        (24, 11, 24, 18, 42, 20, 30, 18, 20, 16, 18, 52),
    )
    history_sheet = workbook.create_sheet("Histórico mensual")
    history_headers = (
        "Mes",
        "Fallas",
        *(label for _, label in SUMMARY_FAMILIES),
        "Sin clasificar",
        "Finalizados",
    )
    history_sheet.append(history_headers)

    for row in _monthly_history_rows(historical_tickets):
        history_sheet.append(row)

    _style_worksheet(
        history_sheet,
        history_headers,
        (18, 12, 14, 14, 14, 14, 14, 14, 16, 16, 14),
    )

    for cell in history_sheet["A"][1:]:
        cell.number_format = '[$-es-MX]mmmm yyyy'

    for family_key, sheet_name in FAMILY_SHEETS:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(FAMILY_HEADERS)
        for ticket in tickets:
            if _snapshot_family_key(ticket) == family_key:
                worksheet.append(_family_row(ticket))
        _style_worksheet(
            worksheet,
            FAMILY_HEADERS,
            (24, 11, 18, 32, 20, 18, 52),
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
