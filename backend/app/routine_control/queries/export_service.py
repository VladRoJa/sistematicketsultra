from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


_BUSINESS_TIMEZONE = ZoneInfo("America/Tijuana")


EXPORT_HEADERS = (
    "ID socio", "PIN", "ID folio", "Nombre", "Email", "Teléfono", "Sucursal", "Fecha de venta",
    "Hora de pago", "Clasificación", "Estado", "Primera rutina", "Última rutina", "Instructor actual",
    "Tipo de asignación", "Incidencias activas", "Evidencias activas", "Versión de estado",
)


def _local_payment_time(value: str | None):
    if not value:
        return None

    parsed = datetime.fromisoformat(value)

    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise ValueError("sale_at_utc debe incluir timezone.")

    return (
        parsed
        .astimezone(_BUSINESS_TIMEZONE)
        .time()
        .replace(tzinfo=None)
    )


def build_members_export(items: list[dict]) -> BytesIO:
    workbook = Workbook(write_only=False)
    worksheet = workbook.active
    worksheet.title = "Control de Rutinas"
    worksheet.append(EXPORT_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in items:
        worksheet.append((
            item["external_member_id"], item.get("pin"), item["external_sale_id"],
            item["member_name"], item["email"], item.get("phone"), item["branch_name"], item["sale_date"],
            _local_payment_time(item.get("sale_at_utc")),
            item["classification_status"], item["current_status"], item["first_routine_at"],
            item["latest_routine_at"], item["current_instructor_name"],
            item["routine_assignment_type"], item["active_incident_count"],
            item["active_evidence_count"], item["status_version"],
        ))
    for row in worksheet.iter_rows(min_row=2):
        for index in (8, 12, 13):
            value = row[index - 1].value
            if isinstance(value, str) and value:
                row[index - 1].value = date.fromisoformat(value)
                row[index - 1].number_format = "yyyy-mm-dd"

        if row[8].value is not None:
            row[8].number_format = "hh:mm:ss"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = (18, 14, 18, 28, 30, 18, 22, 14, 14, 18, 18, 14, 14, 24, 20, 18, 18, 16)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
