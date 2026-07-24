from __future__ import annotations

import hashlib
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from uuid import uuid4

from openpyxl import load_workbook

from .contracts import ProviderArtifact


_SAFE_KEY = re.compile(r"^[a-z0-9][a-z0-9_-]{0,63}$")


class ArtifactStoreError(RuntimeError):
    """Artifact inválido o fuera del almacén autorizado."""


class ArtifactStore:
    def __init__(self, root: str | Path) -> None:
        self.root = Path(root).resolve()
        self.root.mkdir(parents=True, exist_ok=True, mode=0o700)
        self._make_private(self.root, directory=True)

    @staticmethod
    def _make_private(path: Path, *, directory: bool = False) -> None:
        try:
            path.chmod(0o700 if directory else 0o600)
        except OSError:
            # Windows aplica ACLs del directorio; chmod no siempre está disponible.
            pass

    @staticmethod
    def _safe_key(value: str, *, field_name: str) -> str:
        if not _SAFE_KEY.fullmatch(value):
            raise ArtifactStoreError(f"{field_name} contiene caracteres no permitidos.")
        return value

    def _inside_root(self, path: str | Path) -> Path:
        resolved = Path(path).resolve()
        try:
            resolved.relative_to(self.root)
        except ValueError as exc:
            raise ArtifactStoreError("El path está fuera del artifact root.") from exc
        return resolved

    def create_run_directory(self, *, provider_key: str, dataset_key: str) -> Path:
        provider = self._safe_key(provider_key, field_name="provider_key")
        dataset = self._safe_key(dataset_key, field_name="dataset_key")
        run_dir = self.root / provider / dataset / uuid4().hex
        run_dir.mkdir(parents=True, exist_ok=False, mode=0o700)
        self._make_private(run_dir, directory=True)
        return run_dir

    def prepare_download(
        self,
        *,
        run_directory: str | Path,
        source_filename: str,
    ) -> tuple[Path, Path]:
        run_dir = self._inside_root(run_directory)
        if not run_dir.is_dir():
            raise ArtifactStoreError("La carpeta de ejecución no existe.")
        name = str(source_filename or "").strip()
        if (
            not name
            or name != Path(name).name
            or "/" in name
            or "\\" in name
            or Path(name).suffix.lower() != ".xlsx"
        ):
            raise ArtifactStoreError("source_filename debe ser un nombre XLSX seguro.")
        final_path = self._inside_root(run_dir / name)
        partial_path = self._inside_root(run_dir / f".{name}.{uuid4().hex}.partial")
        if final_path.exists() or partial_path.exists():
            raise ArtifactStoreError("El artifact de destino ya existe.")
        return partial_path, final_path

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _validate_xlsx(path: Path, required_headers: Iterable[str]) -> None:
        required = frozenset(required_headers)
        try:
            with path.open("rb") as source:
                workbook = load_workbook(source, read_only=True, data_only=True)
                try:
                    matched = False
                    for worksheet in workbook.worksheets:
                        row = next(
                            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                            (),
                        )
                        headers = {
                            str(value).strip()
                            for value in row
                            if value is not None and str(value).strip()
                        }
                        if required.issubset(headers):
                            matched = True
                            break
                    if not matched:
                        missing = ", ".join(sorted(required))
                        raise ArtifactStoreError(
                            f"El XLSX no contiene los headers requeridos: {missing}."
                        )
                finally:
                    workbook.close()
        except ArtifactStoreError:
            raise
        except Exception as exc:
            raise ArtifactStoreError("El archivo descargado no es un XLSX válido.") from exc

    def finalize_download(
        self,
        *,
        partial_path: str | Path,
        final_path: str | Path,
        provider_key: str,
        dataset_key: str,
        required_headers: Iterable[str],
        extracted_at_utc: datetime,
        business_date_from: date,
        business_date_to: date,
        source_filename: str,
        diagnostic_metadata: Mapping[str, Any] | None = None,
    ) -> ProviderArtifact:
        partial = self._inside_root(partial_path)
        final = self._inside_root(final_path)
        try:
            if not partial.is_file():
                raise ArtifactStoreError("La descarga temporal no existe.")
            if final.exists():
                raise ArtifactStoreError("El artifact final ya existe.")
            self._validate_xlsx(partial, required_headers)
            size_bytes = partial.stat().st_size
            sha256 = self._sha256(partial)
            os.replace(partial, final)
            self._make_private(final)
            return ProviderArtifact(
                provider_key=provider_key,
                dataset_key=dataset_key,
                local_path=final,
                sha256=sha256,
                size_bytes=size_bytes,
                extracted_at_utc=extracted_at_utc,
                business_date_from=business_date_from,
                business_date_to=business_date_to,
                source_filename=source_filename,
                diagnostic_metadata=diagnostic_metadata or {},
            )
        except Exception:
            try:
                partial.unlink(missing_ok=True)
            except OSError:
                pass
            raise

    def discard_incomplete(self, path: str | Path) -> None:
        partial = self._inside_root(path)
        try:
            partial.unlink(missing_ok=True)
        except OSError as exc:
            raise ArtifactStoreError("No se pudo limpiar el artifact incompleto.") from exc

