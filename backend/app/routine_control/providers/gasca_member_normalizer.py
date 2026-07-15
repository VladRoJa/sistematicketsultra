from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.routine_control.domain.commands import UpsertRoutineMemberCommand


_SHEET_NAME = "Socios"
_REQUIRED_HEADERS = frozenset(
    {
        "IDSocio",
        "IDFolio",
        "Sucursal",
        "Nombre",
        "ApellidoPaterno",
        "ApellidoMaterno",
        "Email",
        "FechaPago",
        "FechaCreacion",
    }
)
_EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class GascaNormalizationError(ValueError):
    """Error base del normalizador local de Gasca."""


class GascaInvalidWorkbookError(GascaNormalizationError):
    """El archivo no tiene una estructura XLSX utilizable."""


class GascaMissingHeaderError(GascaNormalizationError):
    """Falta al menos un header obligatorio."""


class GascaInvalidRequiredValueError(GascaNormalizationError):
    """Un valor obligatorio no cumple el contrato local."""


class GascaUnsafeIdentifierError(GascaNormalizationError):
    """Un identificador pudo corromperse o perder precisión."""


class _GascaRowRejected(GascaNormalizationError):
    def __init__(self, reason_code: str, reason_message: str) -> None:
        super().__init__(reason_message)
        self.reason_code = reason_code
        self.reason_message = reason_message


@dataclass(frozen=True, slots=True)
class GascaRejectedRow:
    row_number: int
    reason_code: str
    reason_message: str


@dataclass(frozen=True, slots=True)
class GascaMemberBatch:
    commands: tuple[UpsertRoutineMemberCommand, ...]
    rejected_rows: tuple[GascaRejectedRow, ...]
    total_source_rows: int


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _collapse_spaces(value: Any) -> str:
    return " ".join(str(value or "").split())


def _normalize_email(value: Any) -> tuple[str | None, str | None]:
    original = _clean_text(value)
    if original is None:
        return None, None
    normalized = original.lower()
    if original == "77" or _EMAIL_PATTERN.fullmatch(normalized) is None:
        return original, None
    return original, normalized


def _aware_utc(value: datetime) -> datetime:
    if (
        not isinstance(value, datetime)
        or value.tzinfo is None
        or value.utcoffset() is None
    ):
        raise GascaInvalidRequiredValueError(
            "observed_at_utc debe ser datetime con timezone."
        )
    return value.astimezone(timezone.utc)


def _positive_integer_text(value: Any, *, field_name: str) -> str:
    if isinstance(value, bool):
        raise GascaUnsafeIdentifierError(
            f"{field_name} debe ser un identificador entero positivo seguro."
        )
    if isinstance(value, int):
        if value > 0:
            return str(value)
        raise GascaUnsafeIdentifierError(
            f"{field_name} debe ser un identificador entero positivo seguro."
        )
    if isinstance(value, str) and value.isdigit() and int(value) > 0:
        return str(int(value))
    raise GascaUnsafeIdentifierError(
        f"{field_name} debe ser un identificador entero positivo seguro."
    )


def _exact_folio(value: Any) -> str:
    if not isinstance(value, str) or not value or not value.isdigit():
        raise GascaUnsafeIdentifierError(
            "IDFolio debe conservarse como texto de dígitos sin conversión numérica."
        )
    return value


def _parse_sale_date(value: Any) -> date:
    if not isinstance(value, str):
        raise GascaInvalidRequiredValueError(
            "FechaPago debe usar el formato dd-mm-YYYY HH:mm:ss."
        )
    try:
        return datetime.strptime(value, "%d-%m-%Y %H:%M:%S").date()
    except ValueError as exc:
        raise GascaInvalidRequiredValueError(
            "FechaPago debe usar el formato dd-mm-YYYY HH:mm:ss."
        ) from exc


def _canonical_json_hash(payload: Mapping[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=lambda value: value.isoformat()
        if isinstance(value, (date, datetime))
        else TypeError(f"Tipo no serializable: {type(value).__name__}"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _identity_hash(external_member_id: str, external_sale_id: str) -> str:
    identity = f"gasca|new_member|{external_member_id}|{external_sale_id}"
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()


def _metadata(row: Mapping[str, Any]) -> dict[str, Any] | None:
    metadata: dict[str, Any] = {}
    pin = row.get("Pin")
    if isinstance(pin, int) and not isinstance(pin, bool) and pin > 0:
        metadata["pin"] = str(pin)
    elif isinstance(pin, str) and pin.isdigit() and int(pin) > 0:
        metadata["pin"] = str(int(pin))

    for source_key, metadata_key in (
        ("FechaCreacion", "source_created_at"),
        ("TipoMembresia", "membership_type"),
        ("Tarifa", "rate"),
        ("LugarPago", "payment_location"),
    ):
        value = _clean_text(row.get(source_key))
        if value is not None:
            metadata[metadata_key] = value
    return metadata or None


def normalize_gasca_member_row(
    row: Mapping[str, Any],
    *,
    observed_at_utc: datetime,
    branch_resolver: Callable[[str], int | None],
) -> UpsertRoutineMemberCommand:
    observed_at = _aware_utc(observed_at_utc)
    if not callable(branch_resolver):
        raise GascaInvalidRequiredValueError("branch_resolver debe ser callable.")

    external_member_id = _positive_integer_text(
        row.get("IDSocio"), field_name="IDSocio"
    )
    external_sale_id = _exact_folio(row.get("IDFolio"))
    source_branch_name = _clean_text(row.get("Sucursal"))
    if source_branch_name is None:
        raise GascaInvalidRequiredValueError("Sucursal es obligatoria.")

    sucursal_id = branch_resolver(source_branch_name)
    if sucursal_id is None:
        raise _GascaRowRejected(
            "BRANCH_UNRESOLVED",
            "La sucursal de origen no pudo resolverse.",
        )
    if (
        not isinstance(sucursal_id, int)
        or isinstance(sucursal_id, bool)
        or sucursal_id <= 0
    ):
        raise GascaInvalidRequiredValueError(
            "branch_resolver debe devolver un entero positivo o None."
        )

    member_name = " ".join(
        component
        for component in (
            _collapse_spaces(row.get("Nombre")),
            _collapse_spaces(row.get("ApellidoPaterno")),
            _collapse_spaces(row.get("ApellidoMaterno")),
        )
        if component
    )
    if not member_name:
        raise GascaInvalidRequiredValueError("El nombre del miembro es obligatorio.")

    email_original, email_normalized = _normalize_email(row.get("Email"))
    sale_date = _parse_sale_date(row.get("FechaPago"))
    operational_payload = {
        "external_member_id": external_member_id,
        "external_sale_id": external_sale_id,
        "source_branch_name": source_branch_name,
        "sucursal_id": sucursal_id,
        "member_name": member_name,
        "email_original": email_original,
        "email_normalized": email_normalized,
        "sale_date": sale_date,
    }

    return UpsertRoutineMemberCommand(
        source_system="gasca",
        source_record_id=f"{external_member_id}:{external_sale_id}",
        source_identity_key=_identity_hash(
            external_member_id,
            external_sale_id,
        ),
        external_member_id=external_member_id,
        external_sale_id=external_sale_id,
        sucursal_id=sucursal_id,
        source_branch_name=source_branch_name,
        member_name=member_name,
        email_original=email_original,
        email_normalized=email_normalized,
        sale_date=sale_date,
        source_updated_at_utc=None,
        payload_hash=_canonical_json_hash(operational_payload),
        source_metadata=_metadata(row),
        observed_at_utc=observed_at,
    )


def _headers(worksheet: Any) -> tuple[str | None, ...]:
    headers = tuple(cell.value for cell in worksheet[1])
    if not headers or all(header is None for header in headers):
        raise GascaInvalidWorkbookError("La hoja Socios no contiene headers.")
    if len([header for header in headers if header is not None]) != len(
        {header for header in headers if header is not None}
    ):
        raise GascaInvalidWorkbookError("La hoja Socios contiene headers duplicados.")
    missing = sorted(_REQUIRED_HEADERS.difference(headers))
    if missing:
        raise GascaMissingHeaderError(
            f"Faltan headers obligatorios: {', '.join(missing)}."
        )
    return headers


def _rejection_for_exception(
    row_number: int,
    exc: GascaNormalizationError,
) -> GascaRejectedRow:
    if isinstance(exc, _GascaRowRejected):
        code = exc.reason_code
        message = exc.reason_message
    elif isinstance(exc, GascaUnsafeIdentifierError):
        code = "UNSAFE_IDENTIFIER"
        message = str(exc)
    else:
        code = "INVALID_REQUIRED_VALUE"
        message = str(exc)
    return GascaRejectedRow(row_number, code, message)


def load_gasca_member_commands_from_xlsx(
    path: str | Path,
    *,
    observed_at_utc: datetime,
    branch_resolver: Callable[[str], int | None],
) -> GascaMemberBatch:
    observed_at = _aware_utc(observed_at_utc)
    if not callable(branch_resolver):
        raise GascaInvalidRequiredValueError("branch_resolver debe ser callable.")
    source = None
    try:
        source = Path(path).open("rb")
        workbook = load_workbook(source, data_only=True, read_only=False)
    except (OSError, InvalidFileException, BadZipFile, ValueError) as exc:
        if source is not None:
            source.close()
        raise GascaInvalidWorkbookError("No fue posible abrir el archivo XLSX.") from exc

    try:
        if _SHEET_NAME not in workbook.sheetnames:
            raise GascaInvalidWorkbookError("Falta la hoja obligatoria Socios.")
        worksheet = workbook[_SHEET_NAME]
        headers = _headers(worksheet)
        commands: list[UpsertRoutineMemberCommand] = []
        rejected_rows: list[GascaRejectedRow] = []
        total_source_rows = max(worksheet.max_row - 1, 0)

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=len(headers),
                values_only=True,
            ),
            start=2,
        ):
            row = dict(zip(headers, values))
            try:
                commands.append(
                    normalize_gasca_member_row(
                        row,
                        observed_at_utc=observed_at,
                        branch_resolver=branch_resolver,
                    )
                )
            except GascaNormalizationError as exc:
                rejected_rows.append(_rejection_for_exception(row_number, exc))

        return GascaMemberBatch(
            commands=tuple(commands),
            rejected_rows=tuple(rejected_rows),
            total_source_rows=total_source_rows,
        )
    finally:
        workbook.close()
        source.close()
