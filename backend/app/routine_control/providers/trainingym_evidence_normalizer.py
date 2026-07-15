from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel
from openpyxl.utils.exceptions import InvalidFileException

from app.routine_control.domain.commands import RegisterRoutineEvidenceCommand


_SHEET_NAME = "Export"
_REQUIRED_HEADERS = frozenset(
    {
        "id",
        "Idsocioexterno",
        "Email",
        "Técnico",
        "NºRutinas",
        "NºPesajes",
        "Fecha",
        "Centro Origen",
    }
)
_EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


class TrainingymNormalizationError(ValueError):
    """Error base del normalizador local de Trainingym."""


class TrainingymInvalidWorkbookError(TrainingymNormalizationError):
    """El archivo no tiene una estructura XLSX utilizable."""


class TrainingymMissingHeaderError(TrainingymNormalizationError):
    """Falta al menos un header obligatorio."""


class TrainingymInvalidRequiredValueError(TrainingymNormalizationError):
    """Un valor obligatorio no cumple el contrato local."""


class TrainingymUnsafeIdentifierError(TrainingymNormalizationError):
    """Un identificador no puede normalizarse con seguridad."""


class _TrainingymRowRejected(TrainingymNormalizationError):
    def __init__(self, reason_code: str, reason_message: str) -> None:
        super().__init__(reason_message)
        self.reason_code = reason_code
        self.reason_message = reason_message


@dataclass(frozen=True, slots=True)
class TrainingymRejectedRow:
    row_number: int
    reason_code: str
    reason_message: str


@dataclass(frozen=True, slots=True)
class TrainingymEvidenceBatch:
    commands: tuple[RegisterRoutineEvidenceCommand, ...]
    rejected_rows: tuple[TrainingymRejectedRow, ...]
    total_source_rows: int


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _collapse_spaces(value: Any) -> str:
    return " ".join(str(value or "").split())


def _remove_accents_lower(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.lower())
    return "".join(char for char in decomposed if not unicodedata.combining(char))


def _normalize_person_name(value: Any) -> tuple[str, str]:
    original = _collapse_spaces(value)
    if not original:
        raise TrainingymInvalidRequiredValueError("Técnico es obligatorio.")
    normalized = _collapse_spaces(_remove_accents_lower(original))
    return original, normalized


def _normalize_center(value: Any) -> tuple[str, str]:
    original = _collapse_spaces(value)
    if not original:
        raise TrainingymInvalidRequiredValueError("Centro Origen es obligatorio.")
    normalized = _collapse_spaces(_remove_accents_lower(original)).rstrip(".").rstrip()
    if not normalized:
        raise TrainingymInvalidRequiredValueError("Centro Origen es obligatorio.")
    return original, normalized


def _normalize_email(value: Any) -> tuple[str | None, str | None]:
    original = _clean_text(value)
    if original is None:
        return None, None
    normalized = original.lower()
    if _EMAIL_PATTERN.fullmatch(normalized) is None:
        return original, None
    return original, normalized


def _aware_utc(value: datetime) -> datetime:
    if (
        not isinstance(value, datetime)
        or value.tzinfo is None
        or value.utcoffset() is None
    ):
        raise TrainingymInvalidRequiredValueError(
            "observed_at_utc debe ser datetime con timezone."
        )
    return value.astimezone(timezone.utc)


def _provider_run_id(value: int | None) -> int | None:
    if value is not None and (
        not isinstance(value, int) or isinstance(value, bool)
    ):
        raise TrainingymInvalidRequiredValueError(
            "provider_run_id debe ser int o None."
        )
    return value


def _positive_integer_text(value: Any, *, field_name: str) -> str:
    if isinstance(value, bool):
        raise TrainingymUnsafeIdentifierError(
            f"{field_name} debe ser un identificador entero positivo seguro."
        )
    if isinstance(value, int):
        if value > 0:
            return str(value)
        raise TrainingymUnsafeIdentifierError(
            f"{field_name} debe ser un identificador entero positivo seguro."
        )
    if isinstance(value, str) and value.isdigit() and int(value) > 0:
        return str(int(value))
    raise TrainingymUnsafeIdentifierError(
        f"{field_name} debe ser un identificador entero positivo seguro."
    )


def _external_member_id(value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    normalized = _positive_integer_text(value, field_name="Idsocioexterno")
    return None if normalized == "17" else normalized


def _integer_count(value: Any, *, field_name: str, null_as_zero: bool) -> int:
    if value is None and null_as_zero:
        return 0
    if not isinstance(value, int) or isinstance(value, bool):
        raise TrainingymInvalidRequiredValueError(
            f"{field_name} debe ser un entero."
        )
    return value


def _activity_date(value: Any, *, excel_epoch: datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        raise TrainingymInvalidRequiredValueError("Fecha contiene un valor inválido.")
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch=excel_epoch)
        except (TypeError, ValueError, OverflowError) as exc:
            raise TrainingymInvalidRequiredValueError(
                "Fecha contiene un serial Excel inválido."
            ) from exc
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
    raise TrainingymInvalidRequiredValueError(
        "Fecha debe ser date, datetime o serial Excel."
    )


def _canonical_json_hash(payload: Mapping[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=lambda value: value.isoformat()
        if isinstance(value, (date, datetime))
        else TypeError(f"Tipo no serializable: {type(value).__name__}"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _evidence_identity_hash(
    *,
    provider_member_id: str,
    routine_activity_date: date,
    instructor_name_normalized: str,
    provider_center_key: str,
) -> str:
    payload = {
        "provider_member_id": provider_member_id,
        "routine_activity_date": routine_activity_date,
        "instructor_name_normalized": instructor_name_normalized,
        "provider_center_key": provider_center_key,
    }
    return _canonical_json_hash(payload)


def _metadata(row: Mapping[str, Any]) -> dict[str, Any] | None:
    metadata: dict[str, Any] = {}
    valuation = row.get("Valoración")
    if isinstance(valuation, (int, float)) and not isinstance(valuation, bool):
        metadata["valuation"] = valuation
    sex = _clean_text(row.get("Sexo"))
    if sex is not None:
        metadata["sex"] = sex
    total = row.get("Total Rutinas-Pesaje")
    if isinstance(total, (int, float)) and not isinstance(total, bool):
        metadata["routine_weighing_total"] = total
    return metadata or None


def normalize_trainingym_evidence_row(
    row: Mapping[str, Any],
    *,
    observed_at_utc: datetime,
    provider_run_id: int | None,
    center_resolver: Callable[[str], int | None] | None = None,
    excel_epoch: datetime = WINDOWS_EPOCH,
) -> RegisterRoutineEvidenceCommand:
    observed_at = _aware_utc(observed_at_utc)
    run_id = _provider_run_id(provider_run_id)
    if center_resolver is not None and not callable(center_resolver):
        raise TrainingymInvalidRequiredValueError(
            "center_resolver debe ser callable o None."
        )

    provider_member_id = _positive_integer_text(row.get("id"), field_name="id")
    external_member_id = _external_member_id(row.get("Idsocioexterno"))
    instructor_name, instructor_name_normalized = _normalize_person_name(
        row.get("Técnico")
    )
    if "automat" in instructor_name_normalized:
        raise _TrainingymRowRejected(
            "AUTOMATIC_ROUTINE",
            "La rutina fue asignada por un proceso automático.",
        )

    routine_count = _integer_count(
        row.get("NºRutinas"),
        field_name="NºRutinas",
        null_as_zero=False,
    )
    weighing_count = _integer_count(
        row.get("NºPesajes"),
        field_name="NºPesajes",
        null_as_zero=True,
    )
    if weighing_count < 0:
        raise TrainingymInvalidRequiredValueError(
            "NºPesajes debe ser un entero no negativo."
        )
    if routine_count <= 0 and weighing_count > 0:
        raise _TrainingymRowRejected(
            "WEIGHING_ONLY",
            "La fila contiene pesaje pero no rutina.",
        )
    if routine_count <= 0:
        raise _TrainingymRowRejected(
            "NO_ROUTINE",
            "La fila no contiene una rutina.",
        )

    provider_center_name, provider_center_key = _normalize_center(
        row.get("Centro Origen")
    )
    sucursal_id = (
        center_resolver(provider_center_name)
        if center_resolver is not None
        else None
    )
    if sucursal_id is not None and (
        not isinstance(sucursal_id, int)
        or isinstance(sucursal_id, bool)
        or sucursal_id <= 0
    ):
        raise TrainingymInvalidRequiredValueError(
            "center_resolver debe devolver un entero positivo o None."
        )

    activity_date = _activity_date(row.get("Fecha"), excel_epoch=excel_epoch)
    email_original, email_normalized = _normalize_email(row.get("Email"))
    operational_payload = {
        "provider_member_id": provider_member_id,
        "external_member_id": external_member_id,
        "email_original": email_original,
        "email_normalized": email_normalized,
        "provider_center_key": provider_center_key,
        "provider_center_name": provider_center_name,
        "sucursal_id": sucursal_id,
        "routine_activity_date": activity_date,
        "instructor_name": instructor_name,
        "instructor_name_normalized": instructor_name_normalized,
        "routine_count": routine_count,
        "weighing_count": weighing_count,
    }

    return RegisterRoutineEvidenceCommand(
        provider_key="trainingym",
        provider_member_id=provider_member_id,
        evidence_identity_key=_evidence_identity_hash(
            provider_member_id=provider_member_id,
            routine_activity_date=activity_date,
            instructor_name_normalized=instructor_name_normalized,
            provider_center_key=provider_center_key,
        ),
        external_member_id=external_member_id,
        external_routine_id=None,
        email_original=email_original,
        email_normalized=email_normalized,
        provider_center_key=provider_center_key,
        provider_center_name=provider_center_name,
        sucursal_id=sucursal_id,
        routine_activity_date=activity_date,
        instructor_name=instructor_name,
        instructor_name_normalized=instructor_name_normalized,
        routine_count=routine_count,
        weighing_count=weighing_count,
        provider_run_id=run_id,
        payload_hash=_canonical_json_hash(operational_payload),
        source_metadata=_metadata(row),
        observed_at_utc=observed_at,
    )


def _headers(worksheet: Any) -> tuple[str | None, ...]:
    headers = tuple(cell.value for cell in worksheet[1])
    if not headers or all(header is None for header in headers):
        raise TrainingymInvalidWorkbookError("La hoja Export no contiene headers.")
    if len([header for header in headers if header is not None]) != len(
        {header for header in headers if header is not None}
    ):
        raise TrainingymInvalidWorkbookError(
            "La hoja Export contiene headers duplicados."
        )
    missing = sorted(_REQUIRED_HEADERS.difference(headers))
    if missing:
        raise TrainingymMissingHeaderError(
            f"Faltan headers obligatorios: {', '.join(missing)}."
        )
    return headers


def _non_operational_rejection(
    row_number: int,
    values: tuple[Any, ...],
) -> TrainingymRejectedRow | None:
    if all(value is None or value == "" for value in values):
        return TrainingymRejectedRow(
            row_number,
            "EMPTY_ROW",
            "La fila está completamente vacía.",
        )
    first_value = values[0]
    if isinstance(first_value, str):
        cleaned = first_value.strip()
        if cleaned == "Total":
            return TrainingymRejectedRow(
                row_number,
                "SUMMARY_ROW",
                "La fila corresponde al resumen del reporte.",
            )
        if cleaned.startswith("Filtros aplicados:"):
            return TrainingymRejectedRow(
                row_number,
                "FILTER_DESCRIPTION_ROW",
                "La fila describe los filtros del reporte.",
            )
    return None


def _rejection_for_exception(
    row_number: int,
    exc: TrainingymNormalizationError,
) -> TrainingymRejectedRow:
    if isinstance(exc, _TrainingymRowRejected):
        code = exc.reason_code
        message = exc.reason_message
    elif isinstance(exc, TrainingymUnsafeIdentifierError):
        code = "UNSAFE_IDENTIFIER"
        message = str(exc)
    else:
        code = "INVALID_REQUIRED_VALUE"
        message = str(exc)
    return TrainingymRejectedRow(row_number, code, message)


def load_trainingym_evidence_commands_from_xlsx(
    path: str | Path,
    *,
    observed_at_utc: datetime,
    provider_run_id: int | None,
    center_resolver: Callable[[str], int | None] | None = None,
) -> TrainingymEvidenceBatch:
    observed_at = _aware_utc(observed_at_utc)
    run_id = _provider_run_id(provider_run_id)
    if center_resolver is not None and not callable(center_resolver):
        raise TrainingymInvalidRequiredValueError(
            "center_resolver debe ser callable o None."
        )
    source = None
    try:
        source = Path(path).open("rb")
        workbook = load_workbook(source, data_only=True, read_only=False)
    except (OSError, InvalidFileException, BadZipFile, ValueError) as exc:
        if source is not None:
            source.close()
        raise TrainingymInvalidWorkbookError(
            "No fue posible abrir el archivo XLSX."
        ) from exc

    try:
        if _SHEET_NAME not in workbook.sheetnames:
            raise TrainingymInvalidWorkbookError("Falta la hoja obligatoria Export.")
        worksheet = workbook[_SHEET_NAME]
        headers = _headers(worksheet)
        commands: list[RegisterRoutineEvidenceCommand] = []
        rejected_rows: list[TrainingymRejectedRow] = []
        total_source_rows = max(worksheet.max_row - 1, 0)

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=len(headers),
                values_only=True,
            ),
            start=2,
        ):
            non_operational = _non_operational_rejection(row_number, values)
            if non_operational is not None:
                rejected_rows.append(non_operational)
                continue
            row = dict(zip(headers, values))
            try:
                commands.append(
                    normalize_trainingym_evidence_row(
                        row,
                        observed_at_utc=observed_at,
                        provider_run_id=run_id,
                        center_resolver=center_resolver,
                        excel_epoch=workbook.epoch,
                    )
                )
            except TrainingymNormalizationError as exc:
                rejected_rows.append(_rejection_for_exception(row_number, exc))

        return TrainingymEvidenceBatch(
            commands=tuple(commands),
            rejected_rows=tuple(rejected_rows),
            total_source_rows=total_source_rows,
        )
    finally:
        workbook.close()
        source.close()
