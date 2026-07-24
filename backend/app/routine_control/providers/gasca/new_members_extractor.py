from __future__ import annotations

import time
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from app.routine_control.providers.runtime import (
    ArtifactStore,
    ArtifactStoreError,
    BrowserPhase,
    BrowserRuntime,
    ProviderBrowserError,
    ProviderConfigurationError,
    ProviderExtractionResult,
    ProviderRuntimeConfig,
    provider_lock,
)

from .config import GascaProviderConfig


GASCA_PROVIDER_KEY = "gasca"
GASCA_DATASET_KEY = "new_members"
GASCA_NEW_MEMBERS_FILENAME = "gasca-new-members.xlsx"
GASCA_NEW_MEMBER_HEADERS = frozenset(
    {
        "IDSocio",
        "IDFolio",
        "Sucursal",
        "Nombre",
        "ApellidoPaterno",
        "ApellidoMaterno",
        "Email",
        "FechaPago",
        "FechaCreacion",
    }
)
_GASCA_REPORT_OPTION = "Ventas Nuevas Socios"
_GASCA_DATE_SELECTOR = "#txtFechaIn input.form-control"
_GASCA_DOWNLOAD_SELECTOR = "#btnGenerarKpiVentasClientesNuevosDetallado"
_GASCA_DOWNLOAD_TIMEOUT_MS = 240_000
_GASCA_WORKSHEET = "Socios"


class GascaNewMembersExtractionError(RuntimeError):
    provider_retryable = False

    def __init__(self, error_code: str, error_message: str) -> None:
        super().__init__(error_message)
        self.error_code = error_code
        self.attempts = 1


DownloadOperation = Callable[
    [object, object, GascaProviderConfig, date, date, Path],
    str | None,
]


def login_with_verified_gasca_selectors(page, config: GascaProviderConfig) -> None:
    try:
        page.goto(config.login_url, wait_until="domcontentloaded")
        page.locator("#NombreUsuario").fill(config.user)
        page.locator("#Contrasena").fill(config.password)
        page.locator('button[type="submit"]').click()
        page.wait_for_load_state("domcontentloaded")
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_LOGIN_FAILED",
            "Gasca no permitió completar el acceso.",
        ) from exc


def handle_optional_gasca_home_link(page) -> None:
    try:
        home_link = page.get_by_role("link", name="Ir a Inicio", exact=True)
        if home_link.is_visible():
            home_link.click()
            page.wait_for_load_state("domcontentloaded")
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_REPORT_NAVIGATION_FAILED",
            "Gasca no permitió completar la navegación posterior al acceso.",
        ) from exc


def navigate_to_gasca_kpi(page, config: GascaProviderConfig) -> None:
    try:
        page.goto(config.kpi_url, wait_until="domcontentloaded")
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_REPORT_NAVIGATION_FAILED",
            "Gasca no permitió abrir la pantalla de KPI.",
        ) from exc


def select_gasca_new_members_report(page) -> None:
    try:
        page.locator("select").select_option(label=_GASCA_REPORT_OPTION)
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_REPORT_OPTION_NOT_FOUND",
            "Gasca no mostró la opción requerida del reporte.",
        ) from exc


def configure_gasca_cutoff_date(page, date_to: date) -> None:
    try:
        page.locator(_GASCA_DATE_SELECTOR).fill(date_to.strftime("%m/%d/%Y"))
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_DATE_CONTROL_FAILED",
            "Gasca no permitió configurar la fecha de corte.",
        ) from exc


def download_gasca_new_members_xlsx(page, partial_path: Path) -> str:
    try:
        with page.expect_download(
            timeout=_GASCA_DOWNLOAD_TIMEOUT_MS,
        ) as download_info:
            page.locator(_GASCA_DOWNLOAD_SELECTOR).click()
        download = download_info.value
        download.save_as(str(partial_path))
        source_filename = Path(str(download.suggested_filename or "")).name
        if not source_filename:
            raise ValueError("missing suggested filename")
        return source_filename
    except GascaNewMembersExtractionError:
        raise
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_DOWNLOAD_FAILED",
            "Gasca no produjo una descarga XLSX utilizable.",
        ) from exc


def validate_gasca_new_members_xlsx(partial_path: Path) -> None:
    try:
        if not partial_path.is_file():
            raise ValueError("missing download")
        with partial_path.open("rb") as source:
            workbook = load_workbook(source, read_only=True, data_only=True)
            try:
                if _GASCA_WORKSHEET not in workbook.sheetnames:
                    raise ValueError("missing worksheet")
                worksheet = workbook[_GASCA_WORKSHEET]
                first_row = next(
                    worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                    (),
                )
                headers = {
                    str(value).strip()
                    for value in first_row
                    if value is not None and str(value).strip()
                }
                if not GASCA_NEW_MEMBER_HEADERS.issubset(headers):
                    raise ValueError("missing headers")
            finally:
                workbook.close()
    except Exception as exc:
        raise GascaNewMembersExtractionError(
            "GASCA_VALIDATION_FAILED",
            "El XLSX descargado no cumple el contrato de socios nuevos.",
        ) from exc


def run_productive_gasca_download(
    page,
    tracker,
    config: GascaProviderConfig,
    _date_from: date,
    date_to: date,
    partial_path: Path,
) -> str:
    tracker.set(BrowserPhase.LOGIN)
    login_with_verified_gasca_selectors(page, config)
    tracker.set(BrowserPhase.NAVIGATION)
    handle_optional_gasca_home_link(page)
    navigate_to_gasca_kpi(page, config)
    tracker.set(BrowserPhase.DISCOVERY)
    select_gasca_new_members_report(page)
    tracker.set(BrowserPhase.EXPORT)
    configure_gasca_cutoff_date(page, date_to)
    tracker.set(BrowserPhase.DOWNLOAD)
    return download_gasca_new_members_xlsx(page, partial_path)


def _validate_month_to_date_range(date_from: date, date_to: date) -> None:
    if (
        date_from > date_to
        or date_from.day != 1
        or date_from.year != date_to.year
        or date_from.month != date_to.month
    ):
        raise GascaNewMembersExtractionError(
            "GASCA_DATE_CONTROL_FAILED",
            "Gasca sólo admite un rango mensual iniciado el día 1.",
        )


def _discard_partial(store: ArtifactStore, partial_path: Path) -> None:
    try:
        store.discard_incomplete(partial_path)
    except ArtifactStoreError:
        pass


class GascaNewMembersExtractor:
    def __init__(
        self,
        *,
        download_operation: DownloadOperation | None = None,
        runtime_factory: Callable[[ProviderRuntimeConfig], BrowserRuntime] = BrowserRuntime,
    ) -> None:
        self._uses_productive_operation = download_operation is None
        self._download_operation = download_operation or run_productive_gasca_download
        self._runtime_factory = runtime_factory

    @staticmethod
    def _failed(
        *,
        started: float,
        error_code: str,
        error_message: str,
        attempts: int = 0,
    ) -> ProviderExtractionResult:
        return ProviderExtractionResult(
            succeeded=False,
            artifact=None,
            attempts=attempts,
            elapsed_seconds=round(time.monotonic() - started, 3),
            error_code=error_code,
            error_message=error_message,
        )

    def extract(
        self,
        *,
        date_from: date,
        date_to: date,
        observed_at_utc: datetime,
        headless: bool | None = None,
    ) -> ProviderExtractionResult:
        started = time.monotonic()
        try:
            _validate_month_to_date_range(date_from, date_to)
        except GascaNewMembersExtractionError as exc:
            return self._failed(
                started=started,
                error_code=exc.error_code,
                error_message=str(exc),
            )
        if (
            observed_at_utc.tzinfo is None
            or observed_at_utc.utcoffset() is None
        ):
            return self._failed(
                started=started,
                error_code="INVALID_OBSERVED_AT",
                error_message="observed_at_utc debe incluir timezone.",
            )
        try:
            provider_config = GascaProviderConfig.from_env()
            runtime_config = ProviderRuntimeConfig.from_env(headless=headless)
        except ProviderConfigurationError as exc:
            return self._failed(
                started=started,
                error_code="CONFIG_INVALID",
                error_message=str(exc),
            )

        try:
            store = ArtifactStore(runtime_config.artifact_root)
            run_dir = store.create_run_directory(
                provider_key=GASCA_PROVIDER_KEY,
                dataset_key=GASCA_DATASET_KEY,
            )
            partial, final = store.prepare_download(
                run_directory=run_dir,
                source_filename=GASCA_NEW_MEMBERS_FILENAME,
            )
        except (ArtifactStoreError, OSError):
            return self._failed(
                started=started,
                error_code="GASCA_VALIDATION_FAILED",
                error_message="No fue posible preparar el artifact de socios nuevos.",
            )
        execution = None

        def operation(page, tracker, _attempt):
            if not self._uses_productive_operation:
                tracker.set(BrowserPhase.LOGIN)
            source_name = self._download_operation(
                page,
                tracker,
                provider_config,
                date_from,
                date_to,
                partial,
            )
            tracker.set(BrowserPhase.VALIDATION)
            validate_gasca_new_members_xlsx(partial)
            return source_name or GASCA_NEW_MEMBERS_FILENAME

        try:
            with provider_lock(
                runtime_config.artifact_root,
                provider_key=GASCA_PROVIDER_KEY,
                dataset_key=GASCA_DATASET_KEY,
            ):
                execution = self._runtime_factory(runtime_config).run(operation)
                artifact = store.finalize_download(
                    partial_path=partial,
                    final_path=final,
                    provider_key=GASCA_PROVIDER_KEY,
                    dataset_key=GASCA_DATASET_KEY,
                    required_headers=GASCA_NEW_MEMBER_HEADERS,
                    extracted_at_utc=observed_at_utc.astimezone(timezone.utc),
                    business_date_from=date_from,
                    business_date_to=date_to,
                    source_filename=Path(str(execution.value)).name,
                    diagnostic_metadata={
                        "report_contract": "verified_kpi_new_members_detailed",
                    },
                )
            return ProviderExtractionResult(
                succeeded=True,
                artifact=artifact,
                attempts=execution.attempts,
                elapsed_seconds=execution.elapsed_seconds,
            )
        except GascaNewMembersExtractionError as exc:
            _discard_partial(store, partial)
            return self._failed(
                started=started,
                error_code=exc.error_code,
                error_message=str(exc),
                attempts=int(getattr(exc, "attempts", 1)),
            )
        except ProviderBrowserError as exc:
            _discard_partial(store, partial)
            return self._failed(
                started=started,
                error_code=f"GASCA_{exc.phase.value}_FAILED",
                error_message=str(exc),
                attempts=exc.attempts,
            )
        except (ArtifactStoreError, OSError):
            _discard_partial(store, partial)
            return self._failed(
                started=started,
                error_code="GASCA_VALIDATION_FAILED",
                error_message="No fue posible validar el artifact de socios nuevos.",
                attempts=execution.attempts if execution else 1,
            )
        except Exception:
            _discard_partial(store, partial)
            return self._failed(
                started=started,
                error_code="GASCA_VALIDATION_FAILED",
                error_message="No fue posible finalizar el artifact de socios nuevos.",
                attempts=execution.attempts if execution else 1,
            )

