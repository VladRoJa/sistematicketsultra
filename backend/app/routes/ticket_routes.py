# app\routes\ticket_routes.py

from flask import Blueprint, jsonify, request, send_file, current_app
from flask_cors import CORS
from flask_jwt_extended import get_jwt_identity, jwt_required
from datetime import datetime, timezone, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pytz
from sqlalchemy import or_
from app.models.inventario import InventarioGeneral
from app.utils.email_sender import send_email_html
from app.utils.notify_targets import build_subject, pick_recipients
from app.utils.notify_utils import render_ticket_html, render_ticket_whatsapp_text
from app.utils.ticket_filters import filtrar_tickets_por_usuario
from app.config import Config
from app.models.ticket_model import Ticket
from app.models.user_model import UserORM
from app.extensions import db
from app.utils.error_handler import manejar_error
from dateutil import parser
from sqlalchemy.orm.attributes import flag_modified
from app.utils.datetime_utils import format_datetime 
from app.models.sucursal_model import Sucursal
import os, threading
from app.utils.auth_utils import bloquea_lectores_globales
from decimal import Decimal, InvalidOperation





# ─────────────────────────────────────────────────────────────
# BLUEPRINT: TICKETS
# ─────────────────────────────────────────────────────────────
ticket_bp = Blueprint('tickets', __name__, url_prefix='/api/tickets')

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _send_email_maybe_async(to_list, subject, html):
    """
    Envío de correo: si SMTP_SYNC_DEBUG=1 => envío SIN hilo (bloqueante),
    para que los errores aparezcan en el log del request.
    En caso contrario, lo manda en hilo (comportamiento normal).
    """
    try:
        if os.getenv("SMTP_SYNC_DEBUG", "0") == "1":
            current_app.logger.info("🔎 SMTP_SYNC_DEBUG=1 → envío síncrono (sin hilo)")
            send_email_html(to_list, subject, html)
        else:
            threading.Thread(
                target=send_email_html,
                args=(to_list, subject, html),
                daemon=True
            ).start()
    except Exception as e:
        try:
            current_app.logger.exception("❌ Error enviando correo: %s", e)
        except Exception:
            print("❌ Error enviando correo:", e)


def _es_admin_o_corporativo(user: UserORM) -> bool:
    rol = (user.rol or "").upper()
    if rol in {"ADMINISTRADOR", "SUPER_ADMIN", "EDITOR_CORPORATIVO"}:
        return True
    return user.sucursal_id in (100, 1000)

def _es_jefe_depto(user: UserORM, ticket: Ticket) -> bool:
    # jefe si user.department_id coincide con depto del ticket
    try:
        return bool(user.department_id) and int(user.department_id) == int(ticket.departamento_id)
    except Exception:
        return False

def _es_creador(user: UserORM, ticket: Ticket) -> bool:
    return (user.username or "").lower() == (ticket.username or "").lower()

def _es_gerente_regional(user: UserORM) -> bool:
    return (user.rol or "").upper() == "GERENTE_REGIONAL"



# ─────────────────────────────────────────────────────────────
# RUTA: Crear ticket
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/create', methods=['POST'])
@jwt_required()
def create_ticket():
    try:
        usuario_actual = get_jwt_identity()
        user = UserORM.get_by_id(usuario_actual)
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        data = request.get_json() or {}

        descripcion         = data.get("descripcion")
        departamento_id     = data.get("departamento_id")
        criticidad          = data.get("criticidad")
        categoria           = data.get("categoria")
        subcategoria        = data.get("subcategoria")
        detalle             = data.get("detalle")
        aparato_id          = data.get("aparato_id")
        problema_detectado  = data.get("problema_detectado")
        necesita_refaccion  = data.get("necesita_refaccion", False)
        descripcion_refaccion = data.get("descripcion_refaccion")
        clasificacion_id    = data.get("clasificacion_id")
        sucursal_id_destino = data.get("sucursal_id_destino")

        requiere_aprobacion = bool(data.get("requiere_aprobacion", False))
        # Opción B: NO usamos 'pendiente' como estado principal; siempre nace 'abierto'
        estado_inicial = 'abierto'
        aprobacion_estado = 'pendiente' if requiere_aprobacion else None
        # campos de metadatos iniciales de aprobacion
        aprobador_username = None
        aprobacion_fecha = None
        aprobacion_comentario = None


        try:
            departamento_id = int(departamento_id)
            criticidad = int(criticidad)
        except (TypeError, ValueError):
            return jsonify({"mensaje": "departamento_id y criticidad deben ser numéricos"}), 400

        # ¿quién puede fijar el destino?
        es_admin_corp = (user.rol == "ADMINISTRADOR") or (user.sucursal_id in (1000, 100))
        if es_admin_corp:
            if not sucursal_id_destino:
                return jsonify({"mensaje": "Debes enviar sucursal_id_destino"}), 400
        else:
            # usuario normal: destino = su sucursal
            sucursal_id_destino = user.sucursal_id

        # Requisitos mínimos
        if not descripcion or not departamento_id or not criticidad:
            return jsonify({"mensaje": "Faltan datos obligatorios"}), 400

        # Fuente de jerarquía obligatoria
        if not aparato_id and not clasificacion_id:
            return jsonify({"mensaje": "Debes enviar 'aparato_id' (si viene de inventario) o 'clasificacion_id' (si no)."}), 400

        current_app.logger.info("DEBUG Ticket class: %s.%s", getattr(Ticket, "__module__", "?"), getattr(Ticket, "__name__", "?"))


        # Crear ticket
        nuevo_ticket = Ticket.create_ticket(
            descripcion=descripcion,
            username=user.username,
            sucursal_id=user.sucursal_id,              # creador
            sucursal_id_destino=sucursal_id_destino,   # destino
            departamento_id=departamento_id,
            criticidad=criticidad,
            categoria=categoria,
            subcategoria=subcategoria,
            detalle=detalle,
            aparato_id=aparato_id,
            problema_detectado=problema_detectado,
            necesita_refaccion=necesita_refaccion,
            descripcion_refaccion=descripcion_refaccion,
            url_evidencia=data.get("url_evidencia"),
            ubicacion=data.get("ubicacion"),
            equipo=data.get("equipo"),
            clasificacion_id=clasificacion_id,

            # 👇 NUEVO: estado inicial y metadatos de aprobación RRHH
            estado=estado_inicial,
            requiere_aprobacion=requiere_aprobacion,
            aprobacion_estado=aprobacion_estado,
            aprobador_username=aprobador_username,
            aprobacion_fecha=aprobacion_fecha,
            aprobacion_comentario=aprobacion_comentario
        )

        # ⬇️ Notificación de creación + retorno de destinatarios
        recipients = []
        try:
            if os.getenv("NOTIFY_EMAIL_ON_UPDATE", "true").lower() == "true":
                recipients = pick_recipients(nuevo_ticket, actor_username=user.username, event="create") or []
                if recipients:
                    html = render_ticket_html(nuevo_ticket.to_dict())
                    subject = build_subject(nuevo_ticket, "Creado")
                    current_app.logger.info("📧 create_ticket #%s → notificados=%s", nuevo_ticket.id, recipients)
                    send_email_html(recipients, subject, html)
        except Exception as e:
            try:
                current_app.logger.exception("⚠️ No se pudo enviar correo de creación: %s", e)
            except Exception:
                print("⚠️ No se pudo enviar correo de creación:", e)

        return jsonify({
            "mensaje": "Ticket creado correctamente",
            "ticket_id": nuevo_ticket.id,
            "notificados": recipients
        }), 201

    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener todos los tickets (paginados) - SIMPLE
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/all', methods=['GET'])
@jwt_required()
def get_tickets():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        limit = request.args.get('limit', default=15, type=int)
        offset = request.args.get('offset', default=0, type=int)

        # ¡Usa el helper aquí!
        query = filtrar_tickets_por_usuario(user)

        total_tickets = query.count()
        tickets = query.order_by(Ticket.id.desc()).limit(limit).offset(offset).all()

        return jsonify({
            "mensaje": "Tickets cargados correctamente",
            "tickets": [t.to_dict() for t in tickets],
            "total_tickets": total_tickets
        }), 200

    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener tickets con filtros dinámicos
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/list', methods=['GET'])
@jwt_required()
def list_tickets_with_filters():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        estado          = request.args.get('estado')
        departamento_id = request.args.get('departamento_id')
        criticidad      = request.args.get('criticidad')
        no_paging       = request.args.get('no_paging', default='false').lower() == 'true'
        limit           = request.args.get('limit', default=15, type=int)
        offset          = request.args.get('offset', default=0, type=int)

        # Helper universal (aplica permisos por rol/sucursal)
        query = filtrar_tickets_por_usuario(user)

        if estado:
            query = query.filter_by(estado=estado)

        if departamento_id is not None:
            try:
                departamento_id_int = int(departamento_id)
            except (TypeError, ValueError):
                return jsonify({"mensaje": "departamento_id inválido"}), 400
            query = query.filter_by(departamento_id=departamento_id_int)

        if criticidad is not None:
            try:
                criticidad_int = int(criticidad)
            except (TypeError, ValueError):
                return jsonify({"mensaje": "criticidad inválida"}), 400
            query = query.filter_by(criticidad=criticidad_int)

        total_tickets = query.count()
        if not no_paging:
            query = query.limit(limit).offset(offset)

        tickets = query.order_by(Ticket.id.desc()).all()

        return jsonify({
            "mensaje": "Tickets filtrados",
            "tickets": [t.to_dict() for t in tickets],
            "total_tickets": total_tickets
        }), 200

    except Exception as e:
        return manejar_error(e, "list_tickets_with_filters")



# ─────────────────────────────────────────────────────────────
# RUTA: Actualizar estado de un ticket
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/update/<int:id>', methods=['PUT'])
@jwt_required()
@bloquea_lectores_globales
def update_ticket_status(id):
    try:
        actor = UserORM.get_by_id(get_jwt_identity())
        ticket = Ticket.query.get(id)
        if not ticket:
            return jsonify({"mensaje": "Ticket no encontrado"}), 404

        data = request.get_json() or {}
        estado = data.get("estado")
        fecha_solucion = data.get("fecha_solucion")
        fecha_en_progreso = data.get("fecha_en_progreso")
        historial_nuevo = data.get("historial_fechas", [])
        motivo_cambio = (data.get("motivo_cambio") or "").strip()

        prev_estado   = (ticket.estado or "").strip().lower()
        prev_prog     = ticket.fecha_en_progreso
        prev_sol      = ticket.fecha_solucion
        prev_hist_len = len(ticket.historial_fechas or [])

        if not estado:
            return jsonify({"mensaje": "Estado es requerido"}), 400

        ahora = datetime.now(timezone.utc)

        # ⚠️ BLOQUEO: no permitir "finalizado" directo por este endpoint (doble check)
        if estado == "finalizado":
            # Si usas doble check, no se cierra aquí. Se debe usar /cierre/aceptar-creador
            return jsonify({"mensaje": "No puedes finalizar directamente. Usa el flujo de doble check."}), 400

        # ⚠️ BLOQUEO: si requiere aprobación RRHH y no está aprobado, no permitir avanzar a 'en progreso'
        # (ajusta si quieres bloquear otros estados también)
        if estado in ("en progreso",) and getattr(ticket, "requiere_aprobacion", False):
            if (ticket.aprobacion_estado or "").lower() != "aprobado":
                return jsonify({"mensaje": "Este ticket requiere aprobación del gerente antes de avanzar."}), 400

        # ---- Asignación de estado (ya pasó validaciones) ----
        ticket.estado = estado

        # ---- Asignación de fechas por estado (SIN cerrar aquí) ----
        if estado == "en progreso":
            if fecha_en_progreso:
                try:
                    ticket.fecha_en_progreso = parser.isoparse(fecha_en_progreso).astimezone(timezone.utc)
                except Exception as e:
                    print(f"❌ Error parseando fecha_en_progreso: {e}")
                    ticket.fecha_en_progreso = ahora
            else:
                ticket.fecha_en_progreso = ahora

        # Compromiso / fecha objetivo (no cierre)
        if fecha_solucion:
            try:
                fecha_parsed = parser.isoparse(fecha_solucion)
                ticket.fecha_solucion = fecha_parsed.astimezone(timezone.utc)
            except Exception as e:
                print(f"❌ Error parseando fecha_solucion: {e}")

        # ---- Normalizar historial nuevo ----
        historial_final = ticket.historial_fechas or []
        for entrada in historial_nuevo:
            nueva = entrada.copy()
            for campo in ['fecha', 'fechaCambio']:
                valor = nueva.get(campo)
                if valor:
                    try:
                        nueva[campo] = parser.isoparse(valor).astimezone(timezone.utc).isoformat()
                    except Exception as e:
                        print(f"❌ Error parseando {campo} en ticket #{ticket.id}: {e}")
                        continue

            if motivo_cambio and 'motivo' not in nueva:
                nueva['motivo'] = motivo_cambio

            existe_misma_fecha = False
            if nueva.get("fecha"):
                try:
                    nf = parser.isoparse(nueva["fecha"]).replace(tzinfo=None)
                    existe_misma_fecha = any(
                        parser.isoparse(e["fecha"]).replace(tzinfo=None) == nf
                        for e in historial_final
                        if e.get("fecha")
                    )
                except Exception:
                    existe_misma_fecha = False

            if not existe_misma_fecha:
                historial_final.append(nueva)

        def _key_fecha_cambio(x):
            v = x.get('fechaCambio') or x.get('fecha_cambio') or x.get('fecha')
            try:
                return parser.isoparse(v)
            except Exception:
                return datetime.min.replace(tzinfo=timezone.utc)

        historial_final.sort(key=_key_fecha_cambio, reverse=True)
        ticket.historial_fechas = historial_final
        flag_modified(ticket, 'historial_fechas')
        
                # ── Extra (opcional): actualizar refacción cuando el JEFE fija compromiso via /update
        necesita_ref_x = data.get("necesita_refaccion", None)
        descr_ref_x    = data.get("descripcion_refaccion", None)
        flag_jefe_x    = data.get("refaccion_definida_por_jefe", None)

        def _norm(s):
            return (s or "").strip().lower()

        def _ruta_clas_norm(t: Ticket):
            try:
                jer = t._obtener_jerarquia_clasificacion() or []
                return [_norm(x) for x in jer]
            except Exception:
                return []

        # Permisos + combinaciones
        puede_setear_ref = _es_admin_o_corporativo(actor) or _es_jefe_depto(actor, ticket)
        dep_nom = _norm(ticket.departamento.nombre if ticket.departamento else "")
        es_mantenimiento = dep_nom == "mantenimiento"
        es_sistemas      = dep_nom == "sistemas"

        ruta_clas = _ruta_clas_norm(ticket)
        es_aparatos     = "aparatos" in ruta_clas
        es_dispositivos = "dispositivos" in ruta_clas
        tiene_inventario= bool(ticket.aparato_id or ticket.categoria_inventario_id)

        combo_jefe_ok = (es_mantenimiento and es_aparatos) or (es_sistemas and (es_dispositivos or tiene_inventario))

        if puede_setear_ref and combo_jefe_ok:
            if isinstance(necesita_ref_x, bool):
                ticket.necesita_refaccion = necesita_ref_x
                ticket.refaccion_definida_por_jefe = True if flag_jefe_x is None else bool(flag_jefe_x)
            if descr_ref_x is not None:
                ticket.descripcion_refaccion = (descr_ref_x or None)


        db.session.commit()

        # ---- Notificación por correo (si hay cambios) ----
        notificados = []
        try:
            cambios = []
            if estado and (estado.strip().lower() != prev_estado):
                cambios.append(f"estado: {prev_estado or '—'} → {estado}")
            if ticket.fecha_en_progreso != prev_prog:
                cambios.append("fecha en progreso")
            if ticket.fecha_solucion != prev_sol:
                cambios.append("fecha solución")
            if len(ticket.historial_fechas or []) != prev_hist_len:
                cambios.append("historial")

            if os.getenv("NOTIFY_EMAIL_ON_UPDATE", "true").lower() == "true" and cambios:
                notificados = pick_recipients(ticket, actor.username, event="update") or []
                if notificados:
                    subject_bits = " / ".join(cambios) or "Actualización"
                    subject = build_subject(ticket, subject_bits)
                    html = render_ticket_html(ticket.to_dict())
                    current_app.logger.info("📬 Ticket %s cambios=%s → %s", ticket.id, cambios, notificados)
                    _send_email_maybe_async(notificados, subject, html)
        except Exception as e:
            try:
                current_app.logger.exception("❌ Error notificando actualización de ticket %s: %s", ticket.id, e)
            except Exception:
                print(f"❌ Error notificando ticket {ticket.id}:", e)

        return jsonify({
            "mensaje": f"Ticket {id} actualizado correctamente",
            "notificados": notificados
        }), 200

    except Exception as e:
        db.session.rollback()
        return manejar_error(e, "update_ticket_status")

# ─────────────────────────────────────────────────────────────
# RUTA: Exportar tickets a Excel (con filtros)
# ─────────────────────────────────────────────────────────────



@ticket_bp.route('/export-excel', methods=['GET'])
@jwt_required()
def export_excel():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # ---------------- Query params (filtros) ----------------
        estados        = request.args.getlist('estado')
        departamentos  = request.args.getlist('departamento_id')
        criticidades   = request.args.getlist('criticidad')
        usernames      = request.args.getlist('username')
        categorias     = request.args.getlist('categoria')
        subcategorias  = request.args.getlist('subcategoria')
        detalles       = request.args.getlist('detalle')
        descripciones  = request.args.getlist('descripcion')
        inventarios    = request.args.getlist('inventario')

        fecha_desde      = request.args.get('fecha_desde')
        fecha_hasta      = request.args.get('fecha_hasta')
        fecha_fin_desde  = request.args.get('fecha_fin_desde')
        fecha_fin_hasta  = request.args.get('fecha_fin_hasta')
        fecha_prog_desde = request.args.get('fecha_prog_desde')
        fecha_prog_hasta = request.args.get('fecha_prog_hasta')

        # ---------------- Permisos base ----------------
        query = filtrar_tickets_por_usuario(user)

        # ---------------- Aplicación de filtros ----------------
        if estados:
            query = query.filter(Ticket.estado.in_(estados))
        if departamentos:
            query = query.filter(Ticket.departamento_id.in_([int(d) for d in departamentos]))
        if criticidades:
            query = query.filter(Ticket.criticidad.in_([int(c) for c in criticidades]))
        if usernames:
            query = query.filter(Ticket.username.in_(usernames))

        from sqlalchemy import or_
        def filtrar_con_null(campo, valores):
            condiciones = []
            for v in valores:
                condiciones.append(campo.is_(None) if v == "—" else campo == v)
            return or_(*condiciones)

        if categorias:
            query = query.filter(filtrar_con_null(Ticket.categoria, categorias))
        if subcategorias:
            query = query.filter(filtrar_con_null(Ticket.subcategoria, subcategorias))
        if detalles:
            query = query.filter(filtrar_con_null(Ticket.detalle, detalles))
        if descripciones:
            query = query.filter(filtrar_con_null(Ticket.descripcion, descripciones))

        if inventarios:
            inventario_objs = InventarioGeneral.query.filter(
                InventarioGeneral.nombre.in_(inventarios)
            ).all()
            inventario_ids = [inv.id for inv in inventario_objs]
            if "—" in inventarios:
                query = query.filter(or_(
                    Ticket.aparato_id.in_(inventario_ids) if inventario_ids else False,
                    Ticket.aparato_id.is_(None)
                ))
            else:
                query = query.filter(Ticket.aparato_id.in_(inventario_ids))

        from datetime import datetime
        if fecha_desde:
            query = query.filter(Ticket.fecha_creacion >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        if fecha_hasta:
            query = query.filter(Ticket.fecha_creacion <= datetime.strptime(fecha_hasta, '%Y-%m-%d'))
        if fecha_fin_desde:
            query = query.filter(Ticket.fecha_finalizado >= datetime.strptime(fecha_fin_desde, '%Y-%m-%d'))
        if fecha_fin_hasta:
            query = query.filter(Ticket.fecha_finalizado <= datetime.strptime(fecha_fin_hasta, '%Y-%m-%d'))
        if fecha_prog_desde:
            query = query.filter(Ticket.fecha_en_progreso >= datetime.strptime(fecha_prog_desde, '%Y-%m-%d'))
        if fecha_prog_hasta:
            query = query.filter(Ticket.fecha_en_progreso <= datetime.strptime(fecha_prog_hasta, '%Y-%m-%d'))

        # 👉 Más nuevos arriba (por ID)
        tickets = query.order_by(Ticket.id.desc()).all()

        # ===== Helpers =====
        import pytz
        tz_tijuana = pytz.timezone("America/Tijuana")
        hoy_local = datetime.now(tz_tijuana).date()
        SLA_DIAS_BY_CRIT = {1: 14, 2: 7, 3: 5, 4: 2, 5: 1}

        def to_local_excel(dt):
            """Convierte datetime a hora local (naive) para Excel."""
            if not dt:
                return None
            try:
                if getattr(dt, 'tzinfo', None):
                    dt_local = dt.astimezone(tz_tijuana)
                else:
                    dt_local = dt
                return dt_local.replace(tzinfo=None)
            except Exception:
                return None

        def _to_ts_safe(d):
            try:
                if d is None:
                    return 0.0
                if getattr(d, 'tzinfo', None) is None:
                    return d.timestamp()
                return d.replace(tzinfo=None).timestamp()
            except Exception:
                return 0.0

        def ultimos_tres_motivos(ticket, dict_ticket):
            """Extrae hasta 3 motivos/comentarios del historial (compatibilidad)."""
            motivos = []
            rel = None
            for attr in ("historial_fechas_solucion", "historial_fechas", "cambios_fecha_solucion"):
                if hasattr(ticket, attr):
                    rel = getattr(ticket, attr)
                    break
            if rel:
                try:
                    registros = sorted(
                        list(rel),
                        key=lambda r: _to_ts_safe(getattr(r, "fecha_cambio", None) or getattr(r, "fecha", None)),
                        reverse=True
                    )
                    for r in registros:
                        m = getattr(r, "motivo", None) or getattr(r, "comentario", None) or getattr(r, "razon", None)
                        if m:
                            motivos.append(m)
                except Exception:
                    pass

            if not motivos:
                items = dict_ticket.get("historial_fechas") or []
                if isinstance(items, list):
                    try:
                        registros = sorted(
                            items,
                            key=lambda x: x.get("fechaCambio") or x.get("fecha_cambio") or "",
                            reverse=True
                        )
                        for r in registros:
                            m = r.get("motivo") or r.get("comentario") or r.get("razon")
                            if m:
                                motivos.append(m)
                    except Exception:
                        pass

            motivos = (motivos + ["", "", ""])[:3]
            return motivos[0], motivos[1], motivos[2]

        sucursales_map = {s.sucursal_id: s.sucursal for s in Sucursal.query.all()}

        # ===== Excel =====
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"

        # 🔥 Headers con nuevo estado por_validar + subestado cierre
        headers = [
            "ID", "Aparato/Dispositivo", "Código Interno", "Descripción", "Usuario",
            "Estado", "Estado Cierre", "Días por validar",
            "Criticidad", "Fecha Creación", "Fecha En Progreso",
            "Tiempo Transcurrido", "Deber ser", "Fecha Solución",
            "Comentario 1", "Comentario 2", "Comentario 3",
            "Fecha Finalizado", "Tiempo Solución", "Sucursal (destino)",
            "Departamento", "Categoría", "Subcategoria", "Detalle",
            "Problema Detectado", "Refacción", "Descripción Refacción",
            "Costo solución", "Notas cierre",
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0073C2")
        alt_fill    = PatternFill("solid", fgColor="F2F2F2")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # 🔹 Anchos fijos de columnas
        fixed_widths = {
            "A": 8,
            "B": 22,
            "C": 15,
            "D": 45,
            "E": 18,
            "F": 14,   # Estado
            "G": 18,   # Estado Cierre
            "H": 14,   # Días por validar
            "I": 12,   # Criticidad
            "J": 20,   # Fecha creación
            "K": 20,   # Fecha en progreso
            "L": 14,   # Tiempo transcurrido
            "M": 12,   # Deber ser
            "N": 20,   # Fecha solución
            "O": 25,
            "P": 25,
            "Q": 25,
            "R": 20,   # Fecha finalizado
            "S": 14,   # Tiempo solución
            "T": 20,   # Sucursal
            "U": 20,   # Depto
            "V": 20,   # Categoría
            "W": 20,   # Subcategoría
            "X": 30,   # Detalle
            "Y": 30,   # Problema detectado
            "Z": 12,   # Refacción
            "AA": 32,  # Desc refacción
            "AB": 14,  # Costo
            "AC": 40,  # Notas
        }
        for col, width in fixed_widths.items():
            ws.column_dimensions[col].width = width

        excel_datetime_fmt = 'dd/mm/yyyy hh:mm AM/PM'

        # 🔹 Wrap text
        wrap_cols = ["D", "O", "P", "Q", "X", "Y", "AA", "AC"]  # ajustado a nuevas letras
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # Para hoja "Detenidos"
        detenidos_rows = []

        # Para hoja "DataQuality"
        dq_rows = []

        for idx, ticket in enumerate(tickets, start=2):
            t = ticket.to_dict()

            # Aparato/Dispositivo y código
            if ticket.inventario and ticket.inventario.nombre:
                aparato_nombre = ticket.inventario.nombre
                codigo_interno = ticket.inventario.codigo_interno or "—"
            elif ticket.equipo:
                aparato_nombre = ticket.equipo
                codigo_interno = "—"
            else:
                aparato_nombre = "—"
                codigo_interno = "—"

            # Sucursal destino
            suc_id_dest = ticket.sucursal_id_destino if ticket.sucursal_id_destino is not None else ticket.sucursal_id
            sucursal_nombre_dest = sucursales_map.get(suc_id_dest, "—")

            departamento_txt = ticket.departamento.nombre if ticket.departamento else "—"

            # Fechas (naive, sin tz) para Excel
            f_crea_dt = to_local_excel(ticket.fecha_creacion)
            f_prog_dt = to_local_excel(ticket.fecha_en_progreso)
            f_sol_dt  = to_local_excel(ticket.fecha_solucion)
            f_fin_dt  = to_local_excel(ticket.fecha_finalizado)

            # Para cálculos
            f_crea_d = f_crea_dt.date() if f_crea_dt else None
            f_prog_d = f_prog_dt.date() if f_prog_dt else None
            f_fin_d  = f_fin_dt.date() if f_fin_dt else None

            tiempo_trans = (hoy_local - f_crea_d).days if f_crea_d else ""
            deber_ser    = SLA_DIAS_BY_CRIT.get(ticket.criticidad, "")
            comentario1, comentario2, comentario3 = ultimos_tres_motivos(ticket, t)

            tiempo_sol = (f_fin_d - f_crea_d).days if (f_crea_d and f_fin_d) else "N/A"

            categoria_txt    = t.get("categoria")    or "—"
            subcategoria_txt = t.get("subcategoria") or "—"
            detalle_txt      = t.get("detalle")      or "—"

            # Estado y sub-estado
            estado_txt = (t.get("estado") or "").strip()
            estado_cierre_txt = (t.get("estado_cierre") or "").strip() if isinstance(t.get("estado_cierre"), str) else (ticket.estado_cierre or None)

            # “Por validar” aging (días desde solicitud de cierre)
            dias_por_validar = ""
            if estado_txt == "por_validar":
                # usamos fecha_finalizado como “fecha solicitud cierre”
                if f_fin_d:
                    dias_por_validar = max(0, (hoy_local - f_fin_d).days)

            row = [
                t.get("id"),
                aparato_nombre,
                codigo_interno,
                t.get("descripcion"),
                t.get("username"),

                estado_txt,
                estado_cierre_txt or "—",
                dias_por_validar,

                t.get("criticidad"),
                f_crea_dt,
                f_prog_dt,

                tiempo_trans,
                deber_ser,
                f_sol_dt,

                comentario1,
                comentario2,
                comentario3,

                f_fin_dt,
                tiempo_sol,
                sucursal_nombre_dest,

                departamento_txt,
                categoria_txt,
                subcategoria_txt,
                detalle_txt,

                t.get("problema_detectado"),
                "Sí" if t.get("necesita_refaccion") else "No",
                t.get("descripcion_refaccion"),
                float(ticket.costo_solucion) if ticket.costo_solucion is not None else None,
                ticket.notas_cierre,
            ]
            ws.append(row)

            # Zebra
            if idx % 2 == 0:
                for cell in ws[idx]:
                    cell.fill = alt_fill

            # Formato fechas: J=10, K=11, N=14, R=18
            for col in (10, 11, 14, 18):
                c = ws.cell(row=idx, column=col)
                if c.value:
                    c.number_format = excel_datetime_fmt

            # Wrap text
            for col in wrap_cols:
                c = ws[f"{col}{idx}"]
                c.alignment = wrap_alignment

            # Acumular hoja "Detenidos"
            if estado_txt == "por_validar":
                detenidos_rows.append((
                    t.get("id"),
                    sucursal_nombre_dest,
                    departamento_txt,
                    t.get("criticidad"),
                    t.get("username"),
                    f_fin_dt,
                    dias_por_validar if dias_por_validar != "" else None,
                    float(ticket.costo_solucion) if ticket.costo_solucion is not None else None,
                    ticket.notas_cierre or None,
                ))

            # Acumular DataQuality (fechas inconsistentes)
            #  - En progreso antes que creación
            #  - Finalizado antes que creación
            #  - Tiempo solución negativo
            dq_flags = []
            if f_crea_dt and f_prog_dt and f_prog_dt < f_crea_dt:
                dq_flags.append("fecha_en_progreso < fecha_creacion")
            if f_crea_dt and f_fin_dt and f_fin_dt < f_crea_dt:
                dq_flags.append("fecha_finalizado < fecha_creacion")
            if isinstance(tiempo_sol, (int, float)) and tiempo_sol < 0:
                dq_flags.append("tiempo_solucion_negativo")
            if estado_txt == "en progreso" and not f_sol_dt:
                dq_flags.append("ticket_en_progreso_sin_fecha_compromiso")    

            if dq_flags:
                dq_rows.append((
                    t.get("id"),
                    estado_txt,
                    sucursal_nombre_dest,
                    departamento_txt,
                    t.get("username"),
                    f_crea_dt,
                    f_prog_dt,
                    f_fin_dt,
                    "; ".join(dq_flags),
                ))

        # Filtros y freeze panes
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ─────────────────────────────────────────────────────────────
        # Hoja KPI
        # ─────────────────────────────────────────────────────────────
        ws_kpi = wb.create_sheet(title="KPI")

        # Índices (1-based)
        COL_ESTADO         = headers.index("Estado") + 1
        COL_DEBER          = headers.index("Deber ser") + 1
        COL_T_SOL          = headers.index("Tiempo Solución") + 1
        COL_FECHA_CREACION = headers.index("Fecha Creación") + 1
        COL_CRITICIDAD     = headers.index("Criticidad") + 1
        COL_DIAS_VALIDAR   = headers.index("Días por validar") + 1

        total = abiertos = en_progreso = por_validar = finalizados = 0
        cumplidos_tiempo = 0

        # Métricas extra
        tiempos_solucion_dias = []
        abiertos_criticos = 0
        backlog_aging_buckets = {"0-2": 0, "3-7": 0, "8-14": 0, "15+": 0}
        sum_edad_backlog = 0
        count_backlog = 0

        dias_por_validar_list = []
        tickets_en_riesgo = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            total += 1

            estado = (row[COL_ESTADO-1] or "").strip().lower()
            deber  = row[COL_DEBER-1]
            tsol   = row[COL_T_SOL-1]
            crit   = row[COL_CRITICIDAD-1]
            f_crea = row[COL_FECHA_CREACION-1]
            dv     = row[COL_DIAS_VALIDAR-1]

            if estado == "finalizado":
                finalizados += 1
                if isinstance(tsol, (int, float)):
                    tiempos_solucion_dias.append(tsol)
                    if isinstance(deber, (int, float)) and deber > 0 and tsol <= deber:
                        cumplidos_tiempo += 1

            elif estado == "en progreso":
                en_progreso += 1

            elif estado == "por_validar":
                por_validar += 1
                if isinstance(dv, (int, float)):
                    dias_por_validar_list.append(dv)

            else:
                abiertos += 1

            # Backlog = abierto + en progreso + por_validar
            if estado in ("abierto", "en progreso", "por_validar"):
                if isinstance(crit, (int, float)) and crit >= 4:
                    abiertos_criticos += 1

                if f_crea and hasattr(f_crea, "date"):
                    edad = (hoy_local - f_crea.date()).days
                    edad = max(0, edad)

                    sum_edad_backlog += edad
                    count_backlog += 1

                    if   edad <= 2:  backlog_aging_buckets["0-2"]  += 1
                    elif edad <= 7:  backlog_aging_buckets["3-7"]  += 1
                    elif edad <= 14: backlog_aging_buckets["8-14"] += 1
                    else:            backlog_aging_buckets["15+"]  += 1

                    if isinstance(deber, (int, float)) and deber > 0 and edad > deber:
                        tickets_en_riesgo += 1

        def _mediana(nums):
            if not nums: return 0.0
            v = sorted(nums); n = len(v); m = n // 2
            return float(v[m]) if (n % 2 == 1) else round((v[m-1] + v[m]) / 2.0, 2)

        pct_cumplidos = (cumplidos_tiempo / finalizados * 100.0) if finalizados else 0.0
        prom_tiempo_sol = round(sum(tiempos_solucion_dias)/len(tiempos_solucion_dias), 2) if tiempos_solucion_dias else 0.0
        mediana_tiempo_sol = _mediana(tiempos_solucion_dias)

        edad_promedio_backlog = round(sum_edad_backlog / count_backlog, 2) if count_backlog else 0.0
        prom_dias_validar = round(sum(dias_por_validar_list)/len(dias_por_validar_list), 2) if dias_por_validar_list else 0.0
        mediana_dias_validar = _mediana(dias_por_validar_list)

        ws_kpi.append(["Métrica", "Valor"])
        ws_kpi.append(["Total tickets", total])
        ws_kpi.append(["Abiertos", abiertos])
        ws_kpi.append(["En progreso", en_progreso])
        ws_kpi.append(["Por validar", por_validar])
        ws_kpi.append(["Finalizados", finalizados])
        ws_kpi.append(["% cumplidos en tiempo (solo finalizados)", pct_cumplidos])
        ws_kpi.append(["Tiempo promedio de solución (días)", prom_tiempo_sol])
        ws_kpi.append(["Tiempo mediano de solución (días)", mediana_tiempo_sol])
        ws_kpi.append(["Abiertos críticos (criticidad 4-5) en backlog", abiertos_criticos])
        ws_kpi.append(["Edad promedio del backlog (días)", edad_promedio_backlog])
        ws_kpi.append(["Backlog 0-2 días", backlog_aging_buckets["0-2"]])
        ws_kpi.append(["Backlog 3-7 días", backlog_aging_buckets["3-7"]])
        ws_kpi.append(["Backlog 8-14 días", backlog_aging_buckets["8-14"]])
        ws_kpi.append(["Backlog 15+ días", backlog_aging_buckets["15+"]])
        ws_kpi.append(["Días promedio por validar", prom_dias_validar])
        ws_kpi.append(["Días mediana por validar", mediana_dias_validar])
        ws_kpi.append(["Tickets backlog en riesgo SLA", tickets_en_riesgo])

        from openpyxl.styles import Alignment as KPIAlignment
        hdr = ws_kpi["A1":"B1"][0]
        for c in hdr:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0073C2")

        # Formato de porcentaje en fila de % cumplidos (buscamos la etiqueta)
        for r in range(2, ws_kpi.max_row + 1):
            ws_kpi[f"B{r}"].alignment = KPIAlignment(horizontal="right")
            if ws_kpi[f"A{r}"].value == "% cumplidos en tiempo (solo finalizados)":
                ws_kpi[f"B{r}"].number_format = "0.00"
        for col in ws_kpi.columns:
            mx = max(len(str(c.value or "")) for c in col)
            ws_kpi.column_dimensions[col[0].column_letter].width = mx + 2

        # ─────────────────────────────────────────────────────────────
        # Hoja Tablas (resúmenes) + gráficas
        # ─────────────────────────────────────────────────────────────
        ws_tab = wb.create_sheet(title="Tablas")

        from collections import Counter, defaultdict
        COL_ID             = headers.index("ID") + 1
        COL_SUCURSAL       = headers.index("Sucursal (destino)") + 1
        COL_DEPTO          = headers.index("Departamento") + 1
        COL_CATEGORIA      = headers.index("Categoría") + 1
        COL_ESTADO_T       = headers.index("Estado") + 1
        COL_FECHA_CREACION = headers.index("Fecha Creación") + 1
        COL_FECHA_FINAL    = headers.index("Fecha Finalizado") + 1
        COL_DEBER_T        = headers.index("Deber ser") + 1
        COL_T_SOL_T        = headers.index("Tiempo Solución") + 1
        COL_CRITICIDAD_T   = headers.index("Criticidad") + 1

        sucursal_counts  = defaultdict(lambda: {"total": 0, "abiertos": 0, "en_progreso": 0, "por_validar": 0, "finalizados": 0})
        depto_counts     = defaultdict(lambda: {"total": 0, "abiertos": 0, "en_progreso": 0, "por_validar": 0, "finalizados": 0})
        categoria_counts = defaultdict(lambda: {"total": 0, "abiertos": 0, "en_progreso": 0, "por_validar": 0, "finalizados": 0})

        creados_por_mes = Counter()
        finalizados_por_mes = Counter()
        sla_incumplidos = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue

            tid      = row[COL_ID-1]
            sucursal = row[COL_SUCURSAL-1] or "—"
            depto    = row[COL_DEPTO-1] or "—"
            categoria= row[COL_CATEGORIA-1] or "—"
            estado   = (row[COL_ESTADO_T-1] or "").strip().lower()
            f_crea   = row[COL_FECHA_CREACION-1]
            f_fin    = row[COL_FECHA_FINAL-1]
            deber    = row[COL_DEBER_T-1]
            tsol     = row[COL_T_SOL_T-1]
            crit     = row[COL_CRITICIDAD_T-1]

            # Sucursal
            sucursal_counts[sucursal]["total"] += 1
            if estado == "abierto":
                sucursal_counts[sucursal]["abiertos"] += 1
            elif estado == "en progreso":
                sucursal_counts[sucursal]["en_progreso"] += 1
            elif estado == "por_validar":
                sucursal_counts[sucursal]["por_validar"] += 1
            elif estado == "finalizado":
                sucursal_counts[sucursal]["finalizados"] += 1

            # Departamento
            depto_counts[depto]["total"] += 1
            if estado == "abierto":
                depto_counts[depto]["abiertos"] += 1
            elif estado == "en progreso":
                depto_counts[depto]["en_progreso"] += 1
            elif estado == "por_validar":
                depto_counts[depto]["por_validar"] += 1
            elif estado == "finalizado":
                depto_counts[depto]["finalizados"] += 1

            # Categoría
            categoria_counts[categoria]["total"] += 1
            if estado == "abierto":
                categoria_counts[categoria]["abiertos"] += 1
            elif estado == "en progreso":
                categoria_counts[categoria]["en_progreso"] += 1
            elif estado == "por_validar":
                categoria_counts[categoria]["por_validar"] += 1
            elif estado == "finalizado":
                categoria_counts[categoria]["finalizados"] += 1

            if f_crea and hasattr(f_crea, "strftime"):
                creados_por_mes[f_crea.strftime("%Y-%m")] += 1
            if f_fin and hasattr(f_fin, "strftime"):
                finalizados_por_mes[f_fin.strftime("%Y-%m")] += 1

            # SLA incumplidos (solo finalizados)
            if estado == "finalizado" and isinstance(tsol, (int, float)) and isinstance(deber, (int, float)) and deber > 0:
                if tsol > deber:
                    sla_incumplidos.append((tid, sucursal, depto, crit, deber, tsol))

        def write_section(ws_local, start_row, title, cols, rows_local):
            ws_local.cell(row=start_row, column=1, value=title)
            ws_local.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cols))
            tcell = ws_local.cell(row=start_row, column=1)
            tcell.font = Font(bold=True, color="FFFFFF")
            tcell.fill = PatternFill("solid", fgColor="0073C2")

            hdr_row = start_row + 1
            for j, h in enumerate(cols, start=1):
                c = ws_local.cell(row=hdr_row, column=j, value=h)
                c.font = Font(bold=True)

            r = hdr_row + 1
            for row_vals in rows_local:
                for j, v in enumerate(row_vals, start=1):
                    ws_local.cell(row=r, column=j, value=v)
                r += 1

            return r + 1  # deja línea en blanco

        # Tablas ordenadas
        suc_rows_sorted = sorted(sucursal_counts.items(),  key=lambda x: x[1]["total"], reverse=True)
        dep_rows_sorted = sorted(depto_counts.items(),     key=lambda x: x[1]["total"], reverse=True)
        cat_rows_sorted = sorted(categoria_counts.items(), key=lambda x: x[1]["total"], reverse=True)
        cre_rows  = sorted(creados_por_mes.items())
        fin_rows  = sorted(finalizados_por_mes.items())

        suc_table = [(s, v["total"], v["abiertos"], v["en_progreso"], v["por_validar"], v["finalizados"]) for s, v in suc_rows_sorted]
        dep_table = [(d, v["total"], v["abiertos"], v["en_progreso"], v["por_validar"], v["finalizados"]) for d, v in dep_rows_sorted]
        cat_table = [(c, v["total"], v["abiertos"], v["en_progreso"], v["por_validar"], v["finalizados"]) for c, v in cat_rows_sorted]

        r = 1
        suc_start = r
        r = write_section(ws_tab, r, "Tickets por Sucursal",
                          ["Sucursal", "Cantidad", "Abiertos", "En progreso", "Por validar", "Finalizados"],
                          suc_table)

        dep_start = r
        r = write_section(ws_tab, r, "Tickets por Departamento",
                          ["Departamento", "Cantidad", "Abiertos", "En progreso", "Por validar", "Finalizados"],
                          dep_table)

        cat_start = r
        r = write_section(ws_tab, r, "Tickets por Categoría",
                          ["Categoría", "Cantidad", "Abiertos", "En progreso", "Por validar", "Finalizados"],
                          cat_table)

        r = write_section(ws_tab, r, "Tickets creados por mes", ["Mes", "Cantidad"], cre_rows)
        r = write_section(ws_tab, r, "Tickets finalizados por mes", ["Mes", "Cantidad"], fin_rows)
        r = write_section(ws_tab, r, "Tickets SLA Incumplidos",
                          ["ID", "Sucursal", "Departamento", "Criticidad", "Deber ser", "Tiempo Solución"],
                          sla_incumplidos)

        # Auto-width
        for col_idx in range(1, ws_tab.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws_tab.max_row + 1):
                cell = ws_tab.cell(row=row_idx, column=col_idx)
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws_tab.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        # ── Gráficas: Barras (Abiertos) + Líneas (En progreso / Por validar / Finalizados)
        from openpyxl.chart import BarChart, LineChart, Reference

        def add_combo_chart(ws_local, title, start_row_title, n_rows, anchor):
            if n_rows <= 0:
                return

            header_row = start_row_title + 1
            first_data = header_row + 1
            last_data  = first_data + n_rows - 1

            cats = Reference(ws_local, min_col=1, min_row=first_data, max_row=last_data)

            data_abiertos = Reference(ws_local, min_col=3, min_row=header_row, max_col=3, max_row=last_data)
            bar = BarChart()
            bar.title = title
            bar.y_axis.title = "Cantidad"
            bar.add_data(data_abiertos, titles_from_data=True)
            bar.set_categories(cats)

            data_line = Reference(ws_local, min_col=4, min_row=header_row, max_col=6, max_row=last_data)
            line = LineChart()
            line.add_data(data_line, titles_from_data=True)
            line.set_categories(cats)

            for i, s in enumerate(line.series):
                s.marker.symbol = "circle" if i == 0 else ("triangle" if i == 1 else "diamond")
                s.smooth = False

            bar += line
            bar.height = 12
            bar.width  = 22
            ws_local.add_chart(bar, anchor)

        add_combo_chart(ws_tab, "Tickets por Sucursal",     suc_start, len(suc_table), "H2")
        add_combo_chart(ws_tab, "Tickets por Departamento", dep_start, len(dep_table), "H28")
        add_combo_chart(ws_tab, "Tickets por Categoría",    cat_start, len(cat_table), "H54")

        # ─────────────────────────────────────────────────────────────
        # Hoja Detenidos (Top por_validar)
        # ─────────────────────────────────────────────────────────────
        ws_det = wb.create_sheet(title="Detenidos")

        ws_det.append(["ID", "Sucursal", "Departamento", "Criticidad", "Usuario", "Fecha solicitud cierre", "Días detenido", "Costo", "Notas cierre"])
        for cell in ws_det[1]:
            cell.font = header_font
            cell.fill = header_fill

        # ordenar por días detenido desc (None al final)
        def _dv_key(x):
            dv = x[6]
            return (-dv) if isinstance(dv, (int, float)) else 999999

        detenidos_rows_sorted = sorted(detenidos_rows, key=_dv_key)
        for i, rr in enumerate(detenidos_rows_sorted[:50], start=2):
            ws_det.append(list(rr))
            if i % 2 == 0:
                for cell in ws_det[i]:
                    cell.fill = alt_fill

            # fecha solicitud cierre (col 6)
            c = ws_det.cell(row=i, column=6)
            if c.value:
                c.number_format = excel_datetime_fmt

        # autosize Detenidos
        for col_idx in range(1, ws_det.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws_det.max_row + 1):
                cell = ws_det.cell(row=row_idx, column=col_idx)
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws_det.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        # ─────────────────────────────────────────────────────────────
        # Hoja DataQuality
        # ─────────────────────────────────────────────────────────────
        ws_dq = wb.create_sheet(title="DataQuality")
        ws_dq.append(["ID", "Estado", "Sucursal", "Departamento", "Usuario", "Fecha Creación", "Fecha En Progreso", "Fecha Finalizado", "Problema"])
        for cell in ws_dq[1]:
            cell.font = header_font
            cell.fill = header_fill

        for i, rr in enumerate(dq_rows[:500], start=2):  # limitamos para que no pese demasiado
            ws_dq.append(list(rr))
            if i % 2 == 0:
                for cell in ws_dq[i]:
                    cell.fill = alt_fill

            for col in (6, 7, 8):
                c = ws_dq.cell(row=i, column=col)
                if c.value:
                    c.number_format = excel_datetime_fmt

        for col_idx in range(1, ws_dq.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws_dq.max_row + 1):
                cell = ws_dq.cell(row=row_idx, column=col_idx)
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws_dq.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        # ─────────────────────────────────────────────────────────────
        # Hoja Indicadores 2 (vista agrupada por sucursal)
        # ─────────────────────────────────────────────────────────────
        ws_ind2 = wb.create_sheet(title="Indicadores 2")

        headers_ind2 = [
            "Sucursal",
            "Descripción",
            "Fecha de creación",
            "Fecha compromiso",
            "Comentario de dpto",
            "Tickets en progreso",
        ]
        ws_ind2.append(headers_ind2)

        for cell in ws_ind2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        from collections import defaultdict

        tickets_por_sucursal = defaultdict(list)

        for ticket in tickets:
            t = ticket.to_dict()

            suc_id_dest = ticket.sucursal_id_destino if ticket.sucursal_id_destino is not None else ticket.sucursal_id
            sucursal_nombre_dest = sucursales_map.get(suc_id_dest, "—")

            f_crea_dt = to_local_excel(ticket.fecha_creacion)
            f_sol_dt = to_local_excel(ticket.fecha_solucion)

            comentario1, comentario2, comentario3 = ultimos_tres_motivos(ticket, t)

            estado_txt = (t.get("estado") or "").strip().lower()

            if estado_txt != "en progreso":
                continue

            tickets_por_sucursal[sucursal_nombre_dest].append({
                "sucursal": sucursal_nombre_dest,
                "descripcion": t.get("descripcion") or "",
                "fecha_creacion": f_crea_dt,
                "fecha_compromiso": f_sol_dt,
                "comentario": comentario1 or "",
            })

        fill_sucursal = PatternFill("solid", fgColor="FFC000")
        bold_font = Font(bold=True)

        row_ind2 = 2

        for sucursal in sorted(tickets_por_sucursal.keys()):
            items = tickets_por_sucursal[sucursal]
            total_en_progreso = len(items)

            # Fila resumen por sucursal
            ws_ind2.append([sucursal, "", "", "", "", total_en_progreso])
            for cell in ws_ind2[row_ind2]:
                cell.fill = fill_sucursal
                cell.font = bold_font
                cell.alignment = Alignment(vertical="top")
            row_ind2 += 1

            # Filas detalle
            for item in items:
                ws_ind2.append([
                    item["sucursal"],
                    item["descripcion"],
                    item["fecha_creacion"],
                    item["fecha_compromiso"],
                    item["comentario"],
                    "",
                ])

                if item["fecha_creacion"]:
                    ws_ind2.cell(row=row_ind2, column=3).number_format = "dd/mm/yyyy"
                if item["fecha_compromiso"]:
                    ws_ind2.cell(row=row_ind2, column=4).number_format = "dd/mm/yyyy"

                for col in (1, 2, 5):
                    ws_ind2.cell(row=row_ind2, column=col).alignment = Alignment(wrap_text=True, vertical="top")

                row_ind2 += 1

        # Anchos de columna
        ws_ind2.column_dimensions["A"].width = 24
        ws_ind2.column_dimensions["B"].width = 42
        ws_ind2.column_dimensions["C"].width = 18
        ws_ind2.column_dimensions["D"].width = 18
        ws_ind2.column_dimensions["E"].width = 55
        ws_ind2.column_dimensions["F"].width = 18

        ws_ind2.freeze_panes = "A2"
        ws_ind2.auto_filter.ref = ws_ind2.dimensions


        # ─────────────────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"tickets_exportados_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"mensaje": str(e)}), 500


# ─────────────────────────────────────────────────────────────
# RUTA: migrar tickets a ISO local
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/migrar-historial-local', methods=['POST'])
@jwt_required()
@bloquea_lectores_globales
def migrar_historial_local():


    tz_mx = pytz.timezone("America/Tijuana")
    tickets = Ticket.query.all()
    total_actualizados = 0

    for ticket in tickets:
        historial = ticket.historial_fechas
        if not historial:
            continue

        actualizado = False
        nuevo_historial = []

        for entrada in historial:
            nueva_entrada = entrada.copy()
            for campo in ['fecha', 'fechaCambio']:
                valor = entrada.get(campo)
                if valor and isinstance(valor, str) and '/' in valor and len(valor) == 8:
                    try:
                        fecha_local = datetime.strptime(valor, "%d/%m/%y")
                        fecha_local = tz_mx.localize(datetime.combine(fecha_local.date(), time(hour=7)))
                        fecha_utc = fecha_local.astimezone(timezone.utc)
                        nueva_entrada[campo] = fecha_utc.isoformat()
                        actualizado = True
                    except Exception as e:
                        print(f"❌ Error en ticket #{ticket.id}, campo {campo}: {e}")
            nuevo_historial.append(nueva_entrada)

        if actualizado:
            ticket.historial_fechas = nuevo_historial
            total_actualizados += 1

    if total_actualizados > 0:
        db.session.commit()
        return jsonify({"mensaje": f"✅ Historial actualizado en {total_actualizados} tickets."}), 200
    else:
        return jsonify({"mensaje": "⚠️ No se encontraron entradas para actualizar."}), 200



# ─────────────────────────────────────────────────────────────
# RUTA: eliminar todos tickets
# ─────────────────────────────────────────────────────────────

# Ruta temporal y protegida para borrar todos los tickets


@ticket_bp.route('/eliminar-todos', methods=['DELETE'])
@jwt_required()
def eliminar_todos_los_tickets():
    try:
        cantidad = Ticket.query.delete()
        db.session.commit()
        return jsonify({"mensaje": f"🧨 Se eliminaron {cantidad} tickets."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─────────────────────────────────────────────────────────────
# RUTA: Obtener historial de tickets y reparaciones para un equipo
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/historial-equipo/<int:equipo_id>', methods=['GET'])
@jwt_required()
def historial_equipo(equipo_id):
    """
    Retorna el historial de tickets relacionados a un equipo/aparato (por aparato_id).
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # Si no eres admin, filtra solo tickets de tu sucursal
        query = Ticket.query.filter(Ticket.aparato_id == equipo_id)
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter(Ticket.sucursal_id == user.sucursal_id)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()

        data = []
        for t in tickets:
            data.append({
                "ticket_id": t.id,
                "descripcion": t.descripcion,
                "estado": t.estado,
                "fecha_creacion": t.fecha_creacion.isoformat() if t.fecha_creacion else None,
                "fecha_solucion": t.fecha_solucion.isoformat() if t.fecha_solucion else None,
                "problema_detectado": t.problema_detectado,
                "necesita_refaccion": t.necesita_refaccion,
                "descripcion_refaccion": t.descripcion_refaccion,
                "historial_fechas": t.historial_fechas or [],
                "username": t.username,
                "asignado_a": t.asignado_a,
                "url_evidencia": t.url_evidencia,
                "categoria": t.categoria,
                "subcategoria": t.subcategoria,
                "detalle": t.detalle,
            })
        return jsonify(data), 200
    except Exception as e:
        return manejar_error(e, "historial_equipo")


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener historial completo de tickets por aparato_id
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/historial-por-equipo/<int:aparato_id>', methods=['GET'])
@jwt_required()
def historial_por_equipo(aparato_id):
    """
    Devuelve todos los tickets ligados a un aparato/sistema específico (por aparato_id).
    Si el usuario NO es admin, solo ve los tickets de su sucursal.
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        query = Ticket.query.filter_by(aparato_id=aparato_id)

        # Solo admins pueden ver todos, los demás solo los de su sucursal
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter_by(sucursal_id=user.sucursal_id)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
        return jsonify([t.to_dict() for t in tickets]), 200

    except Exception as e:
        return manejar_error(e, "historial_por_equipo")

# ─────────────────────────────────────────────────────────────
# RUTA: Obtener historial global de tickets con filtros
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/historial-global', methods=['GET'])
@jwt_required()
def historial_global():
    """
    Devuelve el historial completo de tickets de todos los aparatos/sistemas.
    Permite filtrar por tipo (aparato/sistema), sucursal, estado, fecha, etc.
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # Filtros opcionales
        tipo = request.args.get('tipo')  # 'aparato' o 'sistema'
        sucursal_id = request.args.get('sucursal_id', type=int)
        estado = request.args.get('estado')  # abierto/en progreso/finalizado

        query = Ticket.query

        # Filtro por tipo de inventario
        if tipo:
            # Relación con InventarioGeneral (asume que tienes el campo tipo en inventario)
            query = query.join(Ticket.inventario)
            query = query.filter(InventarioGeneral.tipo == tipo)

        # Filtro por sucursal (admins ven todo)
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter_by(sucursal_id=user.sucursal_id)
        elif sucursal_id:
            query = query.filter_by(sucursal_id=sucursal_id)

        # Filtro por estado
        if estado:
            query = query.filter_by(estado=estado)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
        return jsonify([t.to_dict() for t in tickets]), 200

    except Exception as e:
        return manejar_error(e, "historial_global")


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener historial completo de tickets y movimientos por aparato_id
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/historial/<int:aparato_id>', methods=['GET'])
@jwt_required()
def historial_aparato(aparato_id):
    """
    Devuelve el historial completo de un aparato/sistema por su ID:
    - Tickets relacionados
    - Movimientos de inventario relacionados
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # FILTRO: Si no es admin, solo puede ver su sucursal
        filtro_sucursal = []
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            filtro_sucursal = [user.sucursal_id]

        # --- Tickets relacionados ---
        tickets_query = Ticket.query.filter_by(aparato_id=aparato_id)
        if filtro_sucursal:
            tickets_query = tickets_query.filter(Ticket.sucursal_id.in_(filtro_sucursal))
        tickets = tickets_query.order_by(Ticket.fecha_creacion.desc()).all()
        tickets_serializados = [t.to_dict() for t in tickets]

        # --- Movimientos de inventario relacionados ---
        # Un movimiento puede traer varios detalles (productos), pero filtramos por inventario_id == aparato_id
        from app.models.inventario import DetalleMovimiento, MovimientoInventario

        detalles = DetalleMovimiento.query.filter_by(inventario_id=aparato_id).all()
        movimientos = []
        for d in detalles:
            mov = d.movimiento
            # Filtramos por sucursal si aplica
            if filtro_sucursal and mov.sucursal_id not in filtro_sucursal:
                continue
            movimientos.append({
                'id': mov.id,
                'tipo': mov.tipo_movimiento,
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'usuario_id': mov.usuario_id,
                'sucursal_id': mov.sucursal_id,
                'cantidad': d.cantidad,
                'unidad_medida': d.unidad_medida,
                'observaciones': mov.observaciones,
            })

        # --- Respuesta combinada ---
        return jsonify({
            'tickets': tickets_serializados,
            'movimientos': movimientos,
        }), 200

    except Exception as e:
        return manejar_error(e, "historial_aparato")


# ─────────────────────────────────────────────────────────────
# RUTA: Notificar resumen de ticket por email/WhatsApp
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/notify/<int:ticket_id>', methods=['POST'])
@jwt_required()
def notify_ticket(ticket_id):
    """
    Body JSON de ejemplo:
    {
      "emails": ["destino@dominio.com"],
      "channels": ["email"]
    }
    """
    user = UserORM.get_by_id(get_jwt_identity())
    if not user:
        return jsonify({"mensaje":"Usuario no encontrado"}), 404

    data = request.get_json() or {}
    emails  = data.get("emails") or []
    phones  = data.get("phones") or []
    channels = set((data.get("channels") or []))

    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404
    td = t.to_dict()  # usa tu to_dict que ya trae inventario, historial, etc.

    results = {}

    if "email" in channels and emails:
        html = render_ticket_html(td)
        send_email_html(emails, f"Ticket #{t.id} – Resumen", html)
        results["email"] = {"ok": True, "to": emails}


    return jsonify({"mensaje":"Notificaciones enviadas", "results": results}), 200


# ─────────────────────────────────────────────────────────────
# RUTA: Marcar ticket para aprobación RRHH
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/rrhh/solicitar/<int:ticket_id>', methods=['POST'])
@jwt_required()
def rrhh_solicitar(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    # Permite: admin/corporativo o jefe del depto del ticket
    if not (_es_admin_o_corporativo(user) or _es_jefe_depto(user, t)):
        return jsonify({"mensaje":"No autorizado"}), 403

    aprobador_username = (request.json or {}).get("aprobador_username")
    t.marcar_para_aprobacion_rrhh(aprobador_username=aprobador_username)
    return jsonify({"mensaje":"Ticket marcado para aprobación RRHH"}), 200


# ─────────────────────────────────────────────────────────────
# RUTA: Aprobar ticket por RRHH
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/rrhh/aprobar/<int:ticket_id>', methods=['POST'])
@jwt_required()
def rrhh_aprobar(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    # Solo gerente general/regional (trátalo como corporativo/admin) o el username designado como aprobador
    comentario = (request.json or {}).get("comentario")
    if not (_es_admin_o_corporativo(user) or _es_gerente_regional(user) or user.username == t.aprobador_username):
        return jsonify({"mensaje":"No autorizado"}), 403


    t.aprobar_rrhh(user.username, comentario)
    return jsonify({"mensaje":"Aprobado por RRHH/gerencia"}), 200



# ─────────────────────────────────────────────────────────────
# RUTA: Rechazar ticket por RRHH
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/rrhh/rechazar/<int:ticket_id>', methods=['POST'])
@jwt_required()
def rrhh_rechazar(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    comentario = (request.json or {}).get("comentario")
    if not (_es_admin_o_corporativo(user) or _es_gerente_regional(user) or user.username == t.aprobador_username):
        return jsonify({"mensaje":"No autorizado"}), 403


    t.rechazar_rrhh(user.username, comentario)
    return jsonify({"mensaje":"Rechazado por RRHH/gerencia"}), 200



# ─────────────────────────────────────────────────────────────
# RUTA: Establecer compromiso de solución y refacción
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/compromiso/<int:ticket_id>', methods=['PUT'])
@jwt_required()
def set_compromiso(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje": "Ticket no encontrado"}), 404

    if not (_es_admin_o_corporativo(user) or _es_jefe_depto(user, t)):
        return jsonify({"mensaje": "No autorizado"}), 403

    data = request.get_json() or {}
    fecha_solucion = data.get("fecha_solucion")       # ISO string
    necesita_ref   = data.get("necesita_refaccion")   # bool
    desc_ref       = data.get("descripcion_refaccion")

    # ── Validación de combinación permitida para refacción definida por Jefe
    #    Ahora dejamos que cualquier ticket de Mantenimiento o Sistemas
    #    pueda definir refacción desde este endpoint.
    def _norm(s): 
        return (s or "").strip().lower()

    dep_nom = _norm(t.departamento.nombre if t.departamento else "")
    es_mantenimiento = dep_nom == "mantenimiento"
    es_sistemas      = dep_nom == "sistemas"

    # ✅ Cualquier ticket de estos departamentos puede tocar refacción
    combo_jefe_ok = es_mantenimiento or es_sistemas

    # Si otro departamento intenta mandar campos de refacción, lo bloqueamos
    if (necesita_ref is not None or desc_ref is not None) and not combo_jefe_ok:
        return jsonify({
            "mensaje": "Refacción solo se define aquí para tickets de Mantenimiento o Sistemas."
        }), 400



    try:
        # Compromiso de fecha (no cierra el ticket)
        if fecha_solucion:
            try:
                # Usa dateutil para robustez con TZ
                dt = parser.isoparse(fecha_solucion)
                if getattr(dt, "tzinfo", None) is None:
                    # Si viene naive, asúmelo como UTC
                    dt = dt.replace(tzinfo=timezone.utc)
                t.fecha_solucion = dt.astimezone(timezone.utc)
            except Exception as e:
                return jsonify({"mensaje": f"Fecha de solución inválida: {e}"}), 400

        # Campos de refacción (solo si combo_jefe_ok y los mandaron)
        if isinstance(necesita_ref, bool):
            t.necesita_refaccion = necesita_ref
            t.refaccion_definida_por_jefe = True
        if desc_ref is not None:
            t.descripcion_refaccion = (desc_ref or None)

        db.session.commit()
        return jsonify({"mensaje": "Compromiso/Refacción actualizado"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"mensaje": str(e)}), 400



# ─────────────────────────────────────────────────────────────
# RUTA: Aprobar cierre de ticket (jefe de depto)
# ─────────────────────────────────────────────────────────────



@ticket_bp.route('/cierre/solicitar/<int:ticket_id>', methods=['POST'])
@jwt_required()
def cierre_solicitar(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    # Puede iniciar el proceso el jefe o admin; (si quieres, también el asignado)
    if not (_es_admin_o_corporativo(user) or _es_jefe_depto(user, t)):
        return jsonify({"mensaje":"No autorizado"}), 403

    data = request.get_json() or {}
    costo = data.get("costo_solucion")
    notas = data.get("notas_cierre")

    # Validar y asignar costo_solucion
    if costo is not None:
        try:
            # lo convertimos a Decimal usando str para evitar issues con float
            t.costo_solucion = Decimal(str(costo))
        except (InvalidOperation, ValueError):
            return jsonify({"mensaje": "costo_solucion inválido"}), 400
    else:
        t.costo_solucion = None

    # Asignar notas de cierre (texto libre)
    t.notas_cierre = (notas or None)

    # Estado de cierre: queda pendiente aprobación del creador
    t.solicitar_cierre()

    return jsonify({
        "mensaje": "Cierre solicitado (pendiente aprobación del creador)",
        "costo_solucion": float(t.costo_solucion) if t.costo_solucion is not None else None,
        "notas_cierre": t.notas_cierre
    }), 200



# ─────────────────────────────────────────────────────────────
# RUTA: Rechazar cierre de ticket (jefe de depto)
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/cierre/rechazar-jefe/<int:ticket_id>', methods=['POST'])
@jwt_required()
def cierre_rechazar_jefe(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    if not (_es_admin_o_corporativo(user) or _es_jefe_depto(user, t)):
        return jsonify({"mensaje":"No autorizado"}), 403

    data = request.get_json() or {}
    motivo = data.get("motivo")
    nueva_compromiso = data.get("nueva_fecha_solucion")  # ISO
    dt_new = None
    if nueva_compromiso:
        dt_new = datetime.fromisoformat(nueva_compromiso).astimezone(timezone.utc)

    t.rechazar_cierre_jefe(motivo=motivo, nueva_fecha_compromiso=dt_new)
    return jsonify({"mensaje":"Rechazado por jefe; ticket reabierto"}), 200


# ─────────────────────────────────────────────────────────────
# RUTA: Aceptar conformidad y cerrar ticket (creador)
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/cierre/aceptar-creador/<int:ticket_id>', methods=['POST'])
@jwt_required()
def cierre_aceptar_creador(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    if not _es_creador(user, t):
        return jsonify({"mensaje":"Solo el creador puede aceptar la conformidad"}), 403

    t.aceptar_conformidad_creador()
    return jsonify({"mensaje":"Conformidad aceptada; ticket finalizado"}), 200


# ─────────────────────────────────────────────────────────────
# RUTA: Rechazar conformidad y reabrir ticket (creador)
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/cierre/rechazar-creador/<int:ticket_id>', methods=['POST'])
@jwt_required()
def cierre_rechazar_creador(ticket_id):
    user = UserORM.get_by_id(get_jwt_identity())
    t = Ticket.query.get(ticket_id)
    if not t:
        return jsonify({"mensaje":"Ticket no encontrado"}), 404

    if not _es_creador(user, t):
        return jsonify({"mensaje":"Solo el creador puede rechazar la conformidad"}), 403

    data = request.get_json() or {}
    motivo = data.get("motivo")
    nueva_compromiso = data.get("nueva_fecha_solucion")  # ISO
    dt_new = None
    if nueva_compromiso:
        dt_new = datetime.fromisoformat(nueva_compromiso).astimezone(timezone.utc)

    t.rechazar_conformidad_creador(motivo=motivo, nueva_fecha_compromiso=dt_new)
    return jsonify({"mensaje":"Conformidad rechazada; ticket reabierto"}), 200





