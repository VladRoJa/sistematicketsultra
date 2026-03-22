# backend/app/warehouse/services/reporte_direccion_parser.py

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
import re
import unicodedata

from flask import current_app
from openpyxl import load_workbook


REPORTE_DIRECCION_REPORT_TYPE_KEY = "reporte_direccion"
EXPECTED_COLUMN_COUNT = 18

SPANISH_MONTHS = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "set": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}

SPECIAL_NULL_TOKENS = {
    "",
    "-",
    "--",
    "n/a",
    "na",
    "sin datos",
    "sin dato",
    "null",
    "none",
}

SUMMARY_PREFIXES = {"t:", "p:"}

EXPECTED_HEADERS_BY_POSITION: dict[int, set[str]] = {
    0: {"sucursal"},
    1: {"socios activos totales"},
    2: {"socios activos kpi"},
    3: {"socios kpi/m2", "socios kpi m2"},
    4: {"asistencia hoy"},
    5: {"diarios de hoy"},
    6: {"gympass"},
    7: {"totalpass"},
    8: {"pases cortesia", "pases cortesía"},
    9: {"ingreso hoy"},
    10: {"semana en curso"},
    11: {"mes en curso"},
    12: {"mes en curso"},
    13: {"mes en curso"},
    14: {"producto"},
    15: {"mismo mes año anterior", "mismo mes ano anterior"},
    # 16 y 17 se validan como fecha de reporte
}


class InvalidReporteDireccionLayout(RuntimeError):
    """El archivo no cumple el contrato mínimo esperado."""


class BusinessDateNotFound(InvalidReporteDireccionLayout):
    """No fue posible extraer business_date del header."""


class HeaderMismatch(InvalidReporteDireccionLayout):
    """El header no coincide con el layout esperado."""


class NoDetailRowsFound(InvalidReporteDireccionLayout):
    """No se encontraron filas detalle válidas."""


@dataclass(slots=True)
class ParsedRowIssue:
    level: str
    row_index: int
    column_name: str | None
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "level": self.level,
            "row_index": self.row_index,
            "column_name": self.column_name,
            "message": self.message,
        }


def register_reporte_direccion_parser(app) -> None:
    """
    Registra este parser como hook runtime.

    Uso esperado más adelante en init/app factory:
        register_reporte_direccion_parser(app)

    Esto deja resuelto:
        app.config["WAREHOUSE_REPORTE_DIRECCION_PARSER"] = parse_reporte_direccion_snapshot
    """
    app.config["WAREHOUSE_REPORTE_DIRECCION_PARSER"] = parse_reporte_direccion_snapshot


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _ensure_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc)

    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value)
            if parsed.tzinfo is None:
                return parsed.replace(tzinfo=timezone.utc)
            return parsed
        except ValueError:
            pass

    return _utc_now()


def _remove_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c)
    )


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.isoformat()
        return value.isoformat()

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_header(value: Any) -> str:
    text = _normalize_text(value).lower()
    text = _remove_accents(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _is_blank_row(row: list[Any]) -> bool:
    return all(_normalize_text(cell) == "" for cell in row)


def _truncate_or_pad_row(row: list[Any], target_len: int = EXPECTED_COLUMN_COUNT) -> list[Any]:
    if len(row) >= target_len:
        return row[:target_len]
    return row + [None] * (target_len - len(row))


def _parse_business_date_from_header_cell(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = _normalize_text(value)
    if not text:
        raise BusinessDateNotFound("La celda del header de fecha viene vacía.")

    lowered = _remove_accents(text.lower()).strip()

    # Intento ISO / formatos estándar reconocibles por date.fromisoformat no aplica a "21 mar 2026"
    iso_candidate = lowered.replace("/", "-")
    try:
        return date.fromisoformat(iso_candidate)
    except ValueError:
        pass

    # Formato tipo "21 mar 2026"
    match = re.match(r"^(\d{1,2})\s+([a-z]{3,})\.?\s+(\d{4})$", lowered)
    if not match:
        raise BusinessDateNotFound(
            f"No se pudo interpretar la fecha del header: {text!r}"
        )

    day = int(match.group(1))
    month_token = match.group(2)[:3]
    year = int(match.group(3))

    month = SPANISH_MONTHS.get(month_token)
    if month is None:
        raise BusinessDateNotFound(
            f"No se reconoció el mes en la fecha del header: {text!r}"
        )

    return date(year, month, day)


def _validate_headers(header_row: list[Any]) -> date:
    if len(header_row) < EXPECTED_COLUMN_COUNT:
        raise HeaderMismatch(
            f"El header tiene {len(header_row)} columnas; se esperaban al menos {EXPECTED_COLUMN_COUNT}."
        )

    normalized_headers = [_normalize_header(cell) for cell in header_row[:EXPECTED_COLUMN_COUNT]]

    for index, expected_variants in EXPECTED_HEADERS_BY_POSITION.items():
        actual = normalized_headers[index]
        if actual not in expected_variants:
            raise HeaderMismatch(
                f"Header inesperado en columna {index + 1}: {header_row[index]!r}. "
                f"Esperado: {sorted(expected_variants)}"
            )

    business_date_left = _parse_business_date_from_header_cell(header_row[16])
    business_date_right = _parse_business_date_from_header_cell(header_row[17])

    if business_date_left != business_date_right:
        raise BusinessDateNotFound(
            "Las columnas de fecha de Hora Apertura y Hora Clausura no coinciden."
        )

    return business_date_left


def _is_summary_row(first_cell_text: str) -> bool:
    normalized = _remove_accents(first_cell_text.lower()).strip()
    return any(normalized.startswith(prefix) for prefix in SUMMARY_PREFIXES)


def _clean_numeric_text(value: Any) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None

    lowered = _remove_accents(text.lower()).strip()
    if lowered in SPECIAL_NULL_TOKENS:
        return None

    return text.strip()


def _parse_int(value: Any) -> int | None:
    cleaned = _clean_numeric_text(value)
    if cleaned is None:
        return None

    normalized = cleaned.replace(",", "")
    normalized = normalized.replace("$", "")
    normalized = normalized.replace("%", "")
    normalized = normalized.strip()

    if normalized == "":
        return None

    try:
        return int(Decimal(normalized))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"No se pudo convertir a int: {value!r}") from exc


def _parse_float(value: Any) -> float | None:
    cleaned = _clean_numeric_text(value)
    if cleaned is None:
        return None

    normalized = cleaned.replace(",", "")
    normalized = normalized.replace("$", "")
    normalized = normalized.replace("%", "")
    normalized = normalized.strip()

    if normalized == "":
        return None

    try:
        return float(Decimal(normalized))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"No se pudo convertir a float: {value!r}") from exc


def _parse_raw_text(value: Any) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None
    return text


def _parse_detail_row(row: list[Any]) -> dict[str, Any]:
    return {
        "sucursal": _normalize_text(row[0]).strip(),
        "socios_activos_totales": _parse_int(row[1]),
        "socios_activos_kpi": _parse_int(row[2]),
        "socios_kpi_m2": _parse_float(row[3]),
        "asistencia_hoy": _parse_int(row[4]),
        "diarios_hoy": _parse_int(row[5]),
        "gympass": _parse_int(row[6]),
        "totalpass": _parse_int(row[7]),
        "pases_cortesia": _parse_int(row[8]),
        "ingreso_hoy": _parse_float(row[9]),
        "ingreso_acumulado_semana_en_curso": _parse_float(row[10]),
        "ingreso_acumulado_mes_en_curso": _parse_float(row[11]),
        "membresia_domiciliada_mes_en_curso": _parse_float(row[12]),
        "pago_posterior_domiciliado_mes_en_curso": _parse_float(row[13]),
        "producto_pct_venta": _parse_float(row[14]),
        "ingreso_acumulado_mismo_mes_anio_anterior": _parse_float(row[15]),
        "hora_apertura_raw": _parse_raw_text(row[16]),
        "hora_clausura_raw": _parse_raw_text(row[17]),
    }


def _load_rows_from_workbook(file_path: str) -> list[list[Any]]:
    try:
        workbook = load_workbook(
            filename=file_path,
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise InvalidReporteDireccionLayout(
            f"No se pudo abrir el archivo como workbook xlsx: {file_path}"
        ) from exc

    if not workbook.sheetnames:
        raise InvalidReporteDireccionLayout("El workbook no contiene hojas.")

    worksheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()

    if not rows:
        raise InvalidReporteDireccionLayout("La hoja viene vacía.")

    return rows


def parse_reporte_direccion_snapshot(
    *,
    warehouse_upload_id: int,
    file_path: str,
    original_filename: str,
    captured_at: datetime | str | None = None,
    report_type_key: str | None = None,
    content_type: str | None = None,
    storage_path: str | None = None,
) -> dict[str, Any]:
    """
    Parser puro de reporte_direccion.

    Responsabilidades:
    - abrir workbook
    - validar layout esperado
    - extraer business_date desde el header
    - parsear filas detalle
    - devolver estructura normalizada

    No hace:
    - inserts a BD
    - canonicalidad
    - lógica de cierre de mes
    """
    if report_type_key and report_type_key != REPORTE_DIRECCION_REPORT_TYPE_KEY:
        raise InvalidReporteDireccionLayout(
            f"Este parser solo soporta {REPORTE_DIRECCION_REPORT_TYPE_KEY!r}."
        )

    if not isinstance(warehouse_upload_id, int) or warehouse_upload_id <= 0:
        raise InvalidReporteDireccionLayout(
            "'warehouse_upload_id' debe ser entero positivo."
        )

    effective_path = storage_path or file_path
    if not effective_path:
        raise InvalidReporteDireccionLayout("Se requiere 'file_path' o 'storage_path'.")

    path = Path(effective_path)
    if not path.exists():
        raise InvalidReporteDireccionLayout(
            f"El archivo a parsear no existe: {path}"
        )

    rows = _load_rows_from_workbook(str(path))
    header_row = _truncate_or_pad_row(list(rows[0]))
    business_date = _validate_headers(header_row)

    parsed_rows: list[dict[str, Any]] = []
    issues: list[ParsedRowIssue] = []
    row_count_detected = 0
    row_count_valid = 0
    row_count_rejected = 0

    # Excel: fila humana = índice + 1
    for idx, raw_row in enumerate(rows[1:], start=2):
        row = _truncate_or_pad_row(list(raw_row))

        if _is_blank_row(row):
            continue

        first_cell_text = _normalize_text(row[0]).strip()

        if not first_cell_text:
            # Si no trae sucursal, pero parece resumen en otra celda, se ignora.
            non_empty_cells = [_normalize_text(cell) for cell in row if _normalize_text(cell)]
            if non_empty_cells:
                first_non_empty = non_empty_cells[0]
                if _is_summary_row(first_non_empty):
                    continue

            issues.append(
                ParsedRowIssue(
                    level="warning",
                    row_index=idx,
                    column_name="sucursal",
                    message="Fila ignorada sin sucursal válida.",
                )
            )
            continue

        if _is_summary_row(first_cell_text):
            continue

        row_count_detected += 1

        try:
            parsed_row = _parse_detail_row(row)
            parsed_rows.append(parsed_row)
            row_count_valid += 1
        except Exception as exc:
            row_count_rejected += 1
            issues.append(
                ParsedRowIssue(
                    level="error",
                    row_index=idx,
                    column_name=None,
                    message=f"Fila rechazada por error de parseo: {exc}",
                )
            )

    if row_count_valid <= 0:
        raise NoDetailRowsFound(
            "No se encontraron filas detalle válidas en reporte_direccion."
        )

    if row_count_valid + row_count_rejected > row_count_detected:
        raise InvalidReporteDireccionLayout(
            "Inconsistencia interna del parser: valid + rejected > detected."
        )

    effective_captured_at = _ensure_datetime(captured_at)

    current_app.logger.info(
        "Reporte_direccion parsed successfully: warehouse_upload_id=%s business_date=%s valid=%s rejected=%s file=%s content_type=%s",
        warehouse_upload_id,
        business_date.isoformat(),
        row_count_valid,
        row_count_rejected,
        original_filename,
        content_type,
    )

    return {
        "warehouse_upload_id": warehouse_upload_id,
        "report_type_key": REPORTE_DIRECCION_REPORT_TYPE_KEY,
        "business_date": business_date.isoformat(),
        "captured_at": effective_captured_at.isoformat(),
        "row_count_detected": row_count_detected,
        "row_count_valid": row_count_valid,
        "row_count_rejected": row_count_rejected,
        "rows": parsed_rows,
        "issues": [issue.to_dict() for issue in issues],
        "metadata": {
            "parser_name": "parse_reporte_direccion_snapshot",
            "original_filename": original_filename,
            "storage_path": str(path),
            "sheet_index": 0,
        },
    }