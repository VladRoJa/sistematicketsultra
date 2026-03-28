# backend/app/warehouse/services/warehouse_upload_creator_existing_service.py

from __future__ import annotations

from datetime import date, datetime, timezone
import importlib
import inspect
from pathlib import Path
from typing import Any, Callable
import re
import unicodedata

from flask import current_app
from openpyxl import load_workbook

from app.warehouse.services.warehouse_document_upload_service import (
    create_warehouse_document_upload,
)


DEFAULT_SOURCE_KEY = "gasca"
REPORTE_DIRECCION_REPORT_TYPE_KEY = "reporte_direccion"
KPI_DESEMPENO_REPORT_TYPE_KEY = "kpi_desempeno"

SPANISH_MONTHS = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "set": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}


class WarehouseUploadCreatorExistingServiceError(RuntimeError):
    """Error base del bridge hacia el servicio reusable documental."""


class InternalWarehousePeriodResolutionError(
    WarehouseUploadCreatorExistingServiceError
):
    """No fue posible resolver el periodo documental interno."""


def register_warehouse_upload_creator_existing_service_impl(app) -> None:
    """
    Registra esta implementación como bridge real hacia el servicio reusable
    de upload documental de Warehouse.
    """
    app.config["WAREHOUSE_INTERNAL_UPLOAD_CREATOR_IMPL"] = (
        create_warehouse_upload_via_existing_service
    )


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _ensure_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value)
            if parsed.tzinfo is None:
                return parsed.replace(tzinfo=timezone.utc)
            return parsed
        except ValueError as exc:
            raise WarehouseUploadCreatorExistingServiceError(
                f"No se pudo parsear captured_at desde string ISO: {value!r}"
            ) from exc

    if value is None:
        return _utc_now()

    raise WarehouseUploadCreatorExistingServiceError(
        f"Tipo no soportado para captured_at: {type(value).__name__}"
    )


def _remove_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c)
    )


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\xa0", " ")


def _parse_business_date_from_header_cell(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = _normalize_text(value)
    lowered = _remove_accents(text.lower()).strip()

    if not lowered:
        raise InternalWarehousePeriodResolutionError(
            "La celda de fecha del header viene vacía."
        )

    match = re.match(r"^(\d{1,2})\s+([a-z]{3,})\.?\s+(\d{4})$", lowered)
    if not match:
        raise InternalWarehousePeriodResolutionError(
            f"No se pudo interpretar la fecha del header: {text!r}"
        )

    day = int(match.group(1))
    month_token = match.group(2)[:3]
    year = int(match.group(3))

    month = SPANISH_MONTHS.get(month_token)
    if month is None:
        raise InternalWarehousePeriodResolutionError(
            f"No se reconoció el mes en la fecha del header: {text!r}"
        )

    return date(year, month, day)


def _resolve_reporte_direccion_cutoff_date_from_xlsx(
    *,
    file_path: str | None,
    file_bytes: bytes | None,
) -> date:
    if file_bytes is not None:
        from io import BytesIO

        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
    elif file_path:
        workbook = load_workbook(
            filename=file_path,
            data_only=True,
            read_only=True,
        )
    else:
        raise InternalWarehousePeriodResolutionError(
            "Se requiere file_path o file_bytes para resolver cutoff_date de reporte_direccion."
        )

    try:
        if not workbook.sheetnames:
            raise InternalWarehousePeriodResolutionError(
                "El workbook no contiene hojas."
            )

        worksheet = workbook[workbook.sheetnames[0]]
        first_row = next(worksheet.iter_rows(values_only=True), None)
        if not first_row:
            raise InternalWarehousePeriodResolutionError(
                "La hoja está vacía."
            )

        header = list(first_row)
        if len(header) < 18:
            raise InternalWarehousePeriodResolutionError(
                "El header no tiene las 18 columnas esperadas para reporte_direccion."
            )

        left_date = _parse_business_date_from_header_cell(header[16])
        right_date = _parse_business_date_from_header_cell(header[17])

        if left_date != right_date:
            raise InternalWarehousePeriodResolutionError(
                "Las fechas de Hora Apertura y Hora Clausura no coinciden."
            )

        return left_date

    finally:
        workbook.close()


def _resolve_internal_system_user_id() -> int:
    value = current_app.config.get("WAREHOUSE_INTERNAL_SYSTEM_USER_ID")
    if not isinstance(value, int) or value <= 0:
        raise WarehouseUploadCreatorExistingServiceError(
            "Debes configurar app.config['WAREHOUSE_INTERNAL_SYSTEM_USER_ID'] "
            "con un user id válido para los uploads internos."
        )
    return value


def _resolve_period_for_internal_upload(
    *,
    report_type_key: str,
    original_filename: str | None,
    file_path: str | None,
    file_bytes: bytes | None,
    captured_at: datetime,
    metadata: dict[str, Any] | None,
) -> dict[str, Any]:
    metadata = metadata or {}

    # Permite override explícito si en el futuro el productor ya manda periodo claro.
    if metadata.get("cutoff_date"):
        return {
            "cutoff_date": metadata["cutoff_date"],
            "date_from": None,
            "date_to": None,
        }

    if metadata.get("date_from") and metadata.get("date_to"):
        return {
            "cutoff_date": None,
            "date_from": metadata["date_from"],
            "date_to": metadata["date_to"],
        }

    if report_type_key == REPORTE_DIRECCION_REPORT_TYPE_KEY:
        cutoff_date = _resolve_reporte_direccion_cutoff_date_from_xlsx(
            file_path=file_path,
            file_bytes=file_bytes,
        )
        return {
            "cutoff_date": cutoff_date.isoformat(),
            "date_from": None,
            "date_to": None,
        }

    if report_type_key == KPI_DESEMPENO_REPORT_TYPE_KEY:
        cutoff_date = _resolve_kpi_desempeno_cutoff_date_from_filename(
            original_filename=original_filename,
            file_path=file_path,
        )
        return {
            "cutoff_date": cutoff_date.isoformat(),
            "date_from": None,
            "date_to": None,
        }

    raise InternalWarehousePeriodResolutionError(
        f"No hay resolución de periodo documental configurada todavía para "
        f"{report_type_key!r} en uploads internos."
    )
def _resolve_kpi_desempeno_cutoff_date_from_filename(
    *,
    original_filename: str | None,
    file_path: str | None,
) -> date:
    candidate_name = None

    if original_filename:
        candidate_name = Path(original_filename).name
    elif file_path:
        candidate_name = Path(file_path).name

    if not candidate_name:
        raise InternalWarehousePeriodResolutionError(
            "Se requiere original_filename o file_path para resolver cutoff_date de kpi_desempeno."
        )

    match = re.match(
        r"^kpi_desempeno_(\d{4})-(\d{2})-(\d{2})_\d{2}-\d{2}\.xlsx$",
        candidate_name,
        re.IGNORECASE,
    )
    if not match:
        raise InternalWarehousePeriodResolutionError(
            f"No se pudo resolver cutoff_date de kpi_desempeno desde el filename: {candidate_name!r}"
        )

    year = int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))

    return date(year, month, day)


def create_warehouse_upload_via_existing_service(
    *,
    report_type_key: str,
    original_filename: str,
    content_type: str,
    file_path: str | None = None,
    file_bytes: bytes | None = None,
    captured_at: datetime | str | None = None,
    source_key: str | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Bridge real para crear uploads documentales internos reutilizando
    create_warehouse_document_upload().

    Cierra el mismo camino que ya usa el route manual, pero para el pipeline interno.
    """
    effective_captured_at = _ensure_datetime(captured_at)
    uploaded_by_user_id = _resolve_internal_system_user_id()

    period_data = _resolve_period_for_internal_upload(
        report_type_key=report_type_key,
        original_filename=original_filename,
        file_path=file_path,
        file_bytes=file_bytes,
        captured_at=effective_captured_at,
        metadata=metadata,
    )

    current_app.logger.info(
        "Warehouse internal upload creator dispatch: report_type_key=%s original_filename=%s cutoff_date=%s",
        report_type_key,
        original_filename,
        period_data.get("cutoff_date"),
    )

    try:
        result = create_warehouse_document_upload(
            report_type_key=report_type_key,
            original_filename=original_filename,
            content_type=content_type,
            file_path=file_path,
            file_bytes=file_bytes,
            uploaded_by_user_id=uploaded_by_user_id,
            cutoff_date=period_data.get("cutoff_date"),
            date_from=period_data.get("date_from"),
            date_to=period_data.get("date_to"),
            audit_details={
                "upload_origin": "internal_pipeline",
                "source_key": source_key or DEFAULT_SOURCE_KEY,
                **(metadata or {}),
            },
        )
    except Exception as exc:
        raise WarehouseUploadCreatorExistingServiceError(
            f"Falló el upload documental interno para {report_type_key!r}."
        ) from exc

    current_app.logger.info(
        "Warehouse internal upload created successfully: report_type_key=%s warehouse_upload_id=%s",
        report_type_key,
        result["warehouse_upload_id"],
    )

    return {
        "warehouse_upload_id": result["warehouse_upload_id"],
        "upload_status": result.get("upload_status", "created"),
        "metadata": {
            "report_type_key": result.get("report_type_key"),
            "cutoff_date": result.get("cutoff_date"),
            "date_from": result.get("date_from"),
            "date_to": result.get("date_to"),
            "duplicate_detected": result.get("duplicate_detected"),
            "duplicate_upload_id": result.get("duplicate_upload_id"),
        },
    }