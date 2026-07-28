from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from collections.abc import Callable, Mapping
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import pytz
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


VENTAS_NUEVOS_SOCIOS_DETALLE_REPORT_TYPE_KEY = (
    "ventas_nuevos_socios_detalle"
)

SHEET_NAME = "Socios"
SOURCE_TIMEZONE_NAME = "America/Tijuana"

EXPECTED_COLUMNS = (
    "IDSocio",
    "Pin",
    "Sucursal",
    "Nombre",
    "ApellidoPaterno",
    "ApellidoMaterno",
    "Lada",
    "Telefono",
    "Domicilio",
    "Genero",
    "FechaNacimiento",
    "Email",
    "FechaCreacion",
    "Inscripcion",
    "TipoMembresia",
    "Tarifa",
    "Total",
    "FechaPago",
    "FechaRenovacion",
    "FechaFirmaContrato",
    "TipoPago",
    "TipoTarjeta",
    "LugarPago",
    "IDFolio",
    "Pase",
    "Anfitrion",
    "TotalPagado",
)

_EMAIL_PATTERN = re.compile(
    r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
)


class VentasNuevosSociosDetalleParserError(RuntimeError):
    """Error base del parser estructurado."""


class VentasNuevosSociosDetalleLayoutError(
    VentasNuevosSociosDetalleParserError
):
    """El XLSX no cumple el layout contractual."""


class VentasNuevosSociosDetalleContentError(
    VentasNuevosSociosDetalleParserError
):
    """El XLSX no contiene filas estructurables."""


class _RejectedSourceRow(ValueError):
    def __init__(
        self,
        reason_code: str,
        reason_message: str,
    ) -> None:
        super().__init__(reason_message)
        self.reason_code = reason_code
        self.reason_message = reason_message


@dataclass(frozen=True, slots=True)
class VentasNuevosSociosDetalleRejectedRow:
    row_number: int
    reason_code: str
    reason_message: str


@dataclass(frozen=True, slots=True)
class VentasNuevosSociosDetalleParsedRow:
    row_index: int
    row_hash: str

    id_socio: str
    pin: str

    sucursal_raw: str
    sucursal_id: int | None

    nombre: str
    apellido_paterno: str
    apellido_materno: str

    lada: str
    telefono: str
    domicilio: str | None

    genero: str | None
    fecha_nacimiento: date | None
    email: str | None

    fecha_creacion_at: datetime

    inscripcion: str | None
    tipo_membresia: str
    tarifa: str
    total: Decimal | None

    fecha_pago_at: datetime
    fecha_renovacion_at: datetime
    fecha_firma_contrato_at: datetime | None

    tipo_pago_code: int
    tipo_tarjeta_code: int | None
    lugar_pago: str

    id_folio: str
    pase: str | None
    anfitrion: str | None
    total_pagado: Decimal | None

    quality_flags: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class VentasNuevosSociosDetalleParseResult:
    report_type_key: str = (
        VENTAS_NUEVOS_SOCIOS_DETALLE_REPORT_TYPE_KEY
    )

    rows: tuple[
        VentasNuevosSociosDetalleParsedRow,
        ...,
    ] = field(default_factory=tuple)

    rejected_rows: tuple[
        VentasNuevosSociosDetalleRejectedRow,
        ...,
    ] = field(default_factory=tuple)

    row_count: int = 0
    row_count_valid: int = 0
    row_count_rejected: int = 0

    header_columns: tuple[str, ...] = field(
        default_factory=tuple
    )

    quality_flag_counts: dict[str, int] = field(
        default_factory=dict
    )

    metadata: dict[str, Any] = field(
        default_factory=dict
    )


def register_ventas_nuevos_socios_detalle_parser(app) -> None:
    app.config[
        "WAREHOUSE_VENTAS_NUEVOS_SOCIOS_DETALLE_PARSER"
    ] = parse_ventas_nuevos_socios_detalle_xlsx


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ")
    return " ".join(text.split())


def _optional_text(value: Any) -> str | None:
    normalized = _normalize_text(value)
    return normalized or None


def _required_text(
    value: Any,
    *,
    field_name: str,
) -> str:
    normalized = _normalize_text(value)

    if not normalized:
        raise _RejectedSourceRow(
            "INVALID_REQUIRED_VALUE",
            f"{field_name} es obligatorio.",
        )

    return normalized


def _positive_integer_text(
    value: Any,
    *,
    field_name: str,
) -> str:
    if isinstance(value, bool):
        raise _RejectedSourceRow(
            "UNSAFE_IDENTIFIER",
            f"{field_name} no puede ser booleano.",
        )

    if isinstance(value, int) and value > 0:
        return str(value)

    if isinstance(value, str):
        normalized = value.strip()

        if normalized.isdigit() and int(normalized) > 0:
            return str(int(normalized))

    raise _RejectedSourceRow(
        "UNSAFE_IDENTIFIER",
        f"{field_name} debe ser un entero positivo seguro.",
    )


def _exact_folio(value: Any) -> str:
    if not isinstance(value, str):
        raise _RejectedSourceRow(
            "UNSAFE_IDENTIFIER",
            "IDFolio debe llegar como texto.",
        )

    normalized = value.strip()

    if not normalized or not normalized.isdigit():
        raise _RejectedSourceRow(
            "UNSAFE_IDENTIFIER",
            "IDFolio debe conservarse como texto de dígitos.",
        )

    return normalized


def _parse_local_datetime(
    value: Any,
    *,
    field_name: str,
    required: bool,
) -> datetime | None:
    if value in (None, ""):
        if required:
            raise _RejectedSourceRow(
                "INVALID_REQUIRED_VALUE",
                f"{field_name} es obligatorio.",
            )
        return None

    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = datetime.strptime(
                value.strip(),
                "%d-%m-%Y %H:%M:%S",
            )
        except ValueError as exc:
            raise _RejectedSourceRow(
                "INVALID_DATETIME",
                (
                    f"{field_name} debe usar "
                    "DD-MM-YYYY HH:MM:SS."
                ),
            ) from exc
    else:
        raise _RejectedSourceRow(
            "INVALID_DATETIME",
            f"{field_name} tiene un tipo no soportado.",
        )

    if parsed.tzinfo is None:
        source_timezone = pytz.timezone(
            SOURCE_TIMEZONE_NAME
        )
        parsed = source_timezone.localize(parsed)

    return parsed.astimezone(timezone.utc)


def _parse_birth_date(
    value: Any,
) -> tuple[date | None, str | None]:
    if value in (None, ""):
        return None, "BIRTH_DATE_MISSING"

    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = datetime.strptime(
                value.strip(),
                "%d-%m-%Y",
            ).date()
        except ValueError as exc:
            raise _RejectedSourceRow(
                "INVALID_DATE",
                "FechaNacimiento debe usar DD-MM-YYYY.",
            ) from exc
    else:
        raise _RejectedSourceRow(
            "INVALID_DATE",
            "FechaNacimiento tiene un tipo no soportado.",
        )

    if parsed == date(9999, 12, 31):
        return None, "BIRTH_DATE_SENTINEL"

    return parsed, None


def _optional_decimal(
    value: Any,
    *,
    field_name: str,
) -> Decimal | None:
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        raise _RejectedSourceRow(
            "INVALID_AMOUNT",
            f"{field_name} no puede ser booleano.",
        )

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    normalized = (
        str(value)
        .strip()
        .replace("$", "")
        .replace(",", "")
        .replace(" ", "")
    )

    if not normalized:
        return None

    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError) as exc:
        raise _RejectedSourceRow(
            "INVALID_AMOUNT",
            f"{field_name} no pudo convertirse a Decimal.",
        ) from exc


def _required_integer_code(
    value: Any,
    *,
    field_name: str,
) -> int:
    if isinstance(value, bool):
        raise _RejectedSourceRow(
            "INVALID_CODE",
            f"{field_name} no puede ser booleano.",
        )

    try:
        parsed = int(value)
    except Exception as exc:
        raise _RejectedSourceRow(
            "INVALID_CODE",
            f"{field_name} debe ser entero.",
        ) from exc

    return parsed


def _optional_integer_code(
    value: Any,
    *,
    field_name: str,
) -> int | None:
    if value in (None, ""):
        return None

    return _required_integer_code(
        value,
        field_name=field_name,
    )


def _json_default(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    raise TypeError(
        f"Tipo no serializable: {type(value).__name__}"
    )


def _canonical_hash(
    payload: Mapping[str, Any],
) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=_json_default,
    ).encode("utf-8")

    return hashlib.sha256(encoded).hexdigest()


def _resolve_branch_id(
    *,
    sucursal_raw: str,
    branch_resolver: Callable[[str], int | None] | None,
) -> tuple[int | None, str | None]:
    if branch_resolver is None:
        return None, "BRANCH_NOT_RESOLVED"

    resolved = branch_resolver(sucursal_raw)

    if resolved is None:
        return None, "BRANCH_UNRESOLVED"

    if (
        not isinstance(resolved, int)
        or isinstance(resolved, bool)
        or resolved <= 0
    ):
        raise _RejectedSourceRow(
            "INVALID_BRANCH_RESOLVER_RESULT",
            (
                "branch_resolver debe devolver "
                "entero positivo o None."
            ),
        )

    return resolved, None


def _parse_source_row(
    row: Mapping[str, Any],
    *,
    row_number: int,
    branch_resolver: Callable[[str], int | None] | None,
) -> VentasNuevosSociosDetalleParsedRow:
    quality_flags: list[str] = []

    id_socio = _positive_integer_text(
        row.get("IDSocio"),
        field_name="IDSocio",
    )

    pin = _positive_integer_text(
        row.get("Pin"),
        field_name="Pin",
    )

    sucursal_raw = _required_text(
        row.get("Sucursal"),
        field_name="Sucursal",
    )

    sucursal_id, branch_flag = _resolve_branch_id(
        sucursal_raw=sucursal_raw,
        branch_resolver=branch_resolver,
    )

    if branch_flag:
        quality_flags.append(branch_flag)

    nombre = _required_text(
        row.get("Nombre"),
        field_name="Nombre",
    )

    apellido_paterno = _required_text(
        row.get("ApellidoPaterno"),
        field_name="ApellidoPaterno",
    )

    apellido_materno = _required_text(
        row.get("ApellidoMaterno"),
        field_name="ApellidoMaterno",
    )

    lada = _required_text(
        row.get("Lada"),
        field_name="Lada",
    )

    telefono = _required_text(
        row.get("Telefono"),
        field_name="Telefono",
    )

    fecha_nacimiento, birth_flag = _parse_birth_date(
        row.get("FechaNacimiento")
    )

    if birth_flag:
        quality_flags.append(birth_flag)

    email = _optional_text(row.get("Email"))

    if email and _EMAIL_PATTERN.fullmatch(email.lower()) is None:
        quality_flags.append("EMAIL_FORMAT_INVALID")

    fecha_creacion_at = _parse_local_datetime(
        row.get("FechaCreacion"),
        field_name="FechaCreacion",
        required=True,
    )

    total = _optional_decimal(
        row.get("Total"),
        field_name="Total",
    )

    if total is None:
        quality_flags.append("TOTAL_MISSING")

    fecha_pago_at = _parse_local_datetime(
        row.get("FechaPago"),
        field_name="FechaPago",
        required=True,
    )

    fecha_renovacion_at = _parse_local_datetime(
        row.get("FechaRenovacion"),
        field_name="FechaRenovacion",
        required=True,
    )

    fecha_firma_contrato_at = _parse_local_datetime(
        row.get("FechaFirmaContrato"),
        field_name="FechaFirmaContrato",
        required=False,
    )

    tipo_pago_code = _required_integer_code(
        row.get("TipoPago"),
        field_name="TipoPago",
    )

    tipo_tarjeta_code = _optional_integer_code(
        row.get("TipoTarjeta"),
        field_name="TipoTarjeta",
    )

    id_folio = _exact_folio(
        row.get("IDFolio")
    )

    total_pagado = _optional_decimal(
        row.get("TotalPagado"),
        field_name="TotalPagado",
    )

    if total_pagado is None:
        quality_flags.append("TOTAL_PAGADO_MISSING")

    source_payload = {
        "id_socio": id_socio,
        "pin": pin,
        "sucursal_raw": sucursal_raw,
        "nombre": nombre,
        "apellido_paterno": apellido_paterno,
        "apellido_materno": apellido_materno,
        "lada": lada,
        "telefono": telefono,
        "domicilio": _optional_text(
            row.get("Domicilio")
        ),
        "genero": _optional_text(
            row.get("Genero")
        ),
        "fecha_nacimiento": fecha_nacimiento,
        "email": email,
        "fecha_creacion_at": fecha_creacion_at,
        "inscripcion": _optional_text(
            row.get("Inscripcion")
        ),
        "tipo_membresia": _required_text(
            row.get("TipoMembresia"),
            field_name="TipoMembresia",
        ),
        "tarifa": _required_text(
            row.get("Tarifa"),
            field_name="Tarifa",
        ),
        "total": total,
        "fecha_pago_at": fecha_pago_at,
        "fecha_renovacion_at": fecha_renovacion_at,
        "fecha_firma_contrato_at": (
            fecha_firma_contrato_at
        ),
        "tipo_pago_code": tipo_pago_code,
        "tipo_tarjeta_code": tipo_tarjeta_code,
        "lugar_pago": _required_text(
            row.get("LugarPago"),
            field_name="LugarPago",
        ),
        "id_folio": id_folio,
        "pase": _optional_text(
            row.get("Pase")
        ),
        "anfitrion": _optional_text(
            row.get("Anfitrion")
        ),
        "total_pagado": total_pagado,
    }

    return VentasNuevosSociosDetalleParsedRow(
        row_index=row_number,
        row_hash=_canonical_hash(source_payload),
        sucursal_id=sucursal_id,
        quality_flags=tuple(
            sorted(set(quality_flags))
        ),
        **source_payload,
    )


def _open_workbook(
    *,
    file_path: str | None,
    file_bytes: bytes | None,
):
    if not file_path and file_bytes is None:
        raise VentasNuevosSociosDetalleParserError(
            "Se requiere file_path o file_bytes."
        )

    source = (
        BytesIO(file_bytes)
        if file_bytes is not None
        else Path(file_path).open("rb")  # type: ignore[arg-type]
    )

    try:
        workbook = load_workbook(
            source,
            data_only=True,
            read_only=True,
        )
    except (
        OSError,
        InvalidFileException,
        BadZipFile,
        ValueError,
    ) as exc:
        source.close()
        raise VentasNuevosSociosDetalleLayoutError(
            "No se pudo abrir el XLSX."
        ) from exc

    return workbook, source


def _read_headers(worksheet: Any) -> tuple[str, ...]:
    headers = tuple(
        _normalize_text(cell.value)
        for cell in worksheet[1]
    )

    if headers != EXPECTED_COLUMNS:
        raise VentasNuevosSociosDetalleLayoutError(
            "El header no coincide con el contrato. "
            f"Esperado={EXPECTED_COLUMNS!r} "
            f"Recibido={headers!r}"
        )

    return headers


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return not any(
        _normalize_text(value)
        for value in values
    )


def parse_ventas_nuevos_socios_detalle_xlsx(
    *,
    file_path: str | None = None,
    file_bytes: bytes | None = None,
    branch_resolver: Callable[[str], int | None] | None = None,
) -> VentasNuevosSociosDetalleParseResult:
    workbook, source = _open_workbook(
        file_path=file_path,
        file_bytes=file_bytes,
    )

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise VentasNuevosSociosDetalleLayoutError(
                "Falta la hoja obligatoria 'Socios'."
            )

        worksheet = workbook[SHEET_NAME]
        headers = _read_headers(worksheet)

        parsed_rows: list[
            VentasNuevosSociosDetalleParsedRow
        ] = []

        rejected_rows: list[
            VentasNuevosSociosDetalleRejectedRow
        ] = []

        seen_member_ids: set[str] = set()
        detected_rows = 0

        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_col=len(headers),
                values_only=True,
            ),
            start=2,
        ):
            values_tuple = tuple(values)

            if _is_empty_row(values_tuple):
                continue

            detected_rows += 1
            row = dict(zip(headers, values_tuple))

            try:
                parsed_row = _parse_source_row(
                    row,
                    row_number=row_number,
                    branch_resolver=branch_resolver,
                )

                if parsed_row.id_socio in seen_member_ids:
                    raise _RejectedSourceRow(
                        "DUPLICATE_ID_SOCIO",
                        (
                            "IDSocio está duplicado dentro "
                            "del mismo archivo."
                        ),
                    )

                seen_member_ids.add(
                    parsed_row.id_socio
                )
                parsed_rows.append(parsed_row)

            except _RejectedSourceRow as exc:
                rejected_rows.append(
                    VentasNuevosSociosDetalleRejectedRow(
                        row_number=row_number,
                        reason_code=exc.reason_code,
                        reason_message=exc.reason_message,
                    )
                )

        if not parsed_rows:
            raise VentasNuevosSociosDetalleContentError(
                "No se encontraron filas válidas."
            )

        quality_flag_counter = Counter(
            flag
            for parsed_row in parsed_rows
            for flag in parsed_row.quality_flags
        )

        return VentasNuevosSociosDetalleParseResult(
            rows=tuple(parsed_rows),
            rejected_rows=tuple(rejected_rows),
            row_count=detected_rows,
            row_count_valid=len(parsed_rows),
            row_count_rejected=len(rejected_rows),
            header_columns=headers,
            quality_flag_counts=dict(
                sorted(quality_flag_counter.items())
            ),
            metadata={
                "sheet_name": SHEET_NAME,
                "source_timezone": SOURCE_TIMEZONE_NAME,
            },
        )
    finally:
        workbook.close()
        source.close()
