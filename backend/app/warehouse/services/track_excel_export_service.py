# backend\app\warehouse\services\track_excel_export_service.py


from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
from io import BytesIO
from typing import Any, Sequence

from openpyxl.formatting.rule import CellIsRule
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


MONTH_NAMES_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}

TRACK_START_ROW = 4
TRACK_FIRST_COL = 2
TRACK_LAST_COL = 55

CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00;$0.00'
INTEGER_FORMAT = '#,##0'
DECIMAL_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.0%'

ULTRA_ORANGE = "E54525"
ULTRA_BLUE = "2C4595"
ULTRA_CHARCOAL = "1F1F1F"
HEADER_DARK = "111827"

GOOD_FILL_COLOR = "D9F99D"
GOOD_FONT_COLOR = "166534"

WARN_FILL_COLOR = "FEF3C7"
WARN_FONT_COLOR = "92400E"

BAD_FILL_COLOR = "FECACA"
BAD_FONT_COLOR = "991B1B"

NEUTRAL_FILL_COLOR = "F3F4F6"
SEGMENT_BORDER_COLOR = "111827"


TRACK_GROUPS = [
    ("D2:J2", "Ocupación"),
    ("K2:Q2", "C r e c i m i e n t o"),
    ("R2:Y2", "I n g r e s o s    t o t a l e s"),
    ("Z2:AI2", "V e n t a s   n u e v a s     y     R e a c t i v a c i o n e s "),
    ("AJ2:AQ2", "B a j a s  & Churn"),
    ("AR2:AT2", "D o m i c i l i a d o s"),
    ("AU2:AY2", "A  R  P  U"),
    ("AZ2:BC2", "Tienda"),
]

TRACK_BRANCH_ORDER = [
    "VILLAS_DEL_REY",
    "VILLA_VERDE",
    "INDEPENDENCIA",
    "TEC_MXL",
    "SEND_MXL",
    "SAN_LUIS",
    "PABELLON_RTO",
    "MISION_ENS",
    "PASEO_2000",
    "LOMA_BONITA",
    "SANTA_FE",
    "CARROUSEL_TJ",
    "PAPALOTE_TJ",
    "SEND_CUL",
    "SAN_ISIDRO_CUL",
    "AZAHARES_CUL",
    "STA_CATARINA",
    "SEND_SALTILLO",
    "SEND_CHIH",
    "PASEO_LA_PAZ",
    "IXTAPALUCA",
    "INSURGENTES",
    "TLALNEPANTLA",
    "SALTILLO_VILLALTA",
    "METEPEC",
    "LA_VIGA",
]

TRACK_BASE_BRANCHES = set(TRACK_BRANCH_ORDER[:21])
TRACK_NEW_BRANCHES = set(TRACK_BRANCH_ORDER[21:])

TRACK_BRANCH_ORDER_INDEX = {
    branch: index
    for index, branch in enumerate(TRACK_BRANCH_ORDER)
}


BASE_HEADERS = {
    "B": "#",
    "C": "Sucursal",
    "D": "M2 sin  circulaciones",
    "E": "Ocupación 1 {month}",
    "F": "Meta Ocupación 2 px m2 ",
    "G": "Meta Ocupación {month}",
    "H": "Ocupación {day_month}",
    "I": "Dif Inicio de mes VS Día Actual",
    "J": "% Alcance meta Ocupación",
    "K": "Usuarios 1 {month}",
    "L": "Proyección Usuarios activos al cierre {month}",
    "M": "Usuarios activos al {day_month}",
    "N": "DIF INICIO MES",
    "O": "Dif meta  vs Real (##)",
    "P": "% Alcance meta crecimiento",
    "Q": "Alcance usuarios activos",
    "R": "Meta FAYCGO {month}",
    "S": " Ingreso ideal al CIERRE {day_month}",
    "T": " Ingreso Real al {day_month}",
    "U": "Ingreso Real al dia anterior {month} agregadora",
    "V": "Ingreso Real total al {day} de {month}",
    "W": "% alcance de meta al {day_month}",
    "X": "Meta del día",
    "Y": "Dif $$ Real vs Meta",
    "Z": "Meta Clientes nuevos",
    "AA": "Alcance ideal clientes nuevos {day_month}",
    "AB": "Real Clientes nuevos {day_month}",
    "AC": "Diferencia Ideal VS  REAL ",
    "AD": "% alcance nuevos {day_month}",
    "AE": "Meta Reactiv",
    "AF": "Reactivaciones IDEALES AL {day_month}",
    "AG": "Real  Reactivaciones {day_month}",
    "AH": "Meta del día REACTIVACIONES ",
    "AI": "%  Alcance",
    "AJ": "Meta bajas",
    "AK": "Bajas Ideales {day_month}",
    "AL": "Bajas reales {day_month}",
    "AM": "Dif Ideal  VS Real",
    "AN": "Meta Churn",
    "AO": "Churn ideal {day_month}",
    "AP": "Churn real {day_month}",
    "AQ": "DIF CHURN IDEAL VS REAL",
    "AR": "Meta nuevos contratos domiciliados",
    "AS": "Nuevos contratos domiciliados {day_month}",
    "AT": "% alcance {day_month}",
    "AU": "Meta ARPU ",
    "AV": "Real ARPU {day_month}",
    "AW": "ARPU TEORICO ",
    "AX": "Dif $$ ARPU",
    "AY": "% Alcance ARPU",
    "AZ": "Meta venta Tienda",
    "BA": "Real {day_month}",
    "BB": "Diferencia $$",
    "BC": "% alcance real",
}


def build_track_daily_mart_excel(
    *,
    track_date: date,
    generation_mode: str,
    resolved_version: Any,
    rows: Sequence[Any],
) -> bytes:
    """Build an .xlsx export that mirrors the original Track 'Metas OP' columns."""

    workbook = Workbook()
    track_sheet = workbook.active
    track_sheet.title = "Metas OP"

    days_in_month = monthrange(track_date.year, track_date.month)[1]

    _setup_track_sheet(
        worksheet=track_sheet,
        track_date=track_date,
        generation_mode=generation_mode,
        days_in_month=days_in_month,
        resolved_version=resolved_version,
    )

    ordered_rows = _sort_rows_by_track_order(rows)

    base_rows = [
        row
        for row in ordered_rows
        if _normalize_branch_key(getattr(row, "sucursal_canon", "")) in TRACK_BASE_BRANCHES
    ]

    new_rows = [
        row
        for row in ordered_rows
        if _normalize_branch_key(getattr(row, "sucursal_canon", "")) not in TRACK_BASE_BRANCHES
    ]

    current_row = TRACK_START_ROW
    current_index = 1

    base_start_row = current_row
    for row in base_rows:
        _write_track_data_row(
            worksheet=track_sheet,
            excel_row=current_row,
            index=current_index,
            mart_row=row,
            days_in_month=days_in_month,
        )
        current_row += 1
        current_index += 1

    base_end_row = current_row - 1

    new_start_row = current_row
    for row in new_rows:
        _write_track_data_row(
            worksheet=track_sheet,
            excel_row=current_row,
            index=current_index,
            mart_row=row,
            days_in_month=days_in_month,
        )
        current_row += 1
        current_index += 1

    new_end_row = current_row - 1

    base_totals_row = current_row
    _write_subtotal_row(
        worksheet=track_sheet,
        totals_row=base_totals_row,
        label=f"Subtotales {len(base_rows)} GYMS",
        first_data_row=base_start_row,
        last_data_row=base_end_row,
    )

    new_totals_row = current_row + 1
    _write_subtotal_row(
        worksheet=track_sheet,
        totals_row=new_totals_row,
        label="Subtotales Nuevos",
        first_data_row=new_start_row,
        last_data_row=new_end_row,
    )

    general_totals_row = current_row + 2
    _write_general_totals_row(
        worksheet=track_sheet,
        totals_row=general_totals_row,
        base_totals_row=base_totals_row,
        new_totals_row=new_totals_row,
    )

    _apply_track_sheet_formatting(
        worksheet=track_sheet,
        last_row=general_totals_row,
    )

    _build_raw_sheet(
        workbook=workbook,
        rows=ordered_rows,
    )
    _build_info_sheet(
        workbook=workbook,
        track_date=track_date,
        generation_mode=generation_mode,
        resolved_version=resolved_version,
        total_rows=len(ordered_rows),
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _setup_track_sheet(
    *,
    worksheet: Worksheet,
    track_date: date,
    generation_mode: str,
    days_in_month: int,
    resolved_version: Any,
) -> None:
    month_label = MONTH_NAMES_ES[track_date.month]
    day_month_label = f"{track_date.day} {month_label}"

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "D4"

    worksheet["B1"] = "Track"
    worksheet["C1"] = "A dia"
    worksheet["D1"] = track_date.isoformat()

    worksheet.merge_cells("E1:J1")
    worksheet["E1"] = _format_version_update_label(resolved_version)

    worksheet["B2"] = f"=C2/{days_in_month}"
    worksheet["C2"] = track_date.day

    for merge_range, title in TRACK_GROUPS:
        worksheet.merge_cells(merge_range)
        first_cell = merge_range.split(":", 1)[0]
        worksheet[first_cell] = title

    for column_letter, header_template in BASE_HEADERS.items():
        worksheet[f"{column_letter}3"] = header_template.format(
            day=track_date.day,
            month=month_label,
            day_month=day_month_label,
        )


def _write_track_data_row(
    *,
    worksheet: Worksheet,
    excel_row: int,
    index: int,
    mart_row: Any,
    days_in_month: int,
) -> None:
    worksheet[f"B{excel_row}"] = index
    worksheet[f"C{excel_row}"] = _format_branch_label(
        getattr(mart_row, "sucursal_canon", "")
    )

    worksheet[f"D{excel_row}"] = _to_number(getattr(mart_row, "m2_sin_circulaciones", None))
    worksheet[f"E{excel_row}"] = f"=IFERROR(K{excel_row}/D{excel_row},0)"
    worksheet[f"F{excel_row}"] = f"=D{excel_row}*2"
    worksheet[f"G{excel_row}"] = f"=IFERROR(L{excel_row}/D{excel_row},0)"
    worksheet[f"H{excel_row}"] = f"=IFERROR(M{excel_row}/D{excel_row},0)"
    worksheet[f"I{excel_row}"] = f"=G{excel_row}-E{excel_row}"
    worksheet[f"J{excel_row}"] = f"=IFERROR(H{excel_row}/G{excel_row},0)"

    worksheet[f"K{excel_row}"] = _to_number(getattr(mart_row, "usuarios_inicio_mes", None))
    worksheet[f"L{excel_row}"] = _to_number(getattr(mart_row, "proyeccion_usuarios_cierre_mes", None))
    worksheet[f"M{excel_row}"] = _to_number(getattr(mart_row, "usuarios_activos_actual", None))
    worksheet[f"N{excel_row}"] = f"=M{excel_row}-K{excel_row}"
    worksheet[f"O{excel_row}"] = f"=M{excel_row}-L{excel_row}"
    worksheet[f"P{excel_row}"] = f"=IFERROR((M{excel_row}/L{excel_row})-1,0)"
    worksheet[f"Q{excel_row}"] = f"=IFERROR((M{excel_row}/K{excel_row})-1,0)"

    worksheet[f"R{excel_row}"] = _to_number(getattr(mart_row, "meta_faycgo_mes", None))
    worksheet[f"S{excel_row}"] = f"=(R{excel_row}/{days_in_month})*$C$2"
    worksheet[f"T{excel_row}"] = _to_number(getattr(mart_row, "ingreso_real_base_mtd", None))
    worksheet[f"U{excel_row}"] = _to_number(getattr(mart_row, "ingreso_real_agregadora_mtd", None))
    worksheet[f"V{excel_row}"] = f"=T{excel_row}+U{excel_row}"
    worksheet[f"W{excel_row}"] = f"=IFERROR(V{excel_row}/R{excel_row},0)"
    worksheet[f"X{excel_row}"] = f"=V{excel_row}-S{excel_row}"
    worksheet[f"Y{excel_row}"] = f"=V{excel_row}-R{excel_row}"

    worksheet[f"Z{excel_row}"] = _to_number(getattr(mart_row, "meta_clientes_nuevos_mes", None))
    worksheet[f"AA{excel_row}"] = f"=(Z{excel_row}/{days_in_month})*$C$2"
    worksheet[f"AB{excel_row}"] = _to_number(getattr(mart_row, "clientes_nuevos_real_mtd", None))
    worksheet[f"AC{excel_row}"] = f"=AB{excel_row}-AA{excel_row}"
    worksheet[f"AD{excel_row}"] = f"=IFERROR(AB{excel_row}/Z{excel_row},0)"

    worksheet[f"AE{excel_row}"] = _to_number(getattr(mart_row, "meta_reactivaciones_mes", None))
    worksheet[f"AF{excel_row}"] = f"=(AE{excel_row}/{days_in_month})*$C$2"
    worksheet[f"AG{excel_row}"] = _to_number(getattr(mart_row, "reactivaciones_real_mtd", None))
    worksheet[f"AH{excel_row}"] = f"=AF{excel_row}-AG{excel_row}"
    worksheet[f"AI{excel_row}"] = f"=IFERROR(AG{excel_row}/AE{excel_row},0)"

    worksheet[f"AJ{excel_row}"] = _to_number(getattr(mart_row, "meta_bajas_mes", None))
    worksheet[f"AK{excel_row}"] = f"=(AJ{excel_row}/{days_in_month})*$C$2"
    worksheet[f"AL{excel_row}"] = _to_number(getattr(mart_row, "bajas_reales_mtd", None))
    worksheet[f"AM{excel_row}"] = f"=AL{excel_row}-AK{excel_row}"
    worksheet[f"AN{excel_row}"] = f"=IFERROR(AJ{excel_row}/L{excel_row},0)"
    worksheet[f"AO{excel_row}"] = f"=(AN{excel_row}/{days_in_month})*$C$2"
    worksheet[f"AP{excel_row}"] = f"=IFERROR(AL{excel_row}/M{excel_row},0)"
    worksheet[f"AQ{excel_row}"] = f"=AP{excel_row}-AO{excel_row}"

    worksheet[f"AR{excel_row}"] = _to_number(getattr(mart_row, "meta_nuevos_domiciliados_mes", None))
    worksheet[f"AS{excel_row}"] = _to_number(getattr(mart_row, "nuevos_domiciliados_real_mtd", None))
    worksheet[f"AT{excel_row}"] = f"=IFERROR(AS{excel_row}/AR{excel_row},0)"

    worksheet[f"AU{excel_row}"] = _to_number(getattr(mart_row, "meta_arpu_mes", None))
    worksheet[f"AV{excel_row}"] = f"=IFERROR(V{excel_row}/M{excel_row},0)"
    worksheet[f"AW{excel_row}"] = f"=IFERROR(S{excel_row}/M{excel_row},0)"
    worksheet[f"AX{excel_row}"] = f"=AV{excel_row}-AU{excel_row}"
    worksheet[f"AY{excel_row}"] = f"=IFERROR(AV{excel_row}/AU{excel_row},0)"

    worksheet[f"AZ{excel_row}"] = _to_number(getattr(mart_row, "meta_venta_tienda_mes", None))
    worksheet[f"BA{excel_row}"] = _to_number(getattr(mart_row, "venta_tienda_real_mtd", None))
    worksheet[f"BB{excel_row}"] = f"=AZ{excel_row}-BA{excel_row}"
    worksheet[f"BC{excel_row}"] = f"=IFERROR(BA{excel_row}/AZ{excel_row},0)"


def _write_subtotal_row(
    *,
    worksheet: Worksheet,
    totals_row: int,
    label: str,
    first_data_row: int,
    last_data_row: int,
) -> None:
    worksheet[f"C{totals_row}"] = label

    if last_data_row < first_data_row:
        for column_letter in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
            worksheet.cell(row=totals_row, column=column_letter).value = None
        worksheet[f"C{totals_row}"] = label
        return

    sum_columns = [
        "D", "K", "L", "M", "N", "O", "R", "S", "T", "U", "V", "X", "Y",
        "Z", "AA", "AB", "AC", "AE", "AF", "AG", "AH", "AJ", "AK", "AL", "AM",
        "AR", "AS", "AZ", "BA", "BB",
    ]

    for column_letter in sum_columns:
        worksheet[f"{column_letter}{totals_row}"] = (
            f"=SUM({column_letter}{first_data_row}:{column_letter}{last_data_row})"
        )

    worksheet[f"E{totals_row}"] = f"=IFERROR(K{totals_row}/D{totals_row},0)"
    worksheet[f"F{totals_row}"] = f"=D{totals_row}*2"
    worksheet[f"G{totals_row}"] = f"=IFERROR(L{totals_row}/D{totals_row},0)"
    worksheet[f"H{totals_row}"] = f"=IFERROR(M{totals_row}/D{totals_row},0)"
    worksheet[f"I{totals_row}"] = f"=AVERAGE(I{first_data_row}:I{last_data_row})"
    worksheet[f"J{totals_row}"] = f"=AVERAGE(J{first_data_row}:J{last_data_row})"
    worksheet[f"P{totals_row}"] = f"=IFERROR((M{totals_row}/L{totals_row})-1,0)"
    worksheet[f"Q{totals_row}"] = f"=IFERROR((M{totals_row}/K{totals_row})-1,0)"
    worksheet[f"W{totals_row}"] = f"=IFERROR(V{totals_row}/R{totals_row},0)"
    worksheet[f"AD{totals_row}"] = f"=IFERROR(AB{totals_row}/Z{totals_row},0)"
    worksheet[f"AI{totals_row}"] = f"=IFERROR(AG{totals_row}/AE{totals_row},0)"
    worksheet[f"AN{totals_row}"] = f"=IFERROR(AJ{totals_row}/L{totals_row},0)"
    worksheet[f"AO{totals_row}"] = f"=AVERAGE(AO{first_data_row}:AO{last_data_row})"
    worksheet[f"AP{totals_row}"] = f"=IFERROR(AL{totals_row}/M{totals_row},0)"
    worksheet[f"AQ{totals_row}"] = f"=AP{totals_row}-AO{totals_row}"
    worksheet[f"AT{totals_row}"] = f"=IFERROR(AS{totals_row}/AR{totals_row},0)"
    worksheet[f"AU{totals_row}"] = f"=AVERAGE(AU{first_data_row}:AU{last_data_row})"
    worksheet[f"AV{totals_row}"] = f"=IFERROR(V{totals_row}/M{totals_row},0)"
    worksheet[f"AW{totals_row}"] = f"=IFERROR(S{totals_row}/M{totals_row},0)"
    worksheet[f"AX{totals_row}"] = f"=AV{totals_row}-AU{totals_row}"
    worksheet[f"AY{totals_row}"] = f"=IFERROR(AV{totals_row}/AU{totals_row},0)"
    worksheet[f"BC{totals_row}"] = f"=IFERROR(BA{totals_row}/AZ{totals_row},0)"


def _write_general_totals_row(
    *,
    worksheet: Worksheet,
    totals_row: int,
    base_totals_row: int,
    new_totals_row: int,
) -> None:
    worksheet[f"C{totals_row}"] = "TOTAL GENERAL"

    sum_columns = [
        "D", "F", "K", "L", "M", "N", "O", "R", "S", "T", "U", "V", "X", "Y",
        "Z", "AA", "AB", "AC", "AE", "AF", "AG", "AH", "AJ", "AK", "AL", "AM",
        "AR", "AS", "AZ", "BA", "BB",
    ]

    for column_letter in sum_columns:
        worksheet[f"{column_letter}{totals_row}"] = (
            f"={column_letter}{base_totals_row}+{column_letter}{new_totals_row}"
        )

    worksheet[f"E{totals_row}"] = f"=IFERROR(K{totals_row}/D{totals_row},0)"
    worksheet[f"G{totals_row}"] = f"=IFERROR(L{totals_row}/D{totals_row},0)"
    worksheet[f"H{totals_row}"] = f"=IFERROR(M{totals_row}/D{totals_row},0)"
    worksheet[f"I{totals_row}"] = f"=AVERAGE(I{base_totals_row}:I{new_totals_row})"
    worksheet[f"J{totals_row}"] = f"=AVERAGE(J{base_totals_row}:J{new_totals_row})"
    worksheet[f"P{totals_row}"] = f"=IFERROR((M{totals_row}/L{totals_row})-1,0)"
    worksheet[f"Q{totals_row}"] = f"=IFERROR((M{totals_row}/K{totals_row})-1,0)"
    worksheet[f"W{totals_row}"] = f"=IFERROR(V{totals_row}/R{totals_row},0)"
    worksheet[f"AD{totals_row}"] = f"=IFERROR(AB{totals_row}/Z{totals_row},0)"
    worksheet[f"AI{totals_row}"] = f"=IFERROR(AG{totals_row}/AE{totals_row},0)"
    worksheet[f"AN{totals_row}"] = f"=IFERROR(AJ{totals_row}/L{totals_row},0)"
    worksheet[f"AO{totals_row}"] = f"=AVERAGE(AO{base_totals_row}:AO{new_totals_row})"
    worksheet[f"AP{totals_row}"] = f"=IFERROR(AL{totals_row}/M{totals_row},0)"
    worksheet[f"AQ{totals_row}"] = f"=AP{totals_row}-AO{totals_row}"
    worksheet[f"AT{totals_row}"] = f"=IFERROR(AS{totals_row}/AR{totals_row},0)"
    worksheet[f"AU{totals_row}"] = f"=AVERAGE(AU{base_totals_row}:AU{new_totals_row})"
    worksheet[f"AV{totals_row}"] = f"=IFERROR(V{totals_row}/M{totals_row},0)"
    worksheet[f"AW{totals_row}"] = f"=IFERROR(S{totals_row}/M{totals_row},0)"
    worksheet[f"AX{totals_row}"] = f"=AV{totals_row}-AU{totals_row}"
    worksheet[f"AY{totals_row}"] = f"=IFERROR(AV{totals_row}/AU{totals_row},0)"
    worksheet[f"BC{totals_row}"] = f"=IFERROR(BA{totals_row}/AZ{totals_row},0)"

def _apply_track_sheet_formatting(*, worksheet: Worksheet, last_row: int) -> None:
    black_fill = PatternFill("solid", fgColor="1F1F1F")
    red_fill = PatternFill("solid", fgColor="E54525")
    subtotal_fill = PatternFill("solid", fgColor="FBE3DC")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="E54525", bold=True, size=12)
    header_font = Font(color="FFFFFF", bold=True, size=9)
    body_font = Font(color="0F1F3A", size=9)
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    worksheet["B1"].font = title_font
    worksheet.column_dimensions["A"].hidden = True
    worksheet["C1"].font = Font(bold=True)
    worksheet["D1"].font = Font(bold=True)
    worksheet["E1"].font = Font(color="E54525", bold=True)
    worksheet["E1"].alignment = Alignment(horizontal="left", vertical="center")

    for merge_range, _title in TRACK_GROUPS:
        first_cell = merge_range.split(":", 1)[0]
        cell = worksheet[first_cell]
        cell.fill = red_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
        cell = worksheet.cell(row=3, column=col_idx)
        cell.fill = black_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx in range(TRACK_START_ROW, last_row + 1):
        for col_idx in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    _apply_zebra_rows(
        worksheet=worksheet,
        last_row=last_row,
    )

    _apply_summary_row_formatting(
        worksheet=worksheet,
        last_row=last_row,
    )

    for row_idx in range(TRACK_START_ROW, last_row + 1):
        worksheet[f"C{row_idx}"].font = Font(color="0F1F3A", bold=True, size=9)
        worksheet[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")

    _set_number_format(
        worksheet=worksheet,
        column_letters=["R", "S", "T", "U", "V", "X", "Y", "AU", "AV", "AW", "AX", "AZ", "BA", "BB"],
        start_row=TRACK_START_ROW,
        end_row=last_row,
        number_format=CURRENCY_FORMAT,
    )
    _set_number_format(
        worksheet=worksheet,
        column_letters=["B", "K", "L", "M", "N", "O", "Z", "AA", "AB", "AC", "AE", "AF", "AG", "AH", "AJ", "AK", "AL", "AM", "AR", "AS"],
        start_row=TRACK_START_ROW,
        end_row=last_row,
        number_format=INTEGER_FORMAT,
    )
    _set_number_format(
        worksheet=worksheet,
        column_letters=["D", "F"],
        start_row=TRACK_START_ROW,
        end_row=last_row,
        number_format=DECIMAL_FORMAT,
    )
    _set_number_format(
        worksheet=worksheet,
        column_letters=["E", "G", "H", "I", "J", "P", "Q", "W", "AD", "AI", "AN", "AO", "AP", "AQ", "AT", "AY", "BC"],
        start_row=TRACK_START_ROW,
        end_row=last_row,
        number_format=PERCENT_FORMAT,
    )

    widths = {
        "B": 6,
        "C": 24,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 16,
        "K": 14,
        "L": 18,
        "M": 16,
        "R": 16,
        "S": 18,
        "T": 18,
        "U": 20,
        "V": 18,
        "X": 16,
        "Y": 18,
    }

    for col_idx in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = widths.get(column_letter, 14)

    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 48

    for row_idx in range(TRACK_START_ROW, last_row + 1):
        worksheet.row_dimensions[row_idx].height = 22
        

    _apply_track_segment_formatting(
        worksheet=worksheet,
        last_row=last_row,
    )

    _apply_track_conditional_formatting(
        worksheet=worksheet,
        last_row=last_row,
    )

    worksheet.auto_filter.ref = f"B3:BC{last_row}"

def _apply_zebra_rows(
    *,
    worksheet: Worksheet,
    last_row: int,
) -> None:
    zebra_fill = PatternFill("solid", fgColor="F9FAFB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    data_row_counter = 0

    for row_idx in range(TRACK_START_ROW, last_row + 1):
        label = str(worksheet[f"C{row_idx}"].value or "").strip().upper()

        if label.startswith("SUBTOTALES") or label == "TOTAL GENERAL":
            continue

        row_fill = zebra_fill if data_row_counter % 2 else white_fill

        for col_idx in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = row_fill

        data_row_counter += 1


def _apply_summary_row_formatting(
    *,
    worksheet: Worksheet,
    last_row: int,
) -> None:
    subtotal_fill = PatternFill("solid", fgColor="E5E7EB")
    total_fill = PatternFill("solid", fgColor="FBE3DC")
    subtotal_font = Font(color="0F1F3A", bold=True, size=9)
    total_font = Font(color="0F1F3A", bold=True, size=9)

    for row_idx in range(TRACK_START_ROW, last_row + 1):
        label = str(worksheet[f"C{row_idx}"].value or "").strip().upper()

        if label.startswith("SUBTOTALES"):
            row_fill = subtotal_fill
            row_font = subtotal_font
        elif label == "TOTAL GENERAL":
            row_fill = total_fill
            row_font = total_font
        else:
            continue

        for col_idx in range(TRACK_FIRST_COL, TRACK_LAST_COL + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = row_font

def _apply_track_segment_formatting(
    *,
    worksheet: Worksheet,
    last_row: int,
) -> None:
    segment_title_fills = {
        "D": ULTRA_ORANGE,
        "K": ULTRA_BLUE,
        "R": ULTRA_ORANGE,
        "Z": ULTRA_BLUE,
        "AJ": ULTRA_CHARCOAL,
        "AR": ULTRA_BLUE,
        "AU": ULTRA_CHARCOAL,
        "AZ": ULTRA_ORANGE,
    }

    segment_ranges = [
        ("D", "J", "Ocupación"),
        ("K", "Q", "Crecimiento"),
        ("R", "Y", "Ingresos totales"),
        ("Z", "AI", "Ventas nuevas y Reactivaciones"),
        ("AJ", "AQ", "Bajas & Churn"),
        ("AR", "AT", "Domiciliados"),
        ("AU", "AY", "ARPU"),
        ("AZ", "BC", "Tienda"),
    ]

    header_fill = PatternFill("solid", fgColor=HEADER_DARK)
    header_font = Font(color="FFFFFF", bold=True, size=8)
    segment_side = Side(style="medium", color=SEGMENT_BORDER_COLOR)

    for start_col, end_col, _title in segment_ranges:
        start_idx = _column_letter_to_index(start_col)
        end_idx = _column_letter_to_index(end_col)

        for col_idx in range(start_idx, end_idx + 1):
            header_cell = worksheet.cell(row=3, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        title_cell = worksheet[f"{start_col}2"]
        title_cell.fill = PatternFill(
            "solid",
            fgColor=segment_title_fills.get(start_col, ULTRA_CHARCOAL),
        )
        title_cell.font = Font(color="FFFFFF", bold=True, size=9)
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        for row_idx in range(3, last_row + 1):
            left_cell = worksheet[f"{start_col}{row_idx}"]
            right_cell = worksheet[f"{end_col}{row_idx}"]

            left_cell.border = _with_border_side(
                left_cell.border,
                left=segment_side,
            )
            right_cell.border = _with_border_side(
                right_cell.border,
                right=segment_side,
            )

    highlighted_header_columns = {
        "X": ULTRA_ORANGE,
        "Y": ULTRA_ORANGE,
        "AH": "00C853",
        "AM": ULTRA_ORANGE,
        "AQ": ULTRA_ORANGE,
        "AT": ULTRA_BLUE,
        "AX": ULTRA_ORANGE,
        "BB": ULTRA_ORANGE,
        "BC": ULTRA_BLUE,
    }

    for column_letter, color in highlighted_header_columns.items():
        cell = worksheet[f"{column_letter}3"]
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True, size=8)


def _apply_track_conditional_formatting(
    *,
    worksheet: Worksheet,
    last_row: int,
) -> None:
    if last_row < TRACK_START_ROW:
        return

    good_fill = PatternFill("solid", fgColor=GOOD_FILL_COLOR)
    warn_fill = PatternFill("solid", fgColor=WARN_FILL_COLOR)
    bad_fill = PatternFill("solid", fgColor=BAD_FILL_COLOR)

    good_font = Font(color=GOOD_FONT_COLOR, bold=True)
    warn_font = Font(color=WARN_FONT_COLOR, bold=True)
    bad_font = Font(color=BAD_FONT_COLOR, bold=True)

    positive_is_good_columns = [
        "N",   # DIF INICIO MES
        "O",   # Dif meta vs Real
        "P",   # % Alcance meta crecimiento
        "Q",   # Alcance usuarios activos
        "X",   # Meta del día ingresos
        "Y",   # Dif $$ Real vs Meta
        "AC",  # Diferencia clientes nuevos
        "AX",  # Dif $$ ARPU
    ]

    positive_is_bad_columns = [
        "AH",  # Meta del día reactivaciones, positivo = falta
        "AM",  # Bajas sobre ideal, positivo = mal
        "AQ",  # Churn sobre ideal, positivo = mal
        "BB",  # Diferencia tienda, positivo = falta
    ]

    percent_goal_columns = [
        "J",   # % Alcance meta Ocupación
        "W",   # % alcance ingreso
        "AD",  # % alcance nuevos
        "AI",  # % alcance reactivaciones
        "AT",  # % alcance domiciliados
        "AY",  # % alcance ARPU
        "BC",  # % alcance tienda
    ]

    for column_letter in positive_is_good_columns:
        cell_range = f"{column_letter}{TRACK_START_ROW}:{column_letter}{last_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=good_fill,
                font=good_font,
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=bad_fill,
                font=bad_font,
            ),
        )

    for column_letter in positive_is_bad_columns:
        cell_range = f"{column_letter}{TRACK_START_ROW}:{column_letter}{last_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=bad_fill,
                font=bad_font,
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=good_fill,
                font=good_font,
            ),
        )

    for column_letter in percent_goal_columns:
        cell_range = f"{column_letter}{TRACK_START_ROW}:{column_letter}{last_row}"

        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["1"],
                fill=good_fill,
                font=good_font,
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="between",
                formula=["0.8", "0.999999"],
                fill=warn_fill,
                font=warn_font,
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0.8"],
                fill=bad_fill,
                font=bad_font,
            ),
        )


def _with_border_side(
    border: Border,
    *,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(
        left=left or border.left,
        right=right or border.right,
        top=top or border.top,
        bottom=bottom or border.bottom,
    )


def _column_letter_to_index(column_letter: str) -> int:
    result = 0

    for character in column_letter:
        result = result * 26 + ord(character.upper()) - ord("A") + 1

    return result

def _set_number_format(
    *,
    worksheet: Worksheet,
    column_letters: Sequence[str],
    start_row: int,
    end_row: int,
    number_format: str,
) -> None:
    for column_letter in column_letters:
        for row_idx in range(start_row, end_row + 1):
            worksheet[f"{column_letter}{row_idx}"].number_format = number_format


def _build_raw_sheet(*, workbook: Workbook, rows: Sequence[Any]) -> None:
    worksheet = workbook.create_sheet("Daily Mart Raw")
    headers = [
        "track_daily_version_id",
        "track_date",
        "generation_mode",
        "sucursal_canon",
        "target_month",
        "m2_sin_circulaciones",
        "usuarios_inicio_mes",
        "usuarios_activos_actual",
        "proyeccion_usuarios_cierre_mes",
        "meta_faycgo_mes",
        "ingreso_real_base_mtd",
        "ingreso_real_agregadora_mtd",
        "ingreso_real_mtd",
        "meta_clientes_nuevos_mes",
        "clientes_nuevos_real_mtd",
        "meta_reactivaciones_mes",
        "reactivaciones_real_mtd",
        "meta_bajas_mes",
        "bajas_reales_mtd",
        "meta_nuevos_domiciliados_mes",
        "nuevos_domiciliados_real_mtd",
        "meta_arpu_mes",
        "meta_venta_tienda_mes",
        "venta_tienda_real_mtd",
        "source_business_date_desempeno",
        "source_business_date_ingresos",
        "source_business_date_agregadoras",
        "source_business_date_domiciliados",
        "source_business_date_tienda",
    ]

    worksheet.append(headers)

    for row in rows:
        worksheet.append([_serialize_cell_value(getattr(row, header, None)) for header in headers])

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(
            max(len(str(column_cells[0].value or "")) + 2, 12),
            32,
        )


def _build_info_sheet(
    *,
    workbook: Workbook,
    track_date: date,
    generation_mode: str,
    resolved_version: Any,
    total_rows: int,
) -> None:
    worksheet = workbook.create_sheet("Info")
    worksheet.append(["Campo", "Valor"])
    worksheet.append(["Fecha Track", track_date.isoformat()])
    worksheet.append(["Modo", generation_mode])
    worksheet.append(["Versión Track ID", getattr(resolved_version, "id", None)])
    worksheet.append(["Tipo versión", getattr(resolved_version, "version_type", None)])
    worksheet.append(["Status versión", getattr(resolved_version, "status", None)])
    worksheet.append([
        "Generated at UTC",
        _serialize_cell_value(getattr(resolved_version, "generated_at_utc", None)),
    ])
    worksheet.append(["Total sucursales", total_rows])
    worksheet.append(["Fuente", "Suite Ultra / Track Daily Mart"])

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F1F1F")
        cell.font = Font(color="FFFFFF", bold=True)

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 42


def _to_number(value: Any) -> float | int | None:
    if value is None:
        return None

    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        return value

    return None

def _format_version_update_label(resolved_version: Any) -> str:
    version_datetime = (
        getattr(resolved_version, "finished_at_utc", None)
        or getattr(resolved_version, "generated_at_utc", None)
        or getattr(resolved_version, "started_at_utc", None)
    )

    if not isinstance(version_datetime, datetime):
        return "Actualización no disponible"

    local_datetime = _to_tijuana_datetime(version_datetime)

    return f"Actualizado: {local_datetime.strftime('%d/%m/%Y %H:%M')} h"


def _to_tijuana_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)

    return value.astimezone(ZoneInfo("America/Tijuana"))

def _serialize_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if isinstance(value, date):
        return value.isoformat()

    return value


def _format_branch_label(value: Any) -> str:
    return _normalize_branch_key(value).replace("_", " ")

def _sort_rows_by_track_order(rows: Sequence[Any]) -> list[Any]:
    return sorted(
        rows,
        key=lambda item: (
            TRACK_BRANCH_ORDER_INDEX.get(
                _normalize_branch_key(getattr(item, "sucursal_canon", "")),
                999,
            ),
            _normalize_branch_key(getattr(item, "sucursal_canon", "")),
        ),
    )


def _normalize_branch_key(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .upper()
        .replace(" ", "_")
        .replace("-", "_")
    )