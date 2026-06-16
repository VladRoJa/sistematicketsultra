# backend/app/warehouse/jobs/_cobranza_recurrente_rechazados_xlsx.py

from __future__ import annotations

import re
import unicodedata
import zipfile
from collections import defaultdict
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

STATUS_OBJETIVO = "rechazado"
COLUMNA_SUCURSAL = 7
COLUMNA_ESTATUS = 9
COLUMNA_A_ELIMINAR = 12


def letters_a_indice(letters: str) -> int:
    total = 0

    for ch in letters:
        total = total * 26 + (ord(ch) - 64)

    return total


def _leer_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        raw = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ET.fromstring(raw)
    values: list[str] = []

    for si in root.findall(".//m:si", NS):
        texts = si.findall(".//m:t", NS)
        values.append("".join((t.text or "") for t in texts))

    return values


def leer_valor_celda(cell: ET.Element, shared_strings: list[str]) -> str:
    tipo = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)

    if tipo == "inlineStr":
        textos = cell.findall(".//m:t", NS)
        return "".join((t.text or "") for t in textos)

    if value_node is None or value_node.text is None:
        return ""

    raw_value = value_node.text

    if tipo == "s":
        try:
            index = int(raw_value)
            return shared_strings[index] if 0 <= index < len(shared_strings) else raw_value
        except ValueError:
            return raw_value

    return raw_value


def parsear_xlsx_tolerante(ruta_xlsx: Path) -> list[list[str]]:
    """
    Lee el XML interno del XLSX para tolerar archivos con referencias dañadas,
    por ejemplo celdas con r="ANaN".
    """
    with zipfile.ZipFile(ruta_xlsx) as zf:
        shared_strings = _leer_shared_strings(zf)
        worksheet_xml = zf.read("xl/worksheets/sheet1.xml")

    root = ET.fromstring(worksheet_xml)
    rows: list[list[str]] = []

    for row in root.findall(".//m:sheetData/m:row", NS):
        row_map: dict[int, str] = {}
        invalid_values: list[str] = []
        max_valid_idx = 0

        for cell in row.findall("m:c", NS):
            ref = cell.attrib.get("r", "")
            value = leer_valor_celda(cell, shared_strings)
            match = re.fullmatch(r"([A-Z]+)\d+", ref)

            if match:
                col_letters = match.group(1)
                col_idx = letters_a_indice(col_letters)
                row_map[col_idx] = value
                max_valid_idx = max(max_valid_idx, col_idx)
            else:
                invalid_values.append(value)

        if invalid_values:
            next_idx = max(max_valid_idx + 1, 27)
            for value in invalid_values:
                row_map[next_idx] = value
                next_idx += 1

        max_idx = max(row_map.keys(), default=0)
        rows.append([row_map.get(i, "") for i in range(1, max_idx + 1)])

    return rows


def limpiar_nombre_archivo(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r'[\\/:*?"<>|]+', "_", texto)
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto or "SIN_SUCURSAL"


def es_fila_footer(row: list[str]) -> bool:
    valores_no_vacios = [str(v).strip() for v in row if str(v).strip()]

    if len(valores_no_vacios) > 1:
        return False

    if not valores_no_vacios:
        return True

    unico = valores_no_vacios[0]

    return unico.startswith("Reporte generado") or unico.startswith("DEL ")


def ajustar_ancho_columnas(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(cell_value))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def guardar_xlsx(ruta_salida: Path, header: list[str], rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rechazados"

    ws.append(header)

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ajustar_ancho_columnas(ws)
    wb.save(ruta_salida)


def procesar_archivo(ruta_xlsx: Path, carpeta_salida: Path) -> dict:
    rows = parsear_xlsx_tolerante(ruta_xlsx)

    if len(rows) < 3:
        raise ValueError("El archivo no tiene suficientes filas para procesarse.")

    rows = rows[1:-1]
    header = rows[0]
    data_rows = rows[1:]

    data_rows = [r for r in data_rows if not es_fila_footer(r)]

    total_cols = len(header)
    normalizadas = [(r + [""] * total_cols)[:total_cols] for r in data_rows]

    filtradas = [
        r
        for r in normalizadas
        if str(r[COLUMNA_ESTATUS - 1]).strip().lower() == STATUS_OBJETIVO
    ]

    if not filtradas:
        raise ValueError("No se encontraron filas con 'Rechazado' en la columna I.")

    header_final = [
        v
        for i, v in enumerate(header, start=1)
        if i != COLUMNA_A_ELIMINAR
    ]

    filtradas_final = [
        [
            v
            for i, v in enumerate(r, start=1)
            if i != COLUMNA_A_ELIMINAR
        ]
        for r in filtradas
    ]

    agrupadas: dict[str, list[list[str]]] = defaultdict(list)

    for row in filtradas_final:
        sucursal = str(row[COLUMNA_SUCURSAL - 1]).strip() or "SIN_SUCURSAL"
        agrupadas[sucursal].append(row)

    carpeta_salida.mkdir(parents=True, exist_ok=True)

    artifacts: list[dict] = []

    for sucursal, rows_sucursal in sorted(agrupadas.items()):
        nombre_archivo = f"{limpiar_nombre_archivo(sucursal)}.xlsx"
        ruta_salida = carpeta_salida / nombre_archivo

        guardar_xlsx(ruta_salida, header_final, rows_sucursal)

        artifacts.append(
            {
                "sucursal_raw": sucursal,
                "filename": nombre_archivo,
                "path": str(ruta_salida),
                "rows_count": len(rows_sucursal),
            }
        )

    return {
        "total_rows": len(filtradas_final),
        "total_files": len(artifacts),
        "artifacts": artifacts,
    }