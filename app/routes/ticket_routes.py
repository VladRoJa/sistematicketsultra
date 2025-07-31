# app\routes\ticket_routes.py

from flask import Blueprint, jsonify, request, send_file
from flask_cors import CORS
from flask_jwt_extended import get_jwt_identity, jwt_required
from datetime import datetime, timezone, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pytz
from sqlalchemy import or_
from app.models.inventario import InventarioGeneral
from app.utils.ticket_filters import filtrar_tickets_por_usuario
from config import Config
from sqlalchemy import or_
from app.models.ticket_model import Ticket
from app.models.user_model import UserORM
from app.extensions import db
from app.utils.error_handler import manejar_error
from dateutil import parser
from sqlalchemy.orm.attributes import flag_modified
from app.utils.datetime_utils import format_datetime 
from app.models.inventario import InventarioGeneral



# ─────────────────────────────────────────────────────────────
# BLUEPRINT: TICKETS
# ─────────────────────────────────────────────────────────────
ticket_bp = Blueprint('tickets', __name__, url_prefix='/api/tickets')



# ─────────────────────────────────────────────────────────────
# RUTA: Crear ticket
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/create', methods=['POST'])
@jwt_required()
def create_ticket():
    try:
        usuario_actual = get_jwt_identity()
        user = UserORM.get_by_id(usuario_actual)
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        data = request.get_json()

        descripcion = data.get("descripcion")
        departamento_id = data.get("departamento_id")
        criticidad = data.get("criticidad")
        categoria = data.get("categoria")
        subcategoria = data.get("subcategoria")
        detalle = data.get("detalle")
        aparato_id = data.get("aparato_id")
        problema_detectado = data.get("problema_detectado")
        necesita_refaccion = data.get("necesita_refaccion", False)
        descripcion_refaccion = data.get("descripcion_refaccion")
        clasificacion_id = data.get("clasificacion_id")

        if not descripcion or not departamento_id or not criticidad or not categoria or not clasificacion_id:
            return jsonify({"mensaje": "Faltan datos obligatorios"}), 400

        nuevo_ticket = Ticket.create_ticket(
            descripcion=descripcion,
            username=user.username,
            sucursal_id=user.sucursal_id,
            departamento_id=departamento_id,
            criticidad=int(criticidad),
            categoria=categoria,
            subcategoria=subcategoria,
            detalle=detalle,
            aparato_id=aparato_id,
            problema_detectado=problema_detectado,
            necesita_refaccion=necesita_refaccion,
            descripcion_refaccion=descripcion_refaccion,
            url_evidencia=data.get("url_evidencia"),
            ubicacion=data.get("ubicacion"),
            equipo=data.get("equipo"),
            clasificacion_id=clasificacion_id
        )
        
        print("DATA AL CREAR:", data)
        print("Nuevo ticket:", nuevo_ticket.to_dict())

        return jsonify({
            "mensaje": "Ticket creado correctamente",
            "ticket_id": nuevo_ticket.id
        }), 201
        
        


    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener todos los tickets (paginados) - SIMPLE
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/all', methods=['GET'])
@jwt_required()
def get_tickets():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        limit = request.args.get('limit', default=15, type=int)
        offset = request.args.get('offset', default=0, type=int)

        # ¡Usa el helper aquí!
        query = filtrar_tickets_por_usuario(user)

        total_tickets = query.count()
        tickets = query.order_by(Ticket.id.desc()).limit(limit).offset(offset).all()

        return jsonify({
            "mensaje": "Tickets cargados correctamente",
            "tickets": [t.to_dict() for t in tickets],
            "total_tickets": total_tickets
        }), 200

    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Obtener tickets con filtros dinámicos
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/list', methods=['GET'])
@jwt_required()
def list_tickets_with_filters():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        estado = request.args.get('estado')
        departamento_id = request.args.get('departamento_id')
        criticidad = request.args.get('criticidad')
        no_paging = request.args.get('no_paging', default='false').lower() == 'true'
        limit = request.args.get('limit', default=15, type=int)
        offset = request.args.get('offset', default=0, type=int)

        # Helper universal
        query = filtrar_tickets_por_usuario(user)

        if estado:
            query = query.filter_by(estado=estado)
        if departamento_id:
            query = query.filter_by(departamento_id=int(departamento_id))
        if criticidad:
            query = query.filter_by(criticidad=int(criticidad))

        total_tickets = query.count()
        if not no_paging:
            query = query.limit(limit).offset(offset)

        tickets = query.order_by(Ticket.id.desc()).all()

        return jsonify({
            "mensaje": "Tickets filtrados",
            "tickets": [t.to_dict() for t in tickets],
            "total_tickets": total_tickets
        }), 200

    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Actualizar estado de un ticket
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/update/<int:id>', methods=['PUT'])
@jwt_required()
def update_ticket_status(id):
    try:
        ticket = Ticket.query.get(id)
        if not ticket:
            return jsonify({"mensaje": "Ticket no encontrado"}), 404

        data = request.get_json()
        estado = data.get("estado")
        fecha_solucion = data.get("fecha_solucion")
        fecha_en_progreso = data.get("fecha_en_progreso")  
        historial_nuevo = data.get("historial_fechas", [])
        motivo_cambio = data.get("motivo_cambio", "").strip()

        if not estado:
            return jsonify({"mensaje": "Estado es requerido"}), 400

        ticket.estado = estado
        ahora = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

        # ---- Asignación de fechas por estado ----
        if estado == "finalizado":
            ticket.fecha_finalizado = ahora

        if estado == "en progreso":
            # Usar la fecha_en_progreso que viene del frontend, si está
            if fecha_en_progreso:
                try:
                    ticket.fecha_en_progreso = parser.isoparse(fecha_en_progreso).astimezone(timezone.utc)
                except Exception as e:
                    print(f"❌ Error parseando fecha_en_progreso: {e}")
            else:
                ticket.fecha_en_progreso = ahora

        if fecha_solucion:
            try:
                fecha_parsed = parser.isoparse(fecha_solucion)
                ticket.fecha_solucion = fecha_parsed.astimezone(timezone.utc)
            except Exception as e:
                print(f"❌ Error parseando fecha_solucion: {e}")

        # ---- Normalizar historial nuevo ----
        historial_final = ticket.historial_fechas or []

        for entrada in historial_nuevo:
            nueva = entrada.copy()
            for campo in ['fecha', 'fechaCambio']:
                valor = nueva.get(campo)
                if valor:
                    try:
                        nueva[campo] = parser.isoparse(valor).astimezone(timezone.utc).isoformat()
                    except Exception as e:
                        print(f"❌ Error parseando {campo} en ticket #{ticket.id}: {e}")
                        continue

            # Agregar motivo si no viene desde el frontend
            if motivo_cambio and 'motivo' not in nueva:
                nueva['motivo'] = motivo_cambio

            existe_misma_fecha = any(
                parser.isoparse(e.get("fecha")).replace(tzinfo=None) == parser.isoparse(nueva.get("fecha")).replace(tzinfo=None)
                for e in historial_final if e.get("fecha")
            )
            if not existe_misma_fecha:
                historial_final.append(nueva)

        # Ordenar por fecha de cambio, más recientes primero
        historial_final.sort(key=lambda x: parser.isoparse(x['fechaCambio']), reverse=True)
        ticket.historial_fechas = historial_final
        flag_modified(ticket, 'historial_fechas')

        print(f"✅ Historial final para ticket #{ticket.id}: {historial_final}")

        db.session.commit()
        return jsonify({"mensaje": f"Ticket {id} actualizado correctamente"}), 200

    except Exception as e:
        db.session.rollback()
        return manejar_error(e, "update_ticket_status")


# ─────────────────────────────────────────────────────────────
# RUTA: Eliminar ticket
# ─────────────────────────────────────────────────────────────
@ticket_bp.route('/delete/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_ticket(id):
    try:
        ticket = Ticket.query.get(id)
        if not ticket:
            return jsonify({"mensaje": "El ticket no existe"}), 404

        usuario_actual_id = get_jwt_identity()
        usuario_actual = UserORM.get_by_id(usuario_actual_id)

        if ticket.username != usuario_actual.username and usuario_actual.rol != "ADMINISTRADOR":
            return jsonify({"mensaje": "No tienes permiso para eliminar este ticket"}), 403

        db.session.delete(ticket)
        db.session.commit()

        return jsonify({"mensaje": f"Ticket {id} eliminado correctamente"}), 200

    except Exception as e:
        return manejar_error(e)


# ─────────────────────────────────────────────────────────────
# RUTA: Exportar tickets a Excel
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/export-excel', methods=['GET'])
@jwt_required()
def export_excel():
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # --- MULTI-VALORES ---
        estados = request.args.getlist('estado')
        departamentos = request.args.getlist('departamento_id')
        criticidades = request.args.getlist('criticidad')
        usernames = request.args.getlist('username')
        categorias = request.args.getlist('categoria')
        subcategorias = request.args.getlist('subcategoria')
        detalles = request.args.getlist('detalle')
        descripciones = request.args.getlist('descripcion')
        inventarios = request.args.getlist('inventario')

        # --- Fechas ---
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        fecha_fin_desde = request.args.get('fecha_fin_desde')
        fecha_fin_hasta = request.args.get('fecha_fin_hasta')
        fecha_prog_desde = request.args.get('fecha_prog_desde')
        fecha_prog_hasta = request.args.get('fecha_prog_hasta')

        # Helper universal
        query = filtrar_tickets_por_usuario(user)

        # --- MULTI-FILTROS ---
        if estados:
            query = query.filter(Ticket.estado.in_(estados))
        if departamentos:
            query = query.filter(Ticket.departamento_id.in_([int(d) for d in departamentos]))
        if criticidades:
            query = query.filter(Ticket.criticidad.in_([int(c) for c in criticidades]))
        if usernames:
            query = query.filter(Ticket.username.in_(usernames))

        def filtrar_con_null(campo, valores):
            condiciones = []
            for v in valores:
                if v == "—":
                    condiciones.append(campo.is_(None))
                else:
                    condiciones.append(campo == v)
            return or_(*condiciones)

        if categorias:
            query = query.filter(filtrar_con_null(Ticket.categoria, categorias))
        if subcategorias:
            query = query.filter(filtrar_con_null(Ticket.subcategoria, subcategorias))
        if detalles:
            query = query.filter(filtrar_con_null(Ticket.detalle, detalles))
        if descripciones:
            query = query.filter(filtrar_con_null(Ticket.descripcion, descripciones))

        # --- FILTRO POR INVENTARIO ---
        if inventarios:
            inventario_objs = InventarioGeneral.query.filter(InventarioGeneral.nombre.in_(inventarios)).all()
            inventario_ids = [inv.id for inv in inventario_objs]
            if "—" in inventarios:
                query = query.filter(or_(
                    Ticket.aparato_id.in_(inventario_ids) if inventario_ids else False,
                    Ticket.aparato_id.is_(None)
                ))
            else:
                query = query.filter(Ticket.aparato_id.in_(inventario_ids))

        # --- FILTROS DE FECHA ---
        if fecha_desde:
            query = query.filter(Ticket.fecha_creacion >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        if fecha_hasta:
            query = query.filter(Ticket.fecha_creacion <= datetime.strptime(fecha_hasta, '%Y-%m-%d'))
        if fecha_fin_desde:
            query = query.filter(Ticket.fecha_finalizado >= datetime.strptime(fecha_fin_desde, '%Y-%m-%d'))
        if fecha_fin_hasta:
            query = query.filter(Ticket.fecha_finalizado <= datetime.strptime(fecha_fin_hasta, '%Y-%m-%d'))
        if fecha_prog_desde:
            query = query.filter(Ticket.fecha_en_progreso >= datetime.strptime(fecha_prog_desde, '%Y-%m-%d'))
        if fecha_prog_hasta:
            query = query.filter(Ticket.fecha_en_progreso <= datetime.strptime(fecha_prog_hasta, '%Y-%m-%d'))

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()

        # --- EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"

        headers = [
            "ID", "Aparato/Dispositivo", "Descripción", "Usuario", "Estado", "Criticidad",
            "Fecha Creación", "Fecha En Progreso", "Fecha Finalizado", "Fecha Solución",
            "Departamento", "Categoría", "Subcategoría", "Sub-subcategoría",
            "Problema Detectado", "Refacción", "Descripción Refacción"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0073C2")
        alt_fill = PatternFill("solid", fgColor="F2F2F2")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for idx, ticket in enumerate(tickets, start=2):
            t = ticket.to_dict()

            fecha_solucion_corta = ""
            if ticket.fecha_solucion:
                fecha_solucion_corta = ticket.fecha_solucion.astimezone(pytz.timezone("America/Tijuana")).strftime('%d/%m/%Y')

            # 🟢 Lógica combinada para Aparato/Dispositivo
            if ticket.inventario and ticket.inventario.nombre:
                aparato_nombre = ticket.inventario.nombre
            elif ticket.equipo:
                aparato_nombre = ticket.equipo
            else:
                aparato_nombre = "—"

            jerarquia = ticket._obtener_jerarquia_clasificacion() or []

            categoria = jerarquia[0] if len(jerarquia) > 0 else "—"
            subcategoria = jerarquia[1] if len(jerarquia) > 1 else "—"
            subsubcategoria = jerarquia[2] if len(jerarquia) > 2 else "—"
            detalle = jerarquia[3] if len(jerarquia) > 3 else "—"
            
            if hasattr(ticket, "jerarquia_clasificacion"):
                print(f"ID {t.get('id')} jerarquia_clasificacion:", ticket.jerarquia_clasificacion)
            else:
                print(f"ID {t.get('id')} NO TIENE jerarquia_clasificacion")

            print(f"Ticket ID {t.get('id')}, jerarquia_clasificacion: {getattr(ticket, 'jerarquia_clasificacion', None)}")

            ws.append([
                t.get("id"),
                aparato_nombre,   
                t.get("descripcion"),
                t.get("username"),
                t.get("estado"),
                t.get("criticidad"),
                t.get("fecha_creacion"),
                t.get("fecha_en_progreso"),
                t.get("fecha_finalizado"),
                fecha_solucion_corta,
                ticket.departamento.nombre if ticket.departamento else "—",
                categoria,     
                subcategoria,        
                subsubcategoria,    
                detalle,            
                t.get("problema_detectado"),
                "Sí" if t.get("necesita_refaccion") else "No",
                t.get("descripcion_refaccion")
            ])

            

            if idx % 2 == 0:
                for cell in ws[idx]:
                    cell.fill = alt_fill


        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"tickets_exportados_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"mensaje": str(e)}), 500





# ─────────────────────────────────────────────────────────────
# RUTA: migrar tickets a ISO local
# ─────────────────────────────────────────────────────────────


from flask_jwt_extended import jwt_required

@ticket_bp.route('/migrar-historial-local', methods=['POST'])
@jwt_required()
def migrar_historial_local():


    tz_mx = pytz.timezone("America/Tijuana")
    tickets = Ticket.query.all()
    total_actualizados = 0

    for ticket in tickets:
        historial = ticket.historial_fechas
        if not historial:
            continue

        actualizado = False
        nuevo_historial = []

        for entrada in historial:
            nueva_entrada = entrada.copy()
            for campo in ['fecha', 'fechaCambio']:
                valor = entrada.get(campo)
                if valor and isinstance(valor, str) and '/' in valor and len(valor) == 8:
                    try:
                        fecha_local = datetime.strptime(valor, "%d/%m/%y")
                        fecha_local = tz_mx.localize(datetime.combine(fecha_local.date(), time(hour=7)))
                        fecha_utc = fecha_local.astimezone(timezone.utc)
                        nueva_entrada[campo] = fecha_utc.isoformat()
                        actualizado = True
                    except Exception as e:
                        print(f"❌ Error en ticket #{ticket.id}, campo {campo}: {e}")
            nuevo_historial.append(nueva_entrada)

        if actualizado:
            ticket.historial_fechas = nuevo_historial
            total_actualizados += 1

    if total_actualizados > 0:
        db.session.commit()
        return jsonify({"mensaje": f"✅ Historial actualizado en {total_actualizados} tickets."}), 200
    else:
        return jsonify({"mensaje": "⚠️ No se encontraron entradas para actualizar."}), 200


# ─────────────────────────────────────────────────────────────
# RUTA: migrar tickets a ISO railway
# ─────────────────────────────────────────────────────────────

@ticket_bp.route('/migrar-historial-railway', methods=['POST'])
@jwt_required()
def migrar_historial_railway():


    tz_mx = pytz.timezone("America/Tijuana")
    tickets = Ticket.query.all()
    total_actualizados = 0

    for ticket in tickets:
        historial = ticket.historial_fechas
        if not historial:
            continue

        actualizado = False
        nuevo_historial = []

        for entrada in historial:
            nueva_entrada = entrada.copy()
            for campo in ['fecha', 'fechaCambio']:
                valor = entrada.get(campo)
                if valor and isinstance(valor, str) and '/' in valor and len(valor) == 8:
                    try:
                        fecha_local = datetime.strptime(valor, "%d/%m/%y")
                        fecha_local = tz_mx.localize(datetime.combine(fecha_local.date(), time(hour=7)))
                        fecha_utc = fecha_local.astimezone(timezone.utc)
                        nueva_entrada[campo] = fecha_utc.isoformat()
                        actualizado = True
                    except Exception as e:
                        print(f"❌ Error en ticket #{ticket.id}, campo {campo}: {e}")
            nuevo_historial.append(nueva_entrada)

        if actualizado:
            ticket.historial_fechas = nuevo_historial
            total_actualizados += 1

    if total_actualizados > 0:
        db.session.commit()
        return jsonify({"mensaje": f"✅ Historial actualizado en {total_actualizados} tickets (Railway)."}), 200
    else:
        return jsonify({"mensaje": "⚠️ No se encontraron entradas para actualizar (Railway)."}), 200

# ─────────────────────────────────────────────────────────────
# RUTA: eliminar todos tickets
# ─────────────────────────────────────────────────────────────

# Ruta temporal y protegida para borrar todos los tickets


@ticket_bp.route('/eliminar-todos', methods=['DELETE'])
@jwt_required()
def eliminar_todos_los_tickets():
    try:
        cantidad = Ticket.query.delete()
        db.session.commit()
        return jsonify({"mensaje": f"🧨 Se eliminaron {cantidad} tickets."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─────────────────────────────────────────────────────────────
# RUTA: Obtener historial de tickets y reparaciones para un equipo
# ─────────────────────────────────────────────────────────────


@ticket_bp.route('/historial-equipo/<int:equipo_id>', methods=['GET'])
@jwt_required()
def historial_equipo(equipo_id):
    """
    Retorna el historial de tickets relacionados a un equipo/aparato (por aparato_id).
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # Si no eres admin, filtra solo tickets de tu sucursal
        query = Ticket.query.filter(Ticket.aparato_id == equipo_id)
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter(Ticket.sucursal_id == user.sucursal_id)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()

        data = []
        for t in tickets:
            data.append({
                "ticket_id": t.id,
                "descripcion": t.descripcion,
                "estado": t.estado,
                "fecha_creacion": t.fecha_creacion.isoformat() if t.fecha_creacion else None,
                "fecha_solucion": t.fecha_solucion.isoformat() if t.fecha_solucion else None,
                "problema_detectado": t.problema_detectado,
                "necesita_refaccion": t.necesita_refaccion,
                "descripcion_refaccion": t.descripcion_refaccion,
                "historial_fechas": t.historial_fechas or [],
                "username": t.username,
                "asignado_a": t.asignado_a,
                "url_evidencia": t.url_evidencia,
                "categoria": t.categoria,
                "subcategoria": t.subcategoria,
                "detalle": t.detalle,
            })
        return jsonify(data), 200
    except Exception as e:
        return manejar_error(e, "historial_equipo")


@ticket_bp.route('/historial-por-equipo/<int:aparato_id>', methods=['GET'])
@jwt_required()
def historial_por_equipo(aparato_id):
    """
    Devuelve todos los tickets ligados a un aparato/sistema específico (por aparato_id).
    Si el usuario NO es admin, solo ve los tickets de su sucursal.
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        query = Ticket.query.filter_by(aparato_id=aparato_id)

        # Solo admins pueden ver todos, los demás solo los de su sucursal
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter_by(sucursal_id=user.sucursal_id)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
        return jsonify([t.to_dict() for t in tickets]), 200

    except Exception as e:
        return manejar_error(e, "historial_por_equipo")


@ticket_bp.route('/historial-global', methods=['GET'])
@jwt_required()
def historial_global():
    """
    Devuelve el historial completo de tickets de todos los aparatos/sistemas.
    Permite filtrar por tipo (aparato/sistema), sucursal, estado, fecha, etc.
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # Filtros opcionales
        tipo = request.args.get('tipo')  # 'aparato' o 'sistema'
        sucursal_id = request.args.get('sucursal_id', type=int)
        estado = request.args.get('estado')  # abierto/en progreso/finalizado

        query = Ticket.query

        # Filtro por tipo de inventario
        if tipo:
            # Relación con InventarioGeneral (asume que tienes el campo tipo en inventario)
            query = query.join(Ticket.inventario)
            query = query.filter(InventarioGeneral.tipo == tipo)

        # Filtro por sucursal (admins ven todo)
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            query = query.filter_by(sucursal_id=user.sucursal_id)
        elif sucursal_id:
            query = query.filter_by(sucursal_id=sucursal_id)

        # Filtro por estado
        if estado:
            query = query.filter_by(estado=estado)

        tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
        return jsonify([t.to_dict() for t in tickets]), 200

    except Exception as e:
        return manejar_error(e, "historial_global")


@ticket_bp.route('/historial/<int:aparato_id>', methods=['GET'])
@jwt_required()
def historial_aparato(aparato_id):
    """
    Devuelve el historial completo de un aparato/sistema por su ID:
    - Tickets relacionados
    - Movimientos de inventario relacionados
    """
    try:
        user = UserORM.get_by_id(get_jwt_identity())
        if not user:
            return jsonify({"mensaje": "Usuario no encontrado"}), 404

        # FILTRO: Si no es admin, solo puede ver su sucursal
        filtro_sucursal = []
        if not (user.rol == "ADMINISTRADOR" or user.sucursal_id == 1000):
            filtro_sucursal = [user.sucursal_id]

        # --- Tickets relacionados ---
        tickets_query = Ticket.query.filter_by(aparato_id=aparato_id)
        if filtro_sucursal:
            tickets_query = tickets_query.filter(Ticket.sucursal_id.in_(filtro_sucursal))
        tickets = tickets_query.order_by(Ticket.fecha_creacion.desc()).all()
        tickets_serializados = [t.to_dict() for t in tickets]

        # --- Movimientos de inventario relacionados ---
        # Un movimiento puede traer varios detalles (productos), pero filtramos por inventario_id == aparato_id
        from app.models.inventario import DetalleMovimiento, MovimientoInventario

        detalles = DetalleMovimiento.query.filter_by(inventario_id=aparato_id).all()
        movimientos = []
        for d in detalles:
            mov = d.movimiento
            # Filtramos por sucursal si aplica
            if filtro_sucursal and mov.sucursal_id not in filtro_sucursal:
                continue
            movimientos.append({
                'id': mov.id,
                'tipo': mov.tipo_movimiento,
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'usuario_id': mov.usuario_id,
                'sucursal_id': mov.sucursal_id,
                'cantidad': d.cantidad,
                'unidad_medida': d.unidad_medida,
                'observaciones': mov.observaciones,
            })

        # --- Respuesta combinada ---
        return jsonify({
            'tickets': tickets_serializados,
            'movimientos': movimientos,
        }), 200

    except Exception as e:
        return manejar_error(e, "historial_aparato")

